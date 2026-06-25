"""
AbEvaluator — InSynBio AbEngineCore v1.0
=========================================
Universal antibody evaluation entry point for all three workflow types:

  Type 1 — Mouse humanization     → used internally by HumanizationEngine Phase 3+5
  Type 2 — Humanized Ab eval      → compare humanized vs mouse parent
  Type 3 — Fully human Ab eval    → standalone evaluation, no mouse reference

Shared analysis modules (work for ALL types):
  • structure_13param   — 13 structural parameters from PDB / AF2 output
  • cdr_scan            — CDR liability scan (deamidation, isomerization, oxidation, etc.)
  • developability      — SAP, pI, GRAVY, instability index
  • binding_site        — Epitope residues, blocking analysis (requires antigen chain)
  • germline            — Germline origin, SHM count, identity to closest germline

Type-specific additions:
  Type 2 only: delta_vs_mouse (CDR RMSD, angle delta, Vernier packing delta)
  Type 3 only: shm_hotspots, patent_similarity (if enabled)

Usage:
    from core.evaluation import AbEvaluator, AntibodyType

    # Type 3 — Fully human
    ev = AbEvaluator(
        pdb_path="PDL1_Ab2.pdb",
        vh_chain="H", vl_chain="L",
        ab_type=AntibodyType.FULLY_HUMAN,
    )
    result = ev.run(modules=["structure_13param", "cdr_scan", "developability", "germline"])
    result.save("PDL1_Ab2_eval.json")

    # Type 2 — Humanized Ab evaluation
    ev = AbEvaluator(
        pdb_path="humanized.pdb",
        vh_chain="H", vl_chain="L",
        ab_type=AntibodyType.HUMANIZED,
        ref_pdb_path="mouse.pdb",
    )
    result = ev.run(modules=["structure_13param", "cdr_scan", "delta_vs_mouse"])
"""

import json
import sys
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional

_SUITE_ROOT = Path(__file__).resolve().parents[2]


class AntibodyType(str, Enum):
    MOUSE_PARENT         = "mouse_parent"
    HUMANIZED            = "humanized"
    FULLY_HUMAN          = "fully_human"
    VHH                  = "vhh"
    SCFV                 = "scfv"
    IGG_LIKE_BISPECIFIC  = "igg_like_bispecific"
    DOG                  = "dog_caninized"  # Caninized antibody (dog clinical anchors)


# ─────────────────────────────────────────────────────────────────────────────
# Available modules registry
# ─────────────────────────────────────────────────────────────────────────────

ALL_MODULES = {
    "tap": {
        "description": "Therapeutic Antibody Profiler (TAP) metrics: PSH, PPC, PNC, SFvCSP, CDR Length (Raybould 2019).",
        "applies_to": [AntibodyType.MOUSE_PARENT, AntibodyType.HUMANIZED, AntibodyType.FULLY_HUMAN, AntibodyType.VHH],
        "requires_pdb": True,
        "source_script": "core/evaluation/tap.py",
    },
    "structure_13param": {
        "description": "13 structural parameters: ipTM, pTM, pLDDT, interface contacts, VH/VL contacts, pLDDT per region, SASA, etc.",
        "applies_to": [AntibodyType.MOUSE_PARENT, AntibodyType.HUMANIZED, AntibodyType.FULLY_HUMAN, AntibodyType.VHH],
        "requires_pdb": True,
        "source_script": "evaluation_temp/evaluate_pdl1.py → core migration",
    },
    "cdr_scan": {
        "description": "CDR liability scan: deamidation (NG/NS), isomerization (DG/DS), oxidation (M/W), glycosylation (NxS/T), free Cys.",
        "applies_to": [AntibodyType.MOUSE_PARENT, AntibodyType.HUMANIZED, AntibodyType.FULLY_HUMAN,
                       AntibodyType.VHH, AntibodyType.DOG],
        "requires_pdb": False,
        "source_script": "evaluation_temp/scan_cdrs.py → core migration",
    },
    "developability": {
        "description": (
            "Physicochemical profiling: SAP aggregation propensity, pI, GRAVY, instability index, net charge. "
            "Gates evaluated against clinical population benchmarks (humanized_458 / vhh_40 / scfv_84)."
        ),
        "applies_to": [AntibodyType.MOUSE_PARENT, AntibodyType.HUMANIZED, AntibodyType.FULLY_HUMAN,
                       AntibodyType.VHH, AntibodyType.DOG],
        "requires_pdb": False,
        "source_script": "scripts/structure_metrics_humanization.py (partial) → core migration",
    },
    "binding_site": {
        "description": "Epitope residue identification, antigen contact map, blocking analysis (e.g., PD1/PDL1 binding site overlap).",
        "applies_to": [AntibodyType.HUMANIZED, AntibodyType.FULLY_HUMAN],
        "requires_pdb": True,
        "requires_antigen_chain": True,
        "source_script": "evaluation_temp/analyze_pdl1_mechanism.py → core migration",
    },
    "germline": {
        "description": (
            "Closest VH/VL germline vs IMGT aa_translated (species-selectable; primary multi-species source). "
            "OGRDB (human/primate/mouse IG&TR only) is not used here—see data/germlines/ogrdb_human_* only "
            "for ANARCII/confirmed-70 helpers. Optional Fc library stats and fc_probe_seq vs IGHC/IGKC/IGLC AA."
        ),
        "applies_to": [AntibodyType.MOUSE_PARENT, AntibodyType.HUMANIZED, AntibodyType.FULLY_HUMAN,
                       AntibodyType.VHH, AntibodyType.DOG],
        "requires_pdb": False,
        "source_script": "evaluation_temp/analyze_pdl1_germline.py → core migration",
    },
    "dog_scaffold": {
        "description": (
            "Dog caninized antibody scaffold assessment: identity to clinical anchors "
            "(Lokivetmab / Bedinvetmab / Landogrozumab), tier1/tier2 scaffold match, "
            "dog CMC reference from dog_scaffold_cmc_optimization_tier1_tier2_v1.json."
        ),
        "applies_to": [AntibodyType.DOG],
        "requires_pdb": False,
        "reference_population": "dog_clinical_anchors",
        "source_script": "core/evaluation/evaluator.py → _run_dog_scaffold()",
    },
    "delta_vs_mouse": {
        "description": "Type-2 specific: CDR RMSD, VH/VL angle delta, Vernier packing delta vs mouse parent structure.",
        "applies_to": [AntibodyType.HUMANIZED],
        "requires_pdb": True,
        "requires_ref_pdb": True,
        "source_script": "scripts/validate_humanization.py → core migration",
    },
    "cmc_advisor": {
        "description": (
            "Comprehensive developability assessment: 9 key metrics (pI, GRAVY, instability_index, "
            "net_charge_pH7, hydro_patch_max9, charge_patch_max7, agg_motifs, deamidation_sites, "
            "isomerization_sites) benchmarked against clinical reference populations. "
            "Generates targeted optimization suggestions for every metric exceeding established gates."
        ),
        "applies_to": [AntibodyType.MOUSE_PARENT, AntibodyType.HUMANIZED,
                       AntibodyType.FULLY_HUMAN, AntibodyType.DOG],
        "requires_pdb": False,
        "reference_population": "humanized_458",
        "source_script": "core/evaluation/cmc_advisor_module.py",
    },
    "shm_hotspots": {
        "description": "Type-3 specific: somatic hypermutation hotspot identification (WRCY/RGYW motifs) in CDRs.",
        "applies_to": [AntibodyType.FULLY_HUMAN],
        "requires_pdb": False,
        "status": "PLANNED",
    },
    "immunogenicity": {
        "description": (
            "MHC-II T-cell immunogenicity (IEDB 27-allele panel, sliding 15-mer, "
            "cluster analysis, TCIA score) + hydrophilic/hydrophobic surface patch "
            "immunogenicity (Parker scale / SASA-based)."
        ),
        "applies_to": [AntibodyType.MOUSE_PARENT, AntibodyType.HUMANIZED,
                       AntibodyType.FULLY_HUMAN, AntibodyType.VHH],
        "requires_pdb": False,
        "source_script": "core/immunogenicity/mhcii_analyzer.py + surface_immuno.py",
    },
    # ── scFv-specific modules ─────────────────────────────────────────────────
    "scfv_metrics": {
        "description": (
            "scFv-specific metrics: linker length/type/oxidation risk, orientation, "
            "domain-clash risk, aggregation proxy. VH-VL angle + interface when PDB provided. "
            "Results are population-referenced against 84 clinical scFv-like bispecifics."
        ),
        "applies_to": [AntibodyType.SCFV],
        "requires_pdb": False,
        "source_script": "core/evaluation/scfv_metrics.py",
        "reference_population": "scfv_84",
    },
    # ── IgG-like bispecific modules ───────────────────────────────────────────
    "bispecific_arm_cmc": {
        "description": (
            "IgG-like BsAb per-arm CMC (pI, GRAVY, instability, patches, agg motifs), "
            "CDR-aware liability scan (deamidation/isomerization/N-glyc in CDR vs FR), "
            "delta-pI between arms (IEX purification risk). "
            "Results referenced against 232 clinical Fab arms."
        ),
        "applies_to": [AntibodyType.IGG_LIKE_BISPECIFIC],
        "requires_pdb": False,
        "source_script": "core/evaluation/bispecific_metrics.py (BsAbAnalyzer.run_arm_cmc)",
        "reference_population": "humanized_458",
    },
    "bispecific_pairing": {
        "description": (
            "Bispecific format detection (KiH / CrossMab / CommonLC / unknown), "
            "format-specific assembly checks, arm independence score."
        ),
        "applies_to": [AntibodyType.IGG_LIKE_BISPECIFIC],
        "requires_pdb": False,
        "source_script": "core/evaluation/bispecific_metrics.py (BsAbAnalyzer.run_pairing)",
        "reference_population": "humanized_458",
    },
}


# ─────────────────────────────────────────────────────────────────────────────
# Result container
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class EvaluationResult:
    project_name: str
    ab_type: AntibodyType
    modules_run: List[str]
    results: Dict[str, Any] = field(default_factory=dict)
    overall_flags: List[str] = field(default_factory=list)
    generated_at: str = field(default_factory=lambda: datetime.utcnow().isoformat())
    clinical_score: Optional[float] = field(default=None)

    @property
    def overall_status(self) -> str:
        if any("FAIL" in f or "HIGH_RISK" in f for f in self.overall_flags):
            return "FAIL"
        if any("WARN" in f or "MEDIUM_RISK" in f for f in self.overall_flags):
            return "WARN"
        return "PASS"

    def _sync_clinical_score(self) -> None:
        """Sync clinical_score field from developability module result if present."""
        dev = self.results.get("developability") or {}
        if self.clinical_score is None and dev.get("clinical_score") is not None:
            self.clinical_score = dev["clinical_score"]

    # ── Internal helper ──────────────────────────────────────────────────────

    def _executive_summary(self) -> Dict[str, Any]:
        """Build a flat executive-summary block from module results."""
        self._sync_clinical_score()
        dev     = self.results.get("developability", {})
        germ    = self.results.get("germline", {})
        liab    = self.results.get("liabilities", {})
        immuno  = self.results.get("immunogenicity", {})
        struc   = self.results.get("structure", {})
        cmc_adv = self.results.get("cmc_advisor", {})
        adv_gs  = cmc_adv.get("gate_summary", {})
        adv_sug = cmc_adv.get("mutation_suggestions", [])
        return {
            "project":        self.project_name,
            "overall_status": self.overall_status,
            "ab_type":        self.ab_type.value,
            "timestamp":      self.generated_at,
            "modules_run":    self.modules_run,
            "pI_fab":         dev.get("pI_fab_estimate"),
            "GRAVY":          dev.get("GRAVY"),
            "instability_index": dev.get("instability_index"),
            "hydro_patch_max9": dev.get("hydro_patch_max9"),
            "charge_patch_max7": dev.get("charge_patch_max7"),
            "germline_identity_vh": (germ.get("VH") or {}).get("top_match_identity"),
            "germline_identity_vl": (germ.get("VL") or {}).get("top_match_identity"),
            "n_liability_flags":    len(liab.get("flags", [])),
            "immunogenicity_risk":  immuno.get("risk_level", "not_run"),
            "n_immuno_clusters":    immuno.get("n_clusters_high_medium", 0),
            "clinical_score":       self.clinical_score,
            "clinical_population":  dev.get("clinical_population"),
            "cmc_advisor_status":   cmc_adv.get("status"),
            "cmc_n_warn":           adv_gs.get("WARN", 0),
            "cmc_n_fail":           adv_gs.get("FAIL", 0),
            "cmc_n_suggestions":    len(adv_sug),
            "overall_flags":        self.overall_flags,
        }

    # ── Output methods ───────────────────────────────────────────────────────

    def save(self, path: Optional[str] = None) -> Path:
        """Save enriched JSON report (executive_summary + per-module results + _qa)."""
        self._sync_clinical_score()
        target = Path(path) if path else Path(f"{self.project_name}_evaluation.json")
        target.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "abenginecore_version": "1.3.0",
            "executive_summary":    self._executive_summary(),
            "project_name":         self.project_name,
            "ab_type":              self.ab_type.value,
            "overall_status":       self.overall_status,
            "clinical_score":       self.clinical_score,
            "modules_run":          self.modules_run,
            "overall_flags":        self.overall_flags,
            "generated_at":         self.generated_at,
            "results":              self.results,
        }
        if "_qa" in self.results:
            payload["_qa_audit"] = self.results["_qa"]
        with open(target, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)
        print(f"[AbEvaluator] Report saved → {target}")
        return target

    def save_md(self, path: Optional[str] = None) -> Path:
        """Write a human-readable Markdown report."""
        target = Path(path) if path else Path(f"{self.project_name}_evaluation_report.md")
        target.parent.mkdir(parents=True, exist_ok=True)
        with open(target, "w", encoding="utf-8") as f:
            f.write(self._render_md())
        print(f"[AbEvaluator] Markdown report → {target}")
        return target

    def _render_md(self) -> str:
        es      = self._executive_summary()
        dev     = self.results.get("developability", {})
        germ    = self.results.get("germline", {})
        liab    = self.results.get("liabilities", {})
        immuno  = self.results.get("immunogenicity", {})
        cmc_adv = self.results.get("cmc_advisor", {})
        qa      = self.results.get("_qa", {})
        _STATUS_ICON = {"PASS": "✅", "WARN": "⚠️", "FAIL": "❌"}
        icon = _STATUS_ICON.get(self.overall_status, "⚪")

        lines: List[str] = [
            "# AbEvaluator Evaluation Report",
            "",
            f"**Project:** {self.project_name}  ",
            f"**Status:** {icon} **{self.overall_status}**  ",
            f"**Antibody type:** {self.ab_type.value}  ",
            f"**Generated:** {self.generated_at}  ",
            f"**Modules:** {', '.join(self.modules_run)}",
            "",
            "---",
            "",
            "## Executive Summary",
            "",
            "| Metric | Value |",
            "|--------|-------|",
            f"| pI (Fab estimate)    | {es.get('pI_fab') or 'N/A'} |",
            f"| GRAVY                | {es.get('GRAVY') or 'N/A'} |",
            f"| Instability index    | {es.get('instability_index') or 'N/A'} |",
            f"| Hydro patch max9     | {es.get('hydro_patch_max9') or 'N/A'} |",
            f"| Charge patch max7    | {es.get('charge_patch_max7') or 'N/A'} |",
            f"| Germline ID (VH)     | {es.get('germline_identity_vh') or 'N/A'} |",
            f"| Germline ID (VL)     | {es.get('germline_identity_vl') or 'N/A'} |",
            f"| Liability flags      | {es.get('n_liability_flags', 0)} |",
            f"| CMC advisor status   | {es.get('cmc_advisor_status') or 'not_run'} |",
            f"| CMC WARN       | {es.get('cmc_n_warn', 0)} |",
            f"| CMC      | {es.get('cmc_n_suggestions', 0)} |",
            "",
        ]

        if self.overall_flags:
            lines += ["### Flags", ""]
            for flag in self.overall_flags:
                severity = "❌" if "FAIL" in flag or "HIGH" in flag else "⚠️"
                lines.append(f"- {severity} {flag}")
            lines.append("")

        lines += ["---", ""]

        # ── CMC Advisor（ AbRef-458  + ）────────────────────
        if cmc_adv and cmc_adv.get("status") not in ("SKIPPED", "ERROR", None):
            adv_metrics = cmc_adv.get("metrics", {})
            adv_gate_s  = cmc_adv.get("gate_summary", {})
            adv_sug     = cmc_adv.get("mutation_suggestions", [])
            adv_icon    = _STATUS_ICON.get(cmc_adv.get("status", "PASS"), "⚪")

            lines += [
                f"## CMC （ AbRef-458，n=458）",
                "",
                f"**:** {adv_icon} **{cmc_adv.get('status', 'N/A')}**  ",
                f"**:** {cmc_adv.get('reference_population', 'AbRef-458 (n=458)')}  ",
                f"**:** PASS {adv_gate_s.get('PASS', 0)} / "
                f"WARN {adv_gate_s.get('WARN', 0)} / FAIL {adv_gate_s.get('FAIL', 0)}",
                "",
                "|  |  | AbRef-458  |  | p50 | p5 | p95 |",
                "|------|--------|---------------------|------|-----|-----|-----|",
            ]
            for mkey, mdata in adv_metrics.items():
                gate_sym = _STATUS_ICON.get(mdata.get("gate", "PASS"), "⚪")
                val = mdata.get("value")
                val_str = f"{val:.3f}" if isinstance(val, float) else str(val)
                p50 = mdata.get("ref_p50")
                p5  = mdata.get("ref_p5")
                p95 = mdata.get("ref_p95")
                lines.append(
                    f"| {mdata.get('label', mkey)} "
                    f"| {val_str} "
                    f"| {mdata.get('percentile_band', '—')} "
                    f"| {gate_sym} {mdata.get('gate', 'N/A')} "
                    f"| {p50 if p50 is not None else '—'} "
                    f"| {p5 if p5 is not None else '—'} "
                    f"| {p95 if p95 is not None else '—'} |"
                )
            lines.append("")

            if adv_sug:
                lines += ["### （FR-only，CDR/Vernier ）", ""]
                by_cat: Dict[str, List] = {}
                for s in adv_sug:
                    cat = s.get("category", "other")
                    by_cat.setdefault(cat, []).append(s)
                _CAT_LABELS = {
                    "pI_high":       "pI （）",
                    "pI_low":        "pI （）",
                    "hydro_patch":   "",
                    "agg_motif":     " motif",
                    "deamidation":   "（FR）",
                    "isomerization": "（FR）",
                    "instability":   "",
                }
                for cat, cat_sug in by_cat.items():
                    lines += [f"**{_CAT_LABELS.get(cat, cat)}**", ""]
                    for s in cat_sug:
                        sev_icon = "❌" if s.get("severity") == "HIGH" else "⚠️"
                        pos_label = (
                            f"Kabat {s['kabat_pos']}" if s.get("kabat_pos") is not None
                            else f"pos {s['position']}" if "position" in s
                            else f"seq {s['seq_pos']+1}" if "seq_pos" in s
                            else "?"
                        )
                        lines.append(
                            f"- {sev_icon} **{s['chain']} {s.get('region','?')} "
                            f"{s['original']}{pos_label}→{s['suggested']}** "
                            f"[{s.get('severity','?')}]  "
                        )
                        lines.append(f"  {s.get('rationale','')}")
                    lines.append("")
            else:
                lines += ["### ", "", "（ CMC  AbRef-458 ）。", ""]

        elif cmc_adv.get("status") == "SKIPPED":
            lines += ["## CMC ", "", f"> ：{cmc_adv.get('reason', '')}", ""]
        elif cmc_adv.get("status") == "ERROR":
            lines += ["## CMC ", "", f"> ：{cmc_adv.get('error', '')}", ""]

        # ── Developability（， QA audit）─────────────────
        if dev and not cmc_adv:
            lines += [
                "## Developability (CMC)",
                "",
                "| Property | Value | Flags |",
                "|----------|-------|-------|",
                f"| pI       | {dev.get('pI_fab_estimate', 'N/A')} | {'; '.join(f for f in dev.get('flags', []) if 'pI' in f) or '—'} |",
                f"| GRAVY    | {dev.get('GRAVY', 'N/A')} | — |",
                f"| Instability Index | {dev.get('instability_index', 'N/A')} | {'; '.join(f for f in dev.get('flags', []) if 'instab' in f) or '—'} |",
                f"| Net charge pH 7  | {dev.get('net_charge_pH7', 'N/A')} | — |",
                f"| Hydro patch max9 | {dev.get('hydro_patch_max9', 'N/A')} | {'; '.join(f for f in dev.get('flags', []) if 'hydro' in f) or '—'} |",
                f"| Charge patch max7| {dev.get('charge_patch_max7', 'N/A')} | {'; '.join(f for f in dev.get('flags', []) if 'charge' in f) or '—'} |",
                "",
            ]

        # ── Germline ──────────────────────────────────────────────────────────
        if germ:
            lines += ["## Germline Identity", ""]
            for chain, gdata in germ.items():
                if isinstance(gdata, dict):
                    lines.append(
                        f"**{chain}**: top match `{gdata.get('top_match_id', 'N/A')}` "
                        f"— identity {gdata.get('top_match_identity', 'N/A')}"
                    )
            lines.append("")

        # ── Liabilities ───────────────────────────────────────────────────────
        liab_flags = liab.get("flags", [])
        if liab_flags:
            lines += ["## Chemical Liabilities", ""]
            for f in liab_flags:
                lines.append(f"- {f}")
            lines.append("")

        # ── （； TCIA、 ADA ）──
        if immuno and immuno.get("risk_level"):
            summary = immuno.get("summary", {})
            risk    = summary.get("risk_level") or immuno.get("risk_level") or immuno.get("mhcii_risk") or "UNKNOWN"
            n_high  = summary.get("n_high", 0) or immuno.get("n_risk_positions_high", 0)
            n_med   = summary.get("n_medium", 0) or immuno.get("n_risk_positions_medium", 0)
            n_tol   = summary.get("n_tolerated", 0) or immuno.get("n_tolerated", 0)
            n_clust = summary.get("n_clusters", 0) or immuno.get("n_clusters_high_medium", 0)
            action  = summary.get("recommended_action", "")
            risk_icon = _STATUS_ICON.get(risk, "⚪")
            lines += [
                "## ",
                "",
                f"**:** {risk_icon} **{risk}**",
                "",
                "|  |  |",
                "|------|------|",
                f"|  FR  | {n_high} |",
                f"|  FR  | {n_med} |",
                f"|        | {n_tol} |",
                f"|          | {n_clust} |",
                "",
            ]
            if action:
                lines += [f"**：** {action}", ""]

        # ── QA ────────────────────────────────────────────────────────────────
        if qa:
            lines += [
                "## QA Audit",
                "",
                f"| Metric | Value |",
                f"|--------|-------|",
                f"| Status | {qa.get('status', 'N/A')} |",
                f"| PASS   | {qa.get('n_pass', 0)} |",
                f"| WARN   | {qa.get('n_warn', 0)} |",
                f"| FAIL   | {qa.get('n_fail', 0)} |",
                "",
            ]
            qa_checks = qa.get("checks", [])
            if qa_checks:
                lines += ["### QA Findings", ""]
                for chk in qa_checks:
                    lines.append(f"- **{chk.get('level','?')}** `{chk.get('id','?')}`: {chk.get('msg','')}")
                lines.append("")

        lines += ["---", "", "*Report generated by AbEngineCore v1.0 · InSynBio*", ""]
        return "\n".join(lines)

    def print_summary(self):
        print(f"\n{'─'*55}")
        print(f"  AbEvaluator Summary  [{self.ab_type.value}]")
        print(f"  Project : {self.project_name}")
        print(f"  Status  : {self.overall_status}")
        print(f"  Modules : {', '.join(self.modules_run)}")
        if self.overall_flags:
            print(f"  Flags:")
            for flag in self.overall_flags:
                print(f"    • {flag}")
        print(f"{'─'*55}\n")


# ─────────────────────────────────────────────────────────────────────────────
# AbEvaluator
# ─────────────────────────────────────────────────────────────────────────────

class AbEvaluator:
    """
    Universal antibody evaluation engine.

    Shared modules work for Types 1/2/3. Type-specific modules
    are automatically filtered based on ab_type.
    """

    def __init__(
        self,
        project_name: str,
        ab_type: AntibodyType = AntibodyType.FULLY_HUMAN,
        pdb_path: Optional[str] = None,
        vh_chain: str = "H",
        vl_chain: str = "L",
        vh_seq: Optional[str] = None,
        vl_seq: Optional[str] = None,
        ref_pdb_path: Optional[str] = None,
        antigen_chain: Optional[str] = None,
        cdr_seqs: Optional[Dict[str, str]] = None,
        blocking_ref: Optional[Dict[str, int]] = None,
        use_iedb: bool = False,
        immuno_n_clusters: int = 5,
        strict_qa: bool = True,
        # scFv-specific
        scfv_linker_seq: "Optional[str]" = None,
        full_scfv_seq: "Optional[str]" = None,
        scfv_orientation: str = "VH-VL",
        # IgG-like bispecific-specific
        bispecific_arms: "Optional[List[Any]]" = None,
        bispecific_format: "Optional[str]" = None,
        # Germline (IMGT aa_translated): species folder e.g. Homo_sapiens, Mus_musculus, Canis_lupus_familiaris
        germline_species: str = "Homo_sapiens",
        fc_probe_seq: Optional[str] = None,
    ):
        self.project_name  = project_name
        self.ab_type       = AntibodyType(ab_type) if isinstance(ab_type, str) else ab_type
        self.pdb_path      = Path(pdb_path) if pdb_path else None
        self.vh_chain      = vh_chain
        self.vl_chain      = vl_chain
        self.vh_seq        = vh_seq
        self.vl_seq        = vl_seq
        self.ref_pdb_path  = Path(ref_pdb_path) if ref_pdb_path else None
        self.antigen_chain = antigen_chain
        self.cdr_seqs      = cdr_seqs       # {"H1": "GFTFSS...", "H2": ..., "L1": ...}
        self.blocking_ref  = blocking_ref   # {"R113": 94, "M115": 96, ...}
        self.use_iedb_for_immunogenicity = use_iedb   # True = call IEDB API live
        self.immuno_n_clusters = immuno_n_clusters    # k for peptide clustering
        self.strict_qa = strict_qa  # True = raise QAViolation on any QA FAIL
        # scFv fields
        self.scfv_linker_seq  = scfv_linker_seq
        self.full_scfv_seq    = full_scfv_seq
        self.scfv_orientation = scfv_orientation
        # Bispecific fields
        self.bispecific_arms   = bispecific_arms   # List[BsAbArmSpec]
        self.bispecific_format = bispecific_format
        self.germline_species = (germline_species or "Homo_sapiens").replace(" ", "_")
        self.fc_probe_seq = fc_probe_seq.strip() if fc_probe_seq else None

        self._has_biopython = self._probe_import("Bio")

    # ──────────────────────────────────────────────────────────────────────

    def run(
        self,
        modules: Optional[List[str]] = None,
        stored_metrics: Optional[Dict[str, Any]] = None,
    ) -> EvaluationResult:
        """
        Run the requested evaluation modules.

        Args:
            modules: list of module names, or None to run all applicable modules.

        Returns:
            EvaluationResult with per-module results and overall flags.
        """
        applicable = self._applicable_modules()
        requested  = modules if modules is not None else list(applicable.keys())
        to_run     = [m for m in requested if m in applicable]
        skipped    = [m for m in requested if m not in applicable]

        if skipped:
            print(f"[AbEvaluator] Skipped (not applicable for {self.ab_type.value}): {skipped}")

        # A3: Structural linkage — immunogenicity must run before cmc_advisor so that
        # n_hydrophobic_patches (SASA-derived) can override the sequence-level agg_motifs count.
        if "immunogenicity" in to_run and "cmc_advisor" in to_run:
            immuno_idx = to_run.index("immunogenicity")
            cmc_idx    = to_run.index("cmc_advisor")
            if immuno_idx > cmc_idx:
                to_run.insert(cmc_idx, to_run.pop(immuno_idx))
                print("[AbEvaluator] Reordered: immunogenicity moved before cmc_advisor (structural linkage)")

        result = EvaluationResult(
            project_name=self.project_name,
            ab_type=self.ab_type,
            modules_run=to_run,
        )
        stored = stored_metrics or {}

        for module_name in to_run:
            print(f"[AbEvaluator] Running: {module_name}")
            handler = getattr(self, f"_run_{module_name}", None)
            if handler is None:
                result.results[module_name] = {
                    "status": "PLANNED",
                    "note": "Module not yet implemented — scheduled for next release.",
                }
                result.overall_flags.append(f"WARN: {module_name} not yet implemented")
                continue
            try:
                if module_name == "cmc_advisor":
                    # A3: Inject n_hydrophobic_patches from immunogenicity (structural SASA)
                    # to override sequence-level agg_motifs count in the CMC gate.
                    cmc_stored = dict(stored)
                    immuno_res = result.results.get("immunogenicity", {})
                    n_hydro = immuno_res.get("n_hydrophobic_patches")
                    if n_hydro is not None:
                        cmc_stored["_struct_agg_motifs"] = int(n_hydro)
                        print(f"[AbEvaluator] cmc_advisor: injecting _struct_agg_motifs={n_hydro} from immunogenicity SASA")
                    module_result = handler(stored_metrics=cmc_stored)
                else:
                    module_result = handler()
                result.results[module_name] = module_result
                # Collect flags
                for flag in module_result.get("flags", []):
                    result.overall_flags.append(f"{module_name}:{flag}")
            except Exception as e:
                result.results[module_name] = {"status": "ERROR", "error": str(e)}
                result.overall_flags.append(f"FAIL: {module_name} error — {e}")

        result.print_summary()

        # ── Auto QA: validate all module outputs after the run ────────────────
        try:
            from core.qa.pipeline_qa import qa_from_evaluator_result, QAViolation
            qa_report = qa_from_evaluator_result(
                project = self.project_name,
                step    = "evaluation",
                modules = result.results,
                vh_seq  = self.vh_seq,
                vl_seq  = self.vl_seq,
            )
            result.results["_qa"] = {
                "status":       qa_report.status.value,
                "n_pass":       qa_report.n_pass,
                "n_warn":       qa_report.n_warn,
                "n_fail":       qa_report.n_fail,
                "input_hash":   qa_report.input_hash,
                "strict_qa":    self.strict_qa,
                "checks":       [
                    {"id": c.check_id, "level": c.level.value, "msg": c.message}
                    for c in qa_report.checks if c.level.value != "PASS"
                ],
            }
            if qa_report.n_fail > 0:
                flag_msg = (
                    f"QA_FAIL: {qa_report.n_fail} metric(s) outside physical range — "
                    "see results['_qa'] for details"
                )
                result.overall_flags.append(flag_msg)
                if self.strict_qa:
                    fail_msgs = [
                        c.message for c in qa_report.checks if c.level.value == "FAIL"
                    ]
                    raise QAViolation(
                        f"[AbEvaluator] strict_qa=True: {qa_report.n_fail} QA failure(s):\n"
                        + "\n".join(f"  • {m}" for m in fail_msgs)
                        + "\n\nSet strict_qa=False to suppress as warning."
                    )
        except QAViolation:
            raise  # re-raise — do not swallow strict failures
        except Exception as _qa_err:
            result.results["_qa"] = {"status": "QA_ERROR", "error": str(_qa_err)}

        return result

    # ──────────────────────────────────────────────────────────────────────
    # Module implementations
    # ──────────────────────────────────────────────────────────────────────

    def _run_tap(self) -> Dict:
        """Run Therapeutic Antibody Profiler (TAP) analysis."""
        if not self.pdb_path or not self.pdb_path.exists():
            return {"status": "SKIPPED", "reason": "No PDB file provided"}
        
        # Need CDR sequences for TAP
        if not self.cdr_seqs:
             return {"status": "SKIPPED", "reason": "CDR sequences not provided (required for CDR Vicinity)"}

        try:
            from core.evaluation.tap import TAP_Analyzer
            analyzer = TAP_Analyzer(
                pdb_path=str(self.pdb_path),
                vh_chain=self.vh_chain,
                vl_chain=self.vl_chain,
                cdr_seqs=self.cdr_seqs
            )
            results = analyzer.analyze()
            return {
                "status": "PASS",
                "metrics": results,
                "flags": results.get("flags", [])
            }
        except ImportError as e:
             return {"status": "SKIPPED", "reason": f"Dependency missing: {e}"}
        except Exception as e:
            return {"status": "ERROR", "error": str(e), "flags": ["FAIL:TAP_error"]}

    def _run_structure_13param(self) -> Dict:
        """13 structural parameters from PDB."""
        if not self.pdb_path or not self.pdb_path.exists():
            return {"status": "SKIPPED", "reason": "No PDB file provided"}
        if not self._has_biopython:
            return {"status": "SKIPPED", "reason": "BioPython not installed"}
        try:
            sys.path.insert(0, str(_SUITE_ROOT / "scripts"))
            from structure_metrics_humanization import analyze_structure, metrics_to_dict
            metrics = analyze_structure(
                str(self.pdb_path),
                chain_vh=self.vh_chain,
                chain_vl=self.vl_chain,
            )
            md = metrics_to_dict(metrics)
            if md.get("errors"):
                # Treat any structural-metrics error as a hard failure for strict QA runs.
                return {
                    "status": "ERROR",
                    "error": "; ".join(str(e) for e in (md.get("errors") or [])),
                    "metrics": md,
                    "flags": ["FAIL:structure_metrics_error"],
                }
            # Hard check: Vernier 22 dual numbering must be complete when PDB is provided
            vdn = md.get("vernier_dual_numbering") or []
            if isinstance(vdn, list) and len(vdn) != 22:
                return {
                    "status": "ERROR",
                    "error": f"vernier_dual_numbering incomplete: got {len(vdn)}, expected 22",
                    "metrics": md,
                    "flags": ["FAIL:vernier_dual_numbering_incomplete"],
                }
            return {"status": "PASS", "metrics": md, "flags": []}
        except Exception as e:
            return {"status": "ERROR", "error": str(e), "flags": [f"FAIL:structure_error"]}

    def _run_cdr_scan(self) -> Dict:
        """CDR liability scan from sequence."""
        seq = (self.vh_seq or "") + (self.vl_seq or "")
        if not seq:
            return {"status": "SKIPPED", "reason": "No sequence provided"}

        liabilities = []
        flags = []

        # Deamidation: NG, NS
        import re
        for match in re.finditer(r"N[GS]", seq):
            liabilities.append({"type": "deamidation", "pattern": match.group(), "pos": match.start(), "severity": "MEDIUM"})
            flags.append("WARN:deamidation")

        # Isomerization: DG, DS
        for match in re.finditer(r"D[GS]", seq):
            liabilities.append({"type": "isomerization", "pattern": match.group(), "pos": match.start(), "severity": "MEDIUM"})
            flags.append("WARN:isomerization")

        # Glycosylation: N-X-S/T (X != P)
        for match in re.finditer(r"N[^P][ST]", seq):
            liabilities.append({"type": "N-glycosylation", "pattern": match.group(), "pos": match.start(), "severity": "HIGH"})
            flags.append("HIGH_RISK:N_glycosylation")

        # Free Cys — exclude conserved V-region intra-chain disulfide Cys.
        # Standard Kabat positions: VH Kabat 22 (seq ~22) and 92 (seq ~96),
        # VL Kabat 23 (seq ~23) and 88 (seq ~92). These are ALWAYS disulfide-bonded
        # and must NOT be reported as free-Cys risks.
        # We identify them by position in the combined VH+VL sequence.
        vh_seq = self.vh_seq or ""
        vl_seq = self.vl_seq or ""
        vh_len = len(vh_seq)
        # Conserved disulfide positions (0-indexed in combined seq, approximate)
        # VH: residue 22 (pos 21) and residue ~96 (pos 95); VL: residue 23 (pos 22+vh_len) and 88 (pos 91+vh_len)
        _CONSERVED_DISULFIDE_POS = {21, 95, vh_len + 22, vh_len + 91}
        for match in re.finditer(r"C", seq):
            p = match.start()
            if p in _CONSERVED_DISULFIDE_POS:
                continue  # conserved V-region disulfide — not a free-Cys risk
            liabilities.append({"type": "free_Cys_candidate", "pos": p, "severity": "HIGH"})

        return {
            "status": "PASS",
            "total_liabilities": len(liabilities),
            "liabilities": liabilities,
            "flags": list(set(flags)),
        }

    def _run_developability(self) -> Dict:
        """
        CMC/Developability: pI, GRAVY, instability_index, net_charge, SAP proxies.
        All gates are evaluated against the clinical gold standard (1133-antibody database):
          - IgG-like / Humanized / Mouse → ClinicalRuleEngine("humanized_458")
          - VHH                          → ClinicalRuleEngine("vhh_40")
          - scFv                         → ClinicalRuleEngine("scfv_84")
          - DOG caninized                → industry fallback (no dog clinical population yet)
        Industry thresholds (w=0.5) are used only when clinical gates are unavailable.
        """
        from core.evaluation.clinical_rule_engine import get_engine as _get_cre
        import re

        raw_seq = (self.vh_seq or "") + (self.vl_seq or "")
        seq = re.sub(r"[^ACDEFGHIKLMNPQRSTVWY]", "", raw_seq.upper())
        if not seq:
            return {"status": "SKIPPED", "reason": "No valid sequence provided"}

        # Determine clinical population by antibody type
        if self.ab_type == AntibodyType.VHH:
            _cre_population = "vhh_40"
        elif self.ab_type == AntibodyType.SCFV:
            _cre_population = "scfv_84"
        else:
            _cre_population = "humanized_458"

        flags: list = []
        try:
            from Bio.SeqUtils.ProtParam import ProteinAnalysis
            analysis = ProteinAnalysis(seq)
            pi    = analysis.isoelectric_point()
            gravy = analysis.gravy()
            ii    = analysis.instability_index()

            # Net charge at pH 7 (K+R+H*0.1 - D - E)
            try:
                net_charge = analysis.charge_at_pH(7.0)
            except (AttributeError, TypeError):
                net_charge = (
                    seq.count("K") + seq.count("R") + 0.1 * seq.count("H")
                    - seq.count("D") - seq.count("E")
                )

            # SAP proxies: max hydrophobic frac (9-mer), max |charge| (7-mer)
            _HYDROPHOBIC = "AILMFWV"
            _CHARGED_POS = "KR"
            _CHARGED_NEG = "DE"

            def _hydro_frac(s: str) -> float:
                return sum(1 for a in s if a in _HYDROPHOBIC) / len(s) if s else 0.0

            def _abs_charge(s: str) -> float:
                return abs(
                    sum(s.count(x) for x in _CHARGED_POS) + 0.1 * s.count("H")
                    - sum(s.count(x) for x in _CHARGED_NEG)
                )

            hydro_patch_max9 = 0.0
            charge_patch_max7 = 0.0
            if len(seq) >= 9:
                hydro_patch_max9 = max(
                    _hydro_frac(seq[i:i + 9]) for i in range(len(seq) - 8)
                )
            if len(seq) >= 7:
                charge_patch_max7 = max(
                    _abs_charge(seq[i:i + 7]) for i in range(len(seq) - 6)
                )

            # ── Clinical rule engine gates (gold standard, w=1.0) ────────────
            seq_metrics = {
                "pI":               round(pi, 2),
                "GRAVY":            round(gravy, 3),
                "instability_index": round(ii, 1),
                "net_charge_pH7":   round(float(net_charge), 1),
                "hydro_patch_max9": round(hydro_patch_max9, 3),
                "charge_patch_max7": round(charge_patch_max7, 2),
            }
            clinical_score: Optional[float] = None
            percentile_ranks: Dict[str, Any] = {}
            rule_source: str = "clinical"
            cre = _get_cre(_cre_population)
            if cre is not None:
                cre_result = cre.evaluate(seq_metrics)
                clinical_score   = cre_result.clinical_score
                percentile_ranks = cre_result.percentile_ranks
                rule_source      = "clinical"
                for flag in cre_result.flags:
                    if flag:
                        flags.append(flag)
            else:
                rule_source = "industry"
                # Fallback industry gates (only for DOG / unavailable population)
                if pi < 5.5:
                    flags.append(f"WARN:industry:pI_unusually_low ({pi:.2f})")
                if ii > 55:
                    flags.append(f"WARN:industry:instability_index_high ({ii:.1f})")
                if hydro_patch_max9 >= 0.7:
                    flags.append(f"WARN:industry:hydro_patch_max9_high ({hydro_patch_max9:.2f})")
                if charge_patch_max7 >= 7:
                    flags.append(f"WARN:industry:charge_patch_max7_high ({charge_patch_max7:.1f})")

            return {
                "status":            "PASS",
                "pI_fab_estimate":   round(pi, 2),
                "GRAVY":             round(gravy, 3),
                "instability_index": round(ii, 1),
                "net_charge_pH7":    round(float(net_charge), 1),
                "hydro_patch_max9":  round(hydro_patch_max9, 3),
                "charge_patch_max7": round(charge_patch_max7, 2),
                "clinical_population":    _cre_population,
                "clinical_score":         clinical_score,
                "clinical_percentile_ranks": percentile_ranks,
                "rule_source":            rule_source,
                "note": (
                    "pI and instability are sequence-level estimates; confirm post-expression. "
                    f"Gates reference: {_cre_population} clinical population."
                ),
                "flags": flags,
            }
        except ImportError:
            return {"status": "SKIPPED", "reason": "BioPython not installed", "flags": []}
        except Exception as e:
            return {"status": "ERROR", "error": str(e), "flags": []}

    def _run_germline(self) -> Dict:
        """
        Germline analysis using precomputed database (vernier_index_lookup.json).
        For known antibodies: returns Vernier positions, CDR indices, canonical classes.
        For novel sequences: falls back to sequence-level identity vs VH3 germline library.
        Live ANARCI is NOT required.
        """
        if not self.vh_seq:
            return {"status": "SKIPPED", "reason": "No VH sequence provided", "flags": []}

        lookup_path = _SUITE_ROOT / "data" / "humanization_assay" / "vernier_index_lookup.json"
        germline_xl = _SUITE_ROOT / "data" / "humanization_assay" / "thera_human_igG_germline_analysis.xlsx"

        db_match = None
        if germline_xl.exists():
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(germline_xl), read_only=True)
                ws = wb.active
                hdr = [c.value for c in next(ws.iter_rows())]
                name_col = hdr.index("Name") if "Name" in hdr else 0
                vh_col   = hdr.index("VH")   if "VH"   in hdr else 2
                vl_col   = hdr.index("VL")   if "VL"   in hdr else 3
                for row in ws.iter_rows(min_row=2, values_only=True):
                    db_vh = (row[vh_col] or "").strip()
                    if db_vh and self.vh_seq and db_vh == self.vh_seq.strip():
                        db_match = {"name": row[name_col], "db_vl": (row[vl_col] or "").strip()}
                        break
            except Exception:
                pass

        vernier_data = None
        if db_match and lookup_path.exists():
            try:
                import json as _json
                with open(lookup_path, encoding="utf-8") as f:
                    lookup = _json.load(f)
                vernier_data = lookup.get(db_match["name"])
            except Exception:
                pass

        germline_match = self._sequence_germline_identity(self.vh_seq)

        result: Dict[str, Any] = {
            "status":            "PASS",
            "method":            "precomputed_db" if db_match else "sequence_identity",
            "db_antibody_match": db_match["name"] if db_match else None,
            "flags":             [],
            "germline_species":  self.germline_species,
        }
        result.update(germline_match)

        if self.vl_seq and str(self.vl_seq).strip():
            try:
                from ..resources.germline_resources import vl_identity_imgt

                vl_part = vl_identity_imgt(str(self.vl_seq).strip(), self.germline_species)
                result.update(vl_part)
            except Exception as e:
                result["vl_germline_error"] = str(e)

        try:
            from ..resources.germline_resources import fc_probe_identity, summarize_fc_aa_libraries

            result["Fc_libraries"] = summarize_fc_aa_libraries(self.germline_species)
            if self.fc_probe_seq:
                result["Fc_probe"] = fc_probe_identity(self.fc_probe_seq, self.germline_species)
        except Exception as e:
            result["Fc_libraries_error"] = str(e)

        result["VH"] = {
            "top_match":          germline_match.get("closest_vh_germline"),
            "top_match_identity": germline_match.get("vh_germline_identity_pct"),
            "species":            self.germline_species,
        }
        result["VL"] = {
            "top_match":          result.get("closest_vl_germline"),
            "top_match_identity": result.get("vl_germline_identity_pct"),
            "locus":              result.get("vl_germline_locus"),
            "species":            self.germline_species,
        }
        if not (self.vl_seq and str(self.vl_seq).strip()):
            result["VL"]["top_match"] = None
            result["VL"]["top_match_identity"] = None
            result["VL"]["locus"] = None

        if vernier_data:
            result["canonical_classes"]    = vernier_data.get("canonical", {})
            result["vernier_vh_positions"] = vernier_data.get("VH", {})
            result["vernier_vl_positions"] = vernier_data.get("VL", {})
            result["vh_cdr_indices"]       = vernier_data.get("VH_cdr_indices", {})
            result["vl_cdr_indices"]       = vernier_data.get("VL_cdr_indices", {})
        else:
            result["note"] = (
                "Sequence not in precomputed database (842 clinical antibodies). "
                "Install ANARCI for live Vernier zone numbering of novel sequences."
            )
            result["flags"].append(
                "INFO:germline_db_miss — novel sequence; precomputed Vernier data unavailable"
            )
        return result

    def _sequence_germline_identity(self, seq: str) -> Dict:
        """VH sequence-level identity: legacy human VH3 JSON (human only), else IMGT IGHV_aa."""
        germline_path = _SUITE_ROOT / "data" / "germlines" / "human_VH3_germlines.json"
        species = self.germline_species
        if not seq:
            return {"closest_vh_germline": "unknown", "vh_germline_identity_pct": None}
        if species == "Homo_sapiens" and germline_path.exists():
            try:
                import json as _json
                with open(germline_path, encoding="utf-8") as f:
                    germlines = _json.load(f)
                best_name, best_pct = "unknown", 0.0
                items = (germlines.items() if isinstance(germlines, dict)
                         else [(g["name"], g["sequence"]) for g in germlines])
                for gl_name, gl_seq in items:
                    min_len = min(len(seq), len(gl_seq))
                    if min_len < 20:
                        continue
                    pct = sum(a == b for a, b in zip(seq[:min_len], gl_seq[:min_len])) / min_len * 100
                    if pct > best_pct:
                        best_pct, best_name = pct, gl_name
                return {
                    "closest_vh_germline":      best_name,
                    "vh_germline_identity_pct": round(best_pct, 1),
                    "germline_search_db":       "data/germlines/human_VH3_germlines.json",
                }
            except Exception as e:
                return {"closest_vh_germline": "error", "germline_error": str(e)}
        try:
            from ..resources.germline_resources import vh_identity_imgt

            fb = vh_identity_imgt(seq, species)
            if fb:
                return fb
        except Exception as e:
            return {"closest_vh_germline": "error", "germline_error": str(e)}
        return {"closest_vh_germline": "unknown", "vh_germline_identity_pct": None}

    def _run_delta_vs_mouse(self) -> Dict:
        """Type-2: structural delta vs mouse reference."""
        if not self.ref_pdb_path or not self.ref_pdb_path.exists():
            return {"status": "SKIPPED", "reason": "No reference (mouse) PDB provided"}
        if not self.pdb_path or not self.pdb_path.exists():
            return {"status": "SKIPPED", "reason": "No humanized PDB provided"}
        try:
            sys.path.insert(0, str(_SUITE_ROOT / "scripts"))
            import importlib.util
            spec = importlib.util.spec_from_file_location(
                "validate_humanization",
                str(_SUITE_ROOT / "scripts" / "validate_humanization.py")
            )
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            # validate_humanization.py exposes run_comparison()
            if hasattr(mod, "run_comparison"):
                delta = mod.run_comparison(str(self.ref_pdb_path), str(self.pdb_path))
                return {"status": "PASS", "delta": delta, "flags": []}
            return {"status": "PASS", "note": "validate_humanization loaded but run_comparison not found", "flags": []}
        except Exception as e:
            return {"status": "ERROR", "error": str(e), "flags": []}

    def _run_binding_site(self) -> Dict:
        """
        Full binding site + interface metrics analysis.
        Requires: pdb_path + antigen_chain.
        Computes all BioPython-available metrics:
          BSA, H-bonds, salt bridges, hydrophobic contacts, VdW,
          π-π stacking, cation-π, charge complementarity, SC score,
          ΔG estimates, paratope/epitope lists, blocking analysis.
        """
        if not self.antigen_chain:
            return {"status": "SKIPPED",
                    "reason": "No antigen chain specified. Pass antigen_chain= to AbEvaluator.",
                    "flags": []}
        if not self.pdb_path or not self.pdb_path.exists():
            return {"status": "SKIPPED",
                    "reason": f"PDB not found: {self.pdb_path}",
                    "flags": []}
        try:
            from core.evaluation.interface_metrics import compute_interface_metrics
        except ImportError:
            try:
                sys.path.insert(0, str(_SUITE_ROOT))
                from core.evaluation.interface_metrics import compute_interface_metrics
            except ImportError as e:
                return {"status": "ERROR", "error": str(e), "flags": []}

        result = compute_interface_metrics(
            pdb_path      = str(self.pdb_path),
            vh_chain      = self.vh_chain,
            vl_chain      = self.vl_chain,
            ag_chain      = self.antigen_chain,
            cutoff_contact= 5.0,
            cdr_seqs      = getattr(self, "cdr_seqs", None),
            blocking_ref  = getattr(self, "blocking_ref", None),
        )
        result["status"] = result.get("status", "PASS")
        return result

    def _run_immunogenicity(self) -> Dict:
        """
        MHC-II T-cell immunogenicity + surface patch immunogenicity.

        Requires: vh_seq (and optionally vl_seq).
        Optionally uses pdb_path for SASA-based surface analysis.
        Set use_iedb=True on AbEvaluator (or pass immunogenicity_kwargs) to
        enable live IEDB prediction; default is offline heuristic.
        """
        vh = self.vh_seq or ""
        vl = self.vl_seq or ""
        if not vh:
            return {"status": "SKIPPED", "reason": "No VH sequence provided.", "flags": []}

        use_iedb  = getattr(self, "use_iedb_for_immunogenicity", False)
        n_clusters = getattr(self, "immuno_n_clusters", 5)

        try:
            from core.immunogenicity.mhcii_analyzer import MHCII_Analyzer
            from core.immunogenicity.surface_immuno import SurfaceImmunogenicity
        except ImportError as e:
            return {"status": "ERROR", "error": str(e), "flags": []}

        # ── MHC-II analysis ──────────────────────────────────────────────────
        is_vhh = getattr(self, "ab_type", None) == AntibodyType.VHH

        analyzer = MHCII_Analyzer(
            vh_seq=vh, vl_seq=vl, use_iedb=use_iedb, n_clusters=n_clusters,
        )
        mhcii = analyzer.run(is_vhh=is_vhh)

        # ── Surface immunogenicity ───────────────────────────────────────────
        si = SurfaceImmunogenicity(
            pdb_path = str(self.pdb_path) if self.pdb_path and self.pdb_path.exists() else None,
            vh_chain = self.vh_chain, vl_chain = self.vl_chain,
            vh_seq   = vh, vl_seq = vl,
        )
        surf = si.analyze()

        # ── Combined output ──────────────────────────────────────────────────
        from core.immunogenicity.report_writer import format_immunogenicity_json
        summary = format_immunogenicity_json(mhcii)

        return {
            "status":          "PASS",
            "method":          mhcii.method,
            # Pre-computed results-only summary (used by report_writer)
            "summary":         summary,
            # MHC-II (internal audit)
            "tcia_score":      mhcii.tcia_score,
            "mhcii_risk":      mhcii.risk_level,
            "n_epitopes":      len(mhcii.all_epitopes),
            "n_clusters":      mhcii.n_clusters,
            "funnel_stats":    getattr(mhcii, "funnel_stats", {}),
            "top_epitopes":    [
                {"peptide": e.peptide, "chain": e.chain, "region": e.region,
                 "n_alleles": e.n_alleles_hit, "n_strong": e.n_strong,
                 "best_rank": e.best_rank, "best_allele": e.best_allele,
                 "cluster": e.cluster_id, "risk": e.risk}
                for e in mhcii.top_epitopes
            ],
            "cluster_summary": mhcii.cluster_summary,
            "allele_coverage": mhcii.allele_coverage,
            # Surface
            "surface_risk":    surf.overall_surface_risk,
            "surface_method":  surf.method,
            "n_hydrophilic_patches": len(surf.exposed_hydrophilic_patches),
            "n_hydrophobic_patches": len(surf.exposed_hydrophobic_patches),
            "hydrophilic_patches": [
                {"chain": p.chain, "start": p.start, "seq": p.sequence,
                 "region": p.region, "n_charged": p.n_charged, "risk": p.risk}
                for p in surf.exposed_hydrophilic_patches[:10]
            ],
            "hydrophobic_patches": [
                {"chain": p.chain, "start": p.start, "seq": p.sequence,
                 "region": p.region, "risk": p.risk}
                for p in surf.exposed_hydrophobic_patches[:10]
            ],
            "surface_charge_map":       surf.surface_charge_map,
            "surface_hydrophilicity":   surf.hydrophilicity_score,
            "alleles_used":    mhcii.alleles_used,
            "flags": mhcii.flags + surf.flags,
        }

    def _run_cmc_advisor(self, stored_metrics: Optional[Dict[str, Any]] = None) -> Dict:
        """
        Full CMC/developability assessment benchmarked against AbRef-458 (n=458).
        Generates FR-only mutation suggestions for every metric exceeding its gate.
        No ADA / immunogenicity content is included.
        Database-first: reuses stored_metrics when provided.
        """
        if not self.vh_seq:
            return {"status": "SKIPPED", "reason": "No VH sequence provided", "flags": []}
        try:
            from core.evaluation.cmc_advisor_module import run_cmc_advisor
            raw = self._extract_stored_cmc_metrics(stored_metrics or {})
            return run_cmc_advisor(
                vh_seq=self.vh_seq,
                vl_seq=self.vl_seq or "",
                stored_metrics=raw,
                structure_path=str(self.pdb_path) if self.pdb_path and self.pdb_path.exists() else None,
            )
        except Exception as e:
            return {"status": "ERROR", "error": str(e), "flags": [f"FAIL:cmc_advisor_error"]}

    def _extract_stored_cmc_metrics(self, stored: Dict[str, Any]) -> Dict[str, Any]:
        """Extract raw metric values from stored cmc_advisor metrics for database-first reuse."""
        metrics = stored.get("metrics") or stored
        if not isinstance(metrics, dict):
            return {}
        out: Dict[str, Any] = {}
        for k, v in metrics.items():
            if isinstance(v, dict) and "value" in v:
                out[k] = v["value"]
            elif isinstance(v, (int, float, list)):
                out[k] = v
        return out

    def _run_shm_hotspots(self) -> Dict:
        """Type-3: SHM hotspot motifs (WRCY/RGYW) in CDRs."""
        seq = (self.vh_seq or "") + (self.vl_seq or "")
        if not seq:
            return {"status": "SKIPPED", "reason": "No sequence provided"}
        import re
        hotspots = []
        for pattern, name in [
            (r"[AT][AG]C[CT]", "WRCY"),
            (r"[AG]G[CT][AT]", "RGYW"),
        ]:
            for m in re.finditer(pattern, seq):
                hotspots.append({"motif": name, "sequence": m.group(), "pos": m.start()})
        return {
            "status": "PASS",
            "shm_hotspot_count": len(hotspots),
            "hotspots": hotspots,
            "note": "High hotspot density in CDRs may indicate ongoing SHM optimization in vivo.",
            "flags": ["WARN:high_shm_density"] if len(hotspots) > 6 else [],
        }

    # ──────────────────────────────────────────────────────────────────────
    # Utility
    # ──────────────────────────────────────────────────────────────────────


    def _run_scfv_metrics(self) -> Dict:
        """scFv-specific metrics: linker, orientation, aggregation, structure."""
        try:
            from core.evaluation.scfv_metrics import compute_scfv_metrics
        except ImportError as e:
            return {"status": "SKIPPED", "reason": str(e), "flags": []}
        result = compute_scfv_metrics(
            vh_seq        = self.vh_seq,
            vl_seq        = self.vl_seq,
            linker_seq    = self.scfv_linker_seq,
            full_scfv_seq = self.full_scfv_seq,
            orientation   = self.scfv_orientation,
            pdb_path      = str(self.pdb_path) if self.pdb_path else None,
            vh_chain      = self.vh_chain,
            vl_chain      = self.vl_chain,
        )
        # Annotate with clinical reference percentiles
        try:
            from core.evaluation.reference_calibrator import get_calibrator
            cal = get_calibrator()
            if cal:
                ref_keys = {"linker_length", "contact_vh_vl", "linker_end_to_end_A", "junction_vh_vl_A"}
                subset = {k: v for k, v in result.items() if k in ref_keys and isinstance(v, (int, float))}
                ref_ctx = cal.annotate(subset, population="scfv_84")
                result["ref_context"] = {k: v for k, v in ref_ctx.items() if k not in subset}
                result["ref_flags"]   = cal.summary_flags(subset, population="scfv_84")
        except Exception:
            pass
        return result

    def _run_bispecific_arm_cmc(self) -> Dict:
        """Per-arm CMC + CDR liability + delta-pI for IgG-like bispecific."""
        if not self.bispecific_arms:
            return {"status": "SKIPPED", "reason": "No bispecific_arms provided. Pass bispecific_arms=[BsAbArmSpec(...), ...] to AbEvaluator.", "flags": []}
        try:
            from core.evaluation.bispecific_metrics import BsAbAnalyzer
        except ImportError as e:
            return {"status": "SKIPPED", "reason": str(e), "flags": []}
        result = BsAbAnalyzer(self.bispecific_arms, format_hint=self.bispecific_format).run_arm_cmc()
        # Annotate arm-level pI with clinical percentiles
        try:
            from core.evaluation.reference_calibrator import get_calibrator
            cal = get_calibrator()
            if cal:
                for arm_r in result.get("arms", []):
                    cmc = arm_r.get("cmc") or {}
                    subset = {k: v for k, v in cmc.items() if k in {"pI", "GRAVY", "instability_index", "net_charge_pH7"} and isinstance(v, (int, float))}
                    arm_r["ref_context"] = cal.annotate(subset, population="igg_like_232fab")
                    arm_r["ref_flags"]   = cal.summary_flags(subset, population="igg_like_232fab")
        except Exception:
            pass
        return result

    def _run_bispecific_pairing(self) -> Dict:
        """Bispecific format detection and assembly risk checks."""
        if not self.bispecific_arms:
            return {"status": "SKIPPED", "reason": "No bispecific_arms provided.", "flags": []}
        try:
            from core.evaluation.bispecific_metrics import BsAbAnalyzer
        except ImportError as e:
            return {"status": "SKIPPED", "reason": str(e), "flags": []}
        return BsAbAnalyzer(self.bispecific_arms, format_hint=self.bispecific_format).run_pairing()

    def _run_dog_scaffold(self) -> Dict:
        """
        Dog caninized antibody scaffold assessment.

        Scores VH/VL sequences against:
          1. Clinical anchors: Lokivetmab (IL-31RA), Bedinvetmab (NGF), Landogrozumab (myostatin)
             from dog_scaffold_cmc_optimization_tier1_tier2_v1.json
          2. Tier1/Tier2 scaffold shortlist from dog_scaffold_shortlist_tier1_tier2_v1.json
          3. Dog CMC reference (pI, stability class, formulation notes per scaffold)
        """
        if not self.vh_seq:
            return {"status": "SKIPPED", "reason": "No VH sequence provided", "flags": []}

        _DOG_DIR = _SUITE_ROOT / "data" / "germlines" / "canis_lupus_familiaris_ig_aa"
        _SHORTLIST_PATH = _DOG_DIR / "dog_scaffold_shortlist_tier1_tier2_v1.json"
        _CMC_PATH       = _DOG_DIR / "dog_scaffold_cmc_optimization_tier1_tier2_v1.json"
        _DECISION_PATH  = _DOG_DIR / "dog_decision_support_v1.json"

        try:
            import json as _json
            shortlist = _json.loads(_SHORTLIST_PATH.read_text()) if _SHORTLIST_PATH.exists() else {}
            cmc_data  = _json.loads(_CMC_PATH.read_text())       if _CMC_PATH.exists()       else {}
            decision  = _json.loads(_DECISION_PATH.read_text())  if _DECISION_PATH.exists()  else {}
        except Exception as e:
            return {"status": "ERROR", "error": str(e), "flags": []}

        def _identity(a: str, b: str) -> float:
            if not a or not b:
                return 0.0
            n = min(len(a), len(b))
            return sum(1 for i in range(n) if a[i] == b[i]) / n

        # ── Score against clinical anchors ────────────────────────────────────
        anchors = []
        clinical_anchors_data = (
            cmc_data.get("clinical_anchors")
            or cmc_data.get("anchors")
            or decision.get("clinical_anchors", [])
        )
        for anchor in (clinical_anchors_data or []):
            anchor_vh  = anchor.get("vh_sequence") or anchor.get("VH_sequence", "")
            anchor_vl  = anchor.get("vl_sequence") or anchor.get("VL_sequence", "")
            name       = anchor.get("name") or anchor.get("antibody_name", "unknown")
            target     = anchor.get("target", "")
            id_vh      = round(_identity(self.vh_seq, anchor_vh), 3) if anchor_vh else None
            id_vl      = round(_identity(self.vl_seq or "", anchor_vl), 3) if anchor_vl else None
            anchors.append({
                "name":       name,
                "target":     target,
                "vh_identity": id_vh,
                "vl_identity": id_vl,
            })

        # ── Scaffold tier match ───────────────────────────────────────────────
        tier1_scaffolds = shortlist.get("tier1", [])
        tier2_scaffolds = shortlist.get("tier2", [])
        scaffold_tier = "NOVEL"
        scaffold_best = None
        best_id = 0.0
        for sc in tier1_scaffolds + tier2_scaffolds:
            sc_seq = sc.get("vh_sequence") or sc.get("VH_sequence", "")
            if not sc_seq:
                continue
            sc_id = _identity(self.vh_seq, sc_seq)
            if sc_id > best_id:
                best_id = sc_id
                scaffold_best = sc.get("name") or sc.get("id", "")
                scaffold_tier = "tier1" if sc in tier1_scaffolds else "tier2"

        # ── CMC reference ─────────────────────────────────────────────────────
        cmc_note = cmc_data.get("note") or cmc_data.get("summary", "")

        flags: list = []
        if best_id < 0.70:
            flags.append(
                f"WARN:dog_scaffold:low_identity_to_known_scaffold ({best_id:.1%}) — "
                "novel scaffold may require custom CMC development"
            )
        if not anchors:
            flags.append("INFO:dog_scaffold:no_clinical_anchors_loaded")

        return {
            "status":           "PASS" if not any("WARN" in f for f in flags) else "WARN",
            "clinical_anchors": anchors,
            "best_scaffold_match": {
                "name":      scaffold_best,
                "tier":      scaffold_tier,
                "vh_identity": round(best_id, 3),
            },
            "cmc_note":         cmc_note,
            "flags":            flags,
        }

    def _applicable_modules(self) -> Dict:
        """Filter ALL_MODULES to those applicable for this ab_type."""
        return {
            name: info for name, info in ALL_MODULES.items()
            if self.ab_type in info["applies_to"]
        }

    @staticmethod
    def _probe_import(module_name: str) -> bool:
        import importlib.util
        return importlib.util.find_spec(module_name) is not None

    def __repr__(self):
        return (f"AbEvaluator(project={self.project_name!r}, "
                f"type={self.ab_type.value}, pdb={self.pdb_path})")


# ─────────────────────────────────────────────────────────────────────────────
# EvaluationComparison — 4-phase evaluation workflow helper
# ─────────────────────────────────────────────────────────────────────────────

class EvaluationComparison:
    """
    Encapsulates the standard 4-phase antibody developability workflow:

      Phase 1: Evaluate original sequence
      Phase 2: CMC optimization (caller's responsibility — supply optimized sequence)
      Phase 3: Evaluate optimized sequence
      Phase 4: Compare against clinical reference

    Usage
    -----
    ::

        ev_orig = AbEvaluator(project="mumab4d5", vh_seq=vh_orig, vl_seq=vl_orig, ...)
        ev_opt  = AbEvaluator(project="mumab4d5", vh_seq=vh_opt,  vl_seq=vl_opt,  ...)
        ev_ref  = AbEvaluator(project="trastuzumab", vh_seq=vh_ref, vl_seq=vl_ref, ...)

        comp = EvaluationComparison(
            original=ev_orig.run(),
            optimized=ev_opt.run(),
            reference=ev_ref.run(),
            original_label="muMAb4D5 V2 (original)",
            optimized_label="muMAb4D5 V2 (optimized)",
            reference_label="Trastuzumab (clinical)",
        )
        table = comp.diff()
        report = comp.summary_text()
    """

    # Metrics extracted from cmc_advisor output for the diff table
    _DIFF_METRICS = [
        "pI", "GRAVY", "instability_index", "net_charge_pH7",
        "hydro_patch_max9", "charge_patch_max7", "SAP_score", "Fv_charge_asymmetry",
        "agg_motifs", "hydro_cluster_count",
        "glycosylation_sites", "deamidation_sites", "isomerization_sites",
        "oxidation_sites", "free_cys",
    ]

    def __init__(
        self,
        original: "EvaluationResult",
        optimized: "EvaluationResult",
        reference: "EvaluationResult",
        *,
        original_label: str = "Original",
        optimized_label: str = "Optimized",
        reference_label: str = "Reference (clinical)",
    ):
        self.original   = original
        self.optimized  = optimized
        self.reference  = reference
        self.original_label  = original_label
        self.optimized_label = optimized_label
        self.reference_label = reference_label

    # ── helpers ──────────────────────────────────────────────────────────────

    @staticmethod
    def _extract_cmc(result: "EvaluationResult") -> Dict[str, Any]:
        """Extract annotated metrics dict from an EvaluationResult."""
        cmc = result.results.get("cmc_advisor", {})
        return cmc.get("annotated", {})

    @staticmethod
    def _extract_adi(result: "EvaluationResult") -> Optional[float]:
        cmc = result.results.get("cmc_advisor", {})
        return cmc.get("adi")

    @staticmethod
    def _extract_value(annotated: Dict[str, Any], metric: str) -> Optional[float]:
        entry = annotated.get(metric, {})
        val = entry.get("value")
        if isinstance(val, list):
            return float(len(val))
        if val is None:
            return None
        try:
            return float(val)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _extract_gate(annotated: Dict[str, Any], metric: str) -> str:
        return annotated.get(metric, {}).get("gate", "—")

    # ── public API ────────────────────────────────────────────────────────────

    def diff(self) -> List[Dict[str, Any]]:
        """
        Return a list of per-metric comparison rows.

        Each row is a dict with keys:
          metric, label, direction,
          original_value, original_gate,
          optimized_value, optimized_gate, delta,
          reference_value, reference_gate,
          improved (bool | None)
        """
        ann_orig = self._extract_cmc(self.original)
        ann_opt  = self._extract_cmc(self.optimized)
        ann_ref  = self._extract_cmc(self.reference)

        rows: List[Dict[str, Any]] = []
        for metric in self._DIFF_METRICS:
            v_orig = self._extract_value(ann_orig, metric)
            v_opt  = self._extract_value(ann_opt,  metric)
            v_ref  = self._extract_value(ann_ref,  metric)

            direction = ann_orig.get(metric, {}).get("direction",
                        ann_opt.get(metric,  {}).get("direction", "in_range"))
            label     = (ann_orig.get(metric) or ann_opt.get(metric) or {}).get("label", metric)

            # Delta: signed (optimized − original), None if either missing
            delta: Optional[float] = None
            if v_orig is not None and v_opt is not None:
                delta = round(v_opt - v_orig, 4)

            # Did optimized improve over original (direction-aware)?
            improved: Optional[bool] = None
            if delta is not None:
                if direction == "lower_is_better":
                    improved = delta < 0
                else:
                    improved = None  # symmetric: "better" requires domain context

            rows.append({
                "metric":          metric,
                "label":           label,
                "direction":       direction,
                "original_value":  v_orig,
                "original_gate":   self._extract_gate(ann_orig, metric),
                "optimized_value": v_opt,
                "optimized_gate":  self._extract_gate(ann_opt,  metric),
                "delta":           delta,
                "reference_value": v_ref,
                "reference_gate":  self._extract_gate(ann_ref,  metric),
                "improved":        improved,
            })
        return rows

    def adi_summary(self) -> Dict[str, Any]:
        """Return ADI scores for all three phases."""
        from core.cmc.adi_score import adi_interpretation
        adi_orig = self._extract_adi(self.original)
        adi_opt  = self._extract_adi(self.optimized)
        adi_ref  = self._extract_adi(self.reference)
        return {
            "original":  {"adi": adi_orig, "interpretation": adi_interpretation(adi_orig) if adi_orig else None},
            "optimized": {"adi": adi_opt,  "interpretation": adi_interpretation(adi_opt)  if adi_opt  else None},
            "reference": {"adi": adi_ref,  "interpretation": adi_interpretation(adi_ref)  if adi_ref  else None},
            "delta_orig_to_opt": round(adi_opt - adi_orig, 2) if (adi_opt and adi_orig) else None,
            "delta_orig_to_ref": round(adi_ref - adi_orig, 2) if (adi_ref and adi_orig) else None,
        }

    def summary_text(self) -> str:
        """
        Return a concise Markdown summary table comparing all three evaluations.
        """
        rows = self.diff()
        adi  = self.adi_summary()

        lines: List[str] = [
            "## AbEvaluator Comparison Summary",
            "",
            f"| | {self.original_label} | {self.optimized_label} | {self.reference_label} |",
            "|---|---|---|---|",
            f"| **ADI** | {adi['original']['adi']} ({adi['original']['interpretation']}) "
            f"| {adi['optimized']['adi']} ({adi['optimized']['interpretation']}) "
            f"| {adi['reference']['adi']} ({adi['reference']['interpretation']}) |",
            "",
            "### Per-metric comparison",
            "",
            f"| Metric | Direction | {self.original_label} | Gate | "
            f"{self.optimized_label} | Gate | Δ | {self.reference_label} | Gate |",
            "|---|---|---|---|---|---|---|---|---|",
        ]
        for r in rows:
            dir_sym = "↓" if r["direction"] == "lower_is_better" else "↕"
            v_orig  = "—" if r["original_value"]  is None else r["original_value"]
            v_opt   = "—" if r["optimized_value"] is None else r["optimized_value"]
            v_ref   = "—" if r["reference_value"] is None else r["reference_value"]
            delta   = "—" if r["delta"] is None else (f"+{r['delta']}" if r["delta"] > 0 else str(r["delta"]))
            lines.append(
                f"| {r['label']} | {dir_sym} | {v_orig} | {r['original_gate']} "
                f"| {v_opt} | {r['optimized_gate']} | {delta} "
                f"| {v_ref} | {r['reference_gate']} |"
            )
        lines.append("")
        lines.append(
            f"> ADI delta (original → optimized): **{adi['delta_orig_to_opt']}**  "
            f"  ADI delta (original → reference): **{adi['delta_orig_to_ref']}**"
        )
        return "\n".join(lines)

    def __repr__(self) -> str:
        return (f"EvaluationComparison("
                f"orig={self.original_label!r}, "
                f"opt={self.optimized_label!r}, "
                f"ref={self.reference_label!r})")
