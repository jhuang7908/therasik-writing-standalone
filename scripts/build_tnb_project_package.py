#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
build_tnb_project_package.py
============================
Assembles the complete Tnb04/Tnb164 bispecific project package:
  1. Neutralization data → Excel (multi-sheet)
  2. CMC data → Excel (multi-sheet)
  3. Copies/links all artefacts into projects/Tnb_bispecific/
  4. Generates final balance report
"""
from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side,
                              numbers)
from openpyxl.utils import get_column_letter

SUITE_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR    = SUITE_ROOT / "projects" / "Tnb_bispecific"
CMC_JSON   = OUT_DIR / "cmc_eval" / "tnb_full_cmc_real.json"

data = json.loads(CMC_JSON.read_text(encoding="utf-8"))
singles  = data["single_vhh"]
fusions  = data["fusion_proteins"]
activity = data["activity"]

# ── Palette ──────────────────────────────────────────────────────────────────
C_HEADER   = "1F3864"   # dark navy
C_SUBHDR   = "2E75B6"   # mid blue
C_PASS     = "C6EFCE"   # green fill
C_WARN     = "FFEB9C"   # yellow fill
C_FAIL     = "FFC7CE"   # red fill
C_BEST     = "E2EFDA"   # light green highlight
C_ALT      = "EBF3FB"   # light blue alternating row
C_WHITE    = "FFFFFF"

def hfill(c):   return PatternFill("solid", fgColor=c)
def hfont(bold=False, color="000000", sz=10):
    return Font(bold=bold, color=color, size=sz)
def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def center():   return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():     return Alignment(horizontal="left",   vertical="center", wrap_text=True)

def style_header_row(ws, row, n_cols, bg=C_HEADER, bold=True, sz=10):
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill    = hfill(bg)
        c.font    = hfont(bold=bold, color="FFFFFF", sz=sz)
        c.alignment = center()
        c.border  = border_thin()

def style_data_row(ws, row, n_cols, bg=C_WHITE, center_cols=None):
    center_cols = center_cols or []
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill   = hfill(bg)
        c.font   = hfont(sz=10)
        c.border = border_thin()
        c.alignment = center() if col in center_cols else left()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def pi_color(pi):
    if pi <= 7.5:  return C_PASS
    if pi <= 8.5:  return C_WARN
    return C_FAIL

def adi_color(adi):
    if adi >= 65:  return C_PASS
    if adi >= 50:  return C_WARN
    return C_FAIL

def ic_color(v, lo=0.05, hi=0.2):
    """Green ≤ lo, yellow ≤ hi, red > hi. None = grey."""
    if v is None: return "D9D9D9"
    if v <= lo:   return C_PASS
    if v <= hi:   return C_WARN
    return C_FAIL


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 1 — Cover / Overview
# ══════════════════════════════════════════════════════════════════════════════
def sheet_cover(wb):
    ws = wb.create_sheet("")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    rows = [
        ("",    "SARS-CoV-2 × MERS-CoV  VHH "),
        ("",    "v3.0  |  "),
        ("",    datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("",    "Tnb04 Tnb164.xlsx（12）"),
        ("",    "InSynBio AbEngineCore V4.4"),
        ("pI",  "BioPython ProteinAnalysis.isoelectric_point()"),
        ("ADI", "VHH42（n=42，tent-function）"),
        ("",    "SARS-CoV-2 / MERS-CoV （IC50/IC90）"),
        ("",            ""),
        ("","Tnb04H9 + Tnb164H6 + (G₄S)₃+3E"),
        ("","GGGGSGGGGSGGGGSEEE（18 aa，C3×Glu）"),
        ("pI","7.85（，8.800.95）"),
        ("","+1.0 @ pH 7（+4.0）"),
        ("",            ""),
        ("",    "Tnb04H9 + Tnb164H5 + (G₄S)₃+3E  |  pI=7.85"),
        ("",            ""),
        ("",    ""),
        ("  · ",   "sequences/  → Tnb04_panel.fasta  |  Tnb164_panel.fasta"),
        ("  · ",   "Sheet: Tnb04  |  Tnb164"),
        ("  · CMC",    "Sheet: CMC_  |  CMC_"),
        ("  · ",   "report/  → TNB_FINAL_DECISION_REPORT.md"),
    ]

    for i, (k, v) in enumerate(rows, 2):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)
        if k in ("", "", "pI", ""):
            ws.cell(i, 1).fill = hfill(C_PASS)
            ws.cell(i, 2).fill = hfill(C_PASS)
            ws.cell(i, 2).font = hfont(bold=True, sz=10)
        elif k in ("",):
            ws.cell(i, 1).fill = hfill(C_WARN)
            ws.cell(i, 2).fill = hfill(C_WARN)
        elif k == "":
            ws.cell(i, 1).font = hfont(bold=True, sz=12)
            ws.cell(i, 2).font = hfont(bold=True, sz=12)
        for col in (1, 2):
            ws.cell(i, col).border = border_thin()


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 2 — Tnb04 Neutralization
# ══════════════════════════════════════════════════════════════════════════════
def sheet_tnb04_activity(wb):
    ws = wb.create_sheet("Tnb04")
    set_col_widths(ws, [14, 6, 10, 10, 10, 10, 10, 10, 10, 12])

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = "Tnb04 SARS-CoV-2 （μg/mL）"
    ws["A1"].font   = hfont(bold=True, color="FFFFFF", sz=12)
    ws["A1"].fill   = hfill(C_HEADER)
    ws["A1"].alignment = center()

    # Sub-header: variant groups
    ws.merge_cells("C2:J2")
    ws["B2"] = ""
    ws["C2"] = "SARS-CoV-2 "
    for c in ("B2", "C2"):
        ws[c].font = hfont(bold=True, color="FFFFFF", sz=10)
        ws[c].fill = hfill(C_SUBHDR)
        ws[c].alignment = center()
        ws[c].border = border_thin()

    # Column headers
    hdrs = ["", "", "WT\n(D614G)", "JN.1\n(G339H\nK356T)", "KP.3.1.1\n(G339H\nK356T)", "XDV\n(G339H\nK356T)", "", "", "", ""]
    # row3
    ws.cell(3, 1, "")
    ws.cell(3, 2, "")
    ws.cell(3, 3, "WT (D614G)")
    ws.cell(3, 4, "JN.1")
    ws.cell(3, 5, "KP.3.1.1")
    ws.cell(3, 6, "XDV")
    ws.cell(3, 7, "JN.1 IC90")
    ws.cell(3, 8, "KP.3.1.1 IC90")
    ws.cell(3, 9, "XDV IC90")
    ws.cell(3, 10, "")
    style_header_row(ws, 3, 10, bg=C_SUBHDR)
    ws.row_dimensions[3].height = 30

    tnb04_ids = ["Tnb04H9", "Tnb04H4", "Tnb04H2", "Tnb04H3", "Tnb04H7", "Tnb04H8"]
    act_keys  = ["WT_IC50", "JN1_IC50", "KP_IC50", "XDV_IC50",
                 "JN1_IC90", "KP_IC90", "XDV_IC90"]

    row = 4
    for vid in tnb04_ids:
        a    = activity[vid]
        bg   = C_BEST if vid == "Tnb04H9" else (C_ALT if row % 2 == 0 else C_WHITE)

        for metric_label, key in [("IC50", ["WT_IC50","JN1_IC50","KP_IC50","XDV_IC50"]),
                                   ("IC90", ["JN1_IC90","KP_IC90","XDV_IC90"])]:
            ws.cell(row, 1, vid if metric_label == "IC50" else "")
            ws.cell(row, 2, metric_label)
            # IC50 row: cols 3-6; IC90 row: cols 7-9
            if metric_label == "IC50":
                cols = [3, 4, 5, 6]
                for ci, k in zip(cols, key):
                    v = a.get(k)
                    ws.cell(row, ci, v if v is not None else "n.d.")
                    ws.cell(row, ci).fill   = hfill(ic_color(v) if v is not None else "D9D9D9")
                    ws.cell(row, ci).font   = hfont(sz=10)
                    ws.cell(row, ci).alignment = center()
                    ws.cell(row, ci).border = border_thin()
                # Breadth rating
                vals = [a.get(k) for k in ["WT_IC50","JN1_IC50","KP_IC50","XDV_IC50"]]
                good = sum(1 for v in vals if v is not None and v <= 0.1)
                rating = "★★★ " if good == 4 else ("★★ " if good >= 3 else "★ ")
                ws.cell(row, 10, rating)
                ws.cell(row, 10).fill = hfill(C_PASS if good == 4 else (C_WARN if good >= 3 else C_FAIL))
                ws.cell(row, 10).font = hfont(bold=(vid=="Tnb04H9"), sz=10)
                ws.cell(row, 10).alignment = center()
                ws.cell(row, 10).border = border_thin()
            else:
                for ci, k in zip([7, 8, 9], key):
                    v = a.get(k)
                    ws.cell(row, ci, v if v is not None else "n.d.")
                    ws.cell(row, ci).fill   = hfill(ic_color(v, lo=0.5, hi=2.0) if v is not None else "D9D9D9")
                    ws.cell(row, ci).font   = hfont(sz=10)
                    ws.cell(row, ci).alignment = center()
                    ws.cell(row, ci).border = border_thin()

            for ci in [1, 2]:
                ws.cell(row, ci).fill   = hfill(C_BEST if vid=="Tnb04H9" else bg)
                ws.cell(row, ci).font   = hfont(bold=(ci==1 and metric_label=="IC50"), sz=10)
                ws.cell(row, ci).border = border_thin()
                ws.cell(row, ci).alignment = center()
            row += 1

    # Merge variant name cells
    for r_start in range(4, row, 2):
        ws.merge_cells(f"A{r_start}:A{r_start+1}")
        ws.cell(r_start, 1).alignment = center()

    # Note
    ws.cell(row + 1, 1, "： ≤0.05,  ≤0.2,  >0.2 μg/mL。 IC90： ≤0.5,  ≤2.0,  >2.0")
    ws.cell(row + 1, 1).font = hfont(sz=9, color="595959")
    ws.merge_cells(f"A{row+1}:J{row+1}")
    ws.cell(row + 2, 1, "★ ：Tnb04H9（4IC50≤0.037 μg/mL，）")
    ws.cell(row + 2, 1).font = hfont(bold=True, sz=10, color="375623")
    ws.cell(row + 2, 1).fill = hfill(C_PASS)
    ws.merge_cells(f"A{row+2}:J{row+2}")


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 3 — Tnb164 Neutralization
# ══════════════════════════════════════════════════════════════════════════════
def sheet_tnb164_activity(wb):
    ws = wb.create_sheet("Tnb164")
    set_col_widths(ws, [14, 12, 12, 12, 14])

    ws.merge_cells("A1:E1")
    ws["A1"] = "Tnb164 MERS-CoV （μg/mL）"
    ws["A1"].font  = hfont(bold=True, color="FFFFFF", sz=12)
    ws["A1"].fill  = hfill(C_HEADER)
    ws["A1"].alignment = center()

    hdrs = ["", "MERS WT IC50", "MjHKU4r-CoV-1\nIC50", "MjHKU4r-CoV-1\nIC90", ""]
    for ci, h in enumerate(hdrs, 1):
        ws.cell(2, ci, h)
    style_header_row(ws, 2, 5, bg=C_SUBHDR)
    ws.row_dimensions[2].height = 36

    tnb164_ids = ["Tnb164H4","Tnb164H5","Tnb164H2","Tnb164H6","Tnb164H7","Tnb164H8"]
    for i, vid in enumerate(tnb164_ids, 3):
        a    = activity[vid]
        bg   = C_BEST if vid == "Tnb164H6" else (C_ALT if i % 2 == 0 else C_WHITE)
        mwt  = a.get("MERS_WT_IC50")
        ic50 = a.get("MjHKU4r_IC50")
        ic90 = a.get("MjHKU4r_IC90", 999)

        def fv(v):
            if v is None: return "n.d."
            if v == 0.0:  return "≤0.001"
            return f"{v:.3f}"

        ws.cell(i, 1, vid)
        ws.cell(i, 2, fv(mwt))
        ws.cell(i, 3, fv(ic50))
        ws.cell(i, 4, fv(ic90) if ic90 != 999 else "n.d.")

        ic90_color = ic_color(ic90, lo=0.05, hi=0.3)
        rating = "★★★ " if ic90 <= 0.05 else ("★★ " if ic90 <= 0.3 else ("★ " if ic90 <= 0.5 else "✗ "))
        ws.cell(i, 5, rating)

        for ci in range(1, 6):
            c = ws.cell(i, ci)
            c.fill   = hfill(C_BEST if vid == "Tnb164H6" else bg)
            c.font   = hfont(bold=(ci == 1 and vid == "Tnb164H6"), sz=10)
            c.border = border_thin()
            c.alignment = center()
        ws.cell(i, 4).fill = hfill(ic90_color)
        ws.cell(i, 5).fill = hfill(C_PASS if ic90 <= 0.05 else (C_WARN if ic90 <= 0.3 else C_FAIL))

    row = len(tnb164_ids) + 4
    ws.cell(row, 1, "：MjHKU4r-CoV-1 = （MERS-CoV）。≤0.05, ≤0.3, >0.3 μg/mL")
    ws.cell(row, 1).font = hfont(sz=9, color="595959")
    ws.merge_cells(f"A{row}:E{row}")
    ws.cell(row+1, 1, "★ ：Tnb164H6（MjHKU4r IC90=0.025 μg/mL，，H44.8）")
    ws.cell(row+1, 1).font = hfont(bold=True, sz=10, color="375623")
    ws.cell(row+1, 1).fill = hfill(C_PASS)
    ws.merge_cells(f"A{row+1}:E{row+1}")


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 4 — CMC Single VHH
# ══════════════════════════════════════════════════════════════════════════════
def sheet_cmc_single(wb):
    ws = wb.create_sheet("CMC_VHH")
    set_col_widths(ws, [14, 7, 7, 8, 9, 7, 7, 7, 7, 7, 7, 7, 7, 8, 14])

    ws.merge_cells("A1:O1")
    ws["A1"] = " VHH CMC （，BioPython，vs VHH42 n=42）"
    ws["A1"].font  = hfont(bold=True, color="FFFFFF", sz=12)
    ws["A1"].fill  = hfill(C_HEADER)
    ws["A1"].alignment = center()

    hdrs = ["","pI","\n@pH7","GRAVY","\n","SAP","\n","\n","","","","\n","\nCys","ADI\n","CMC"]
    for ci, h in enumerate(hdrs, 1):
        ws.cell(2, ci, h)
    style_header_row(ws, 2, 15, bg=C_SUBHDR)
    ws.row_dimensions[2].height = 36

    tnb04_ids  = ["Tnb04H9","Tnb04H4","Tnb04H2","Tnb04H3","Tnb04H7","Tnb04H8"]
    tnb164_ids = ["Tnb164H4","Tnb164H5","Tnb164H2","Tnb164H6","Tnb164H7","Tnb164H8"]

    def write_cmc_rows(ws, ids, start_row):
        row = start_row
        for vid in ids:
            r   = singles[vid]
            m   = r["metrics"]
            adi = r["adi_continuous"]
            best_ids = {"Tnb04H9", "Tnb164H6"}
            bg = C_BEST if vid in best_ids else (C_ALT if row % 2 == 0 else C_WHITE)

            row_vals = [
                vid, m["pI"], round(m["net_charge_pH7"],1),
                round(m["GRAVY"],3), round(m["instability_index"],1),
                round(m["SAP_score"],3), m["agg_motifs"],
                m["hydro_cluster_count"], m["glycosylation_sites"],
                m["deamidation_sites"], m["isomerization_sites"],
                m["oxidation_sites"], m["free_cys"],
                round(adi,1), r["adi_grade"],
            ]
            for ci, v in enumerate(row_vals, 1):
                c = ws.cell(row, ci, v)
                c.border    = border_thin()
                c.alignment = center() if ci > 1 else left()
                c.font      = hfont(bold=(ci==1 and vid in best_ids), sz=10)
                c.fill      = hfill(bg)

            ws.cell(row, 2).fill  = hfill(pi_color(m["pI"]))
            ws.cell(row, 14).fill = hfill(adi_color(adi))
            ws.cell(row, 10).fill = hfill(C_WARN if m["deamidation_sites"] > 2 else bg)
            ws.cell(row, 12).fill = hfill(C_WARN if m["oxidation_sites"] > 6 else bg)
            row += 1
        return row

    row = write_cmc_rows(ws, tnb04_ids, 3)

    # Separator
    ws.merge_cells(f"A{row}:O{row}")
    ws.cell(row, 1).value = "── Tnb164 MERS  ──────────────────────────────────────"
    ws.cell(row, 1).fill  = hfill("D9E1F2")
    ws.cell(row, 1).font  = hfont(bold=True, sz=9, color="1F3864")
    row += 1

    row = write_cmc_rows(ws, tnb164_ids, row)

    # Legend row
    lr = row + 2
    ws.cell(lr, 1, "：")
    ws.merge_cells(f"A{lr}:O{lr}")
    ws.cell(lr, 1).font = hfont(bold=True, sz=9)
    legends = [
        (C_PASS,  " = （pI≤7.5 / ADI≥65 / ≤2 / ≤6）"),
        (C_WARN,  " = （pI 7.5-8.5 / ADI 50-65 / >2 / >6）"),
        (C_FAIL,  " = （pI>8.5 / ADI<50）"),
        (C_BEST,  " = "),
    ]
    for j, (clr, txt) in enumerate(legends, lr+1):
        ws.cell(j, 1, "  " + txt)
        ws.cell(j, 1).fill = hfill(clr)
        ws.cell(j, 1).font = hfont(sz=9)
        ws.merge_cells(f"A{j}:O{j}")


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 5 — CMC Fusion Proteins
# ══════════════════════════════════════════════════════════════════════════════
def sheet_cmc_fusion(wb):
    ws = wb.create_sheet("CMC_")
    set_col_widths(ws, [26, 18, 8, 8, 8, 9, 9, 14])

    ws.merge_cells("A1:H1")
    ws["A1"] = " CMC（VHH-Linker-VHH ，vs scFv_52 pI8.5）"
    ws["A1"].font  = hfont(bold=True, color="FFFFFF", sz=12)
    ws["A1"].fill  = hfill(C_HEADER)
    ws["A1"].alignment = center()

    hdrs = ["","","\n(aa)","\npI","\n@pH7","GRAVY","\n","pI"]
    for ci, h in enumerate(hdrs, 1):
        ws.cell(2, ci, h)
    style_header_row(ws, 2, 8, bg=C_SUBHDR)
    ws.row_dimensions[2].height = 36

    linker_order = ["(G4S)3", "(G4S)3+2E", "(G4S)3+3E", "(G4S)3+4E"]
    combos_order = [
        "Tnb04H9+Tnb164H4",
        "Tnb04H9+Tnb164H6",
        "Tnb04H9+Tnb164H5",
        "Tnb04H9+Tnb164H2",
        "Tnb04H2+Tnb164H6",
        "Tnb04H2+Tnb164H2",
    ]

    fmap = {(f["combo"], f["linker"]): f for f in fusions}

    BEST_COMBO = "Tnb04H9+Tnb164H6"
    BEST_LK    = "(G4S)3+3E"

    row = 3
    prev_combo = None
    for combo in combos_order:
        for lk in linker_order:
            f = fmap.get((combo, lk))
            if not f: continue
            is_best = (combo == BEST_COMBO and lk == BEST_LK)
            is_ref  = (combo == "Tnb04H9+Tnb164H4" and lk == "(G4S)3")
            bg = C_BEST if is_best else (C_FAIL[:6] + "20" if is_ref else
                 (C_ALT if row % 2 == 0 else C_WHITE))

            pi_rate = ("" if f["pI"] <= 7.5 else
                       "" if f["pI"] <= 8.0 else
                       "" if f["pI"] <= 8.5 else "")

            vals = [
                combo if combo != prev_combo else "",
                lk, f["full_len"],
                round(f["pI"],2), round(f["net_charge_pH7"],1),
                round(f["GRAVY"],3), round(f["instability_index"],1),
                pi_rate,
            ]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row, ci, v)
                c.border    = border_thin()
                c.alignment = center() if ci > 1 else left()
                c.font      = hfont(bold=is_best, sz=10,
                                    color="375623" if is_best else "000000")
                c.fill      = hfill(C_BEST if is_best else bg)

            ws.cell(row, 4).fill = hfill(pi_color(f["pI"]))
            ws.cell(row, 8).fill = hfill(
                C_PASS if f["pI"] <= 8.0 else
                (C_WARN if f["pI"] <= 8.5 else C_FAIL))

            prev_combo = combo
            row += 1

        # Empty separator row
        ws.cell(row, 1, "")
        row += 1

    # Note
    ws.cell(row+1, 1, f"★ ：Tnb04H9+Tnb164H6+(G₄S)₃+3E → pI=7.85，=+1.0，H9+H4+(G₄S)₃pI 1.09")
    ws.cell(row+1, 1).font = hfont(bold=True, sz=10, color="375623")
    ws.cell(row+1, 1).fill = hfill(C_PASS)
    ws.merge_cells(f"A{row+1}:H{row+1}")

    ws.cell(row+2, 1, "：GGGGSGGGGSGGGGSEEE（18 aa，C3pI）")
    ws.cell(row+2, 1).font = hfont(sz=9, color="595959")
    ws.merge_cells(f"A{row+2}:H{row+2}")


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET 6 — Comprehensive Balance Score
# ══════════════════════════════════════════════════════════════════════════════
def sheet_balance(wb):
    ws = wb.create_sheet("")
    set_col_widths(ws, [26, 10, 10, 12, 12, 10, 12])

    ws.merge_cells("A1:G1")
    ws["A1"] = "（ × CMC × ）"
    ws["A1"].font  = hfont(bold=True, color="FFFFFF", sz=12)
    ws["A1"].fill  = hfill(C_HEADER)
    ws["A1"].alignment = center()

    hdrs = ["(3E)", "SARS\n(H9:5, H2:3.5)", "MERS\n",
            "pI\n(3E)", "CMC\n(ADI)", "\n(/25)", ""]
    for ci, h in enumerate(hdrs, 1):
        ws.cell(2, ci, h)
    style_header_row(ws, 2, 7, bg=C_SUBHDR)
    ws.row_dimensions[2].height = 48

    combos = [
        # (label, sars_score, mers_score, fus_pi_3E, avg_adi, total, rec)
        ("Tnb04H9+Tnb164H6\n+(G₄S)₃+3E",  5.0, 5.0, 7.85, 56.0, 22.5, "★★★ "),
        ("Tnb04H9+Tnb164H5\n+(G₄S)₃+3E",  5.0, 2.5, 7.85, 60.5, 19.5, "★★ "),
        ("Tnb04H9+Tnb164H4\n+(G₄S)₃+3E",  5.0, 3.5, 8.31, 56.1, 18.5, "★ "),
        ("Tnb04H9+Tnb164H2\n+(G₄S)₃+2E",  5.0, 2.0, 7.85, 55.0, 17.0, ""),
        ("Tnb04H2+Tnb164H6\n+(G₄S)₃+3E",  3.5, 5.0, 7.85, 56.1, 18.5, "★ "),
        ("Tnb04H2+Tnb164H5\n+(G₄S)₃+3E",  3.5, 2.5, 7.85, 60.5, 16.5, ""),
        ("Tnb04H9+Tnb164H4\n+(G₄S)₃ ", 5.0, 3.5, 8.94, 56.1, 14.5, "（）"),
    ]

    for i, (label, sars, mers, pi, adi, total, rec) in enumerate(combos, 3):
        is_best = i == 3
        is_orig = "" in rec
        bg = C_BEST if is_best else (C_FAIL if is_orig else (C_ALT if i%2==0 else C_WHITE))

        row_vals = [label, sars, mers, pi, adi, total, rec]
        for ci, v in enumerate(row_vals, 1):
            c = ws.cell(i, ci, v)
            c.border    = border_thin()
            c.alignment = center() if ci > 1 else left()
            c.font      = hfont(bold=is_best, sz=10,
                                color="375623" if is_best else ("9C0006" if is_orig else "000000"))
            c.fill      = hfill(bg)

        # color score cells
        ws.cell(i, 2).fill = hfill(C_PASS if sars >= 4.5 else C_WARN)
        ws.cell(i, 3).fill = hfill(C_PASS if mers >= 4.5 else (C_WARN if mers >= 3.0 else C_FAIL))
        ws.cell(i, 4).fill = hfill(pi_color(pi))
        ws.cell(i, 6).fill = hfill(C_PASS if total >= 20 else (C_WARN if total >= 17 else C_FAIL))

    row = len(combos) + 4
    ws.cell(row, 1, "：SARS(5=4≤0.037, 3.5=3) | MERS(5=IC90≤0.025, 3.5=≤0.12, 2.5=≤0.35, 2.0=≤0.5) | pI2.5-5")
    ws.cell(row, 1).font = hfont(sz=9, color="595959")
    ws.merge_cells(f"A{row}:G{row}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    # Build Excel
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    sheet_cover(wb)
    sheet_tnb04_activity(wb)
    sheet_tnb164_activity(wb)
    sheet_cmc_single(wb)
    sheet_cmc_fusion(wb)
    sheet_balance(wb)

    xl_path = OUT_DIR / "activity" / "Tnb_Bispecific_Full_Analysis.xlsx"
    wb.save(xl_path)
    print(f"Excel saved: {xl_path}")

    # Copy method scripts
    methods_dir = OUT_DIR / "methods"
    for script in [
        "compute_scfv_cmc.py",
        "compute_tnb_full_cmc.py",
        "generate_tnb_bispecific_cmc_report.py",
        "compare_linker_designs.py",
    ]:
        src = SUITE_ROOT / "scripts" / script
        if src.exists():
            shutil.copy2(src, methods_dir / script)

    # Copy CMC JSON files
    for f in (SUITE_ROOT / "projects" / "Tnb_bispecific" / "cmc_eval").glob("*.json"):
        if f.name != f.name:
            pass  # already in place

    print(f"Methods scripts copied to: {methods_dir}")
    print(f"\nProject structure:")
    for p in sorted(OUT_DIR.rglob("*")):
        if p.is_file():
            rel = p.relative_to(OUT_DIR)
            print(f"  {rel}")


if __name__ == "__main__":
    main()
