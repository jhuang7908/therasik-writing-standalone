#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Build Therapeutic Evidence Layer for Framework Library

Processes Thera-SAbDab therapeutic antibody sequences to compute framework usage
statistics and link evidence to our framework library entries.

Requirements:
- ANARCII installed (pip install anarcii)
- Thera-SAbDab CSV/XLSX file in core/data/thera_sabdab/
- IMGT reference FASTA files (for germline matching)
- Framework library YAML files with Tier1/Tier2 tags
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import yaml

try:
    import openpyxl  # For XLSX support
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from core.numbering.imgt_anarcii import imgt_number_anarcii, IMGTNumberingError
    from core.vhh_humanization import split_regions, IMGT_REGIONS
    HAS_ANARCII = True
except ImportError as e:
    HAS_ANARCII = False
    print(f"Warning: ANARCII not available: {e}")

# IMGT region boundaries (FR1-FR3 for framework matching)
IMGT_FRAMEWORK_REGIONS = {
    "FR1": {"start": 1, "end": 26},
    "FR2": {"start": 39, "end": 55},
    "FR3": {"start": 66, "end": 104},
}


def calculate_sequence_identity(seq1: str, seq2: str) -> float:
    """Calculate sequence identity between two sequences."""
    if not seq1 or not seq2:
        return 0.0
    
    L = min(len(seq1), len(seq2))
    if L == 0:
        return 0.0
    
    same = sum(1 for i in range(L) if seq1[i] == seq2[i])
    return same / L


def extract_fr1_fr3_from_numbering(numbering_rows: List[Dict[str, Any]]) -> Optional[str]:
    """Extract FR1-FR3 concatenated sequence from ANARCII numbering results."""
    try:
        regions = split_regions(numbering_rows)
        fr1 = regions.get("FR1", "")
        fr2 = regions.get("FR2", "")
        fr3 = regions.get("FR3", "")
        return fr1 + fr2 + fr3
    except Exception:
        return None


def match_to_germline(
    therapeutic_fr: str,
    germline_fr_map: Dict[str, str],
) -> Tuple[Optional[str], float]:
    """
    Match therapeutic FR sequence to best-matching IMGT germline.
    
    Args:
        therapeutic_fr: FR1-FR3 concatenated sequence from therapeutic antibody
        germline_fr_map: {germline_id: fr1_fr3_sequence} mapping
        
    Returns:
        (best_match_germline_id, identity_score)
    """
    if not therapeutic_fr:
        return None, 0.0
    
    best_match = None
    best_identity = 0.0
    
    for germline_id, germline_fr in germline_fr_map.items():
        if not germline_fr or germline_fr == "TODO":
            continue
        
        identity = calculate_sequence_identity(therapeutic_fr, germline_fr)
        if identity > best_identity:
            best_identity = identity
            best_match = germline_id
    
    return best_match, best_identity


def load_thera_sabdab_csv(csv_path: Path) -> List[Dict[str, Any]]:
    """Load Thera-SAbDab CSV file."""
    entries = []
    
    if not csv_path.exists():
        return entries
    
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            entries.append(row)
    
    return entries


def load_thera_sabdab_xlsx(xlsx_path: Path) -> List[Dict[str, Any]]:
    """Load Thera-SAbDab XLSX file."""
    entries = []
    
    if not HAS_XLSX:
        print("Warning: openpyxl not installed. Install with: pip install openpyxl")
        return entries
    
    if not xlsx_path.exists():
        return entries
    
    from openpyxl import load_workbook
    
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    # Read header row
    headers = [cell.value for cell in ws[1]]
    
    # Read data rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = {}
        for i, value in enumerate(row):
            if i < len(headers):
                entry[headers[i]] = value
        if entry:
            entries.append(entry)
    
    return entries


def extract_sequences_from_entry(entry: Dict[str, Any]) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Extract VH, VL sequences and INN name from Thera-SAbDab entry.
    
    Common column names in Thera-SAbDab:
    - VH sequence: "VH_sequence", "Heavy_V_Sequence", "VH", etc.
    - VL sequence: "VL_sequence", "Light_V_Sequence", "VL", "VK", "VL_chain", etc.
    - INN: "INN", "Name", "Antibody_Name", "Therapeutic_Name", etc.
    
    Returns:
        (vh_sequence, vl_sequence, inn_name)
    """
    vh_seq = None
    vl_seq = None
    inn_name = None
    
    # Try various column name patterns
    vh_keys = ["VH_sequence", "Heavy_V_Sequence", "VH", "Heavy_Chain", "Heavy_V"]
    vl_keys = ["VL_sequence", "Light_V_Sequence", "VL", "VK", "VL_chain", "Light_Chain", "Light_V"]
    inn_keys = ["INN", "Name", "Antibody_Name", "Therapeutic_Name", "Antibody", "Drug_Name"]
    
    # Find VH sequence
    for key in vh_keys:
        if key in entry and entry[key]:
            vh_seq = str(entry[key]).strip().upper()
            if vh_seq and len(vh_seq) > 50:  # Reasonable VH length
                break
    
    # Find VL sequence
    for key in vl_keys:
        if key in entry and entry[key]:
            vl_seq = str(entry[key]).strip().upper()
            if vl_seq and len(vl_seq) > 40:  # Reasonable VL length
                break
    
    # Find INN name
    for key in inn_keys:
        if key in entry and entry[key]:
            inn_name = str(entry[key]).strip()
            if inn_name:
                break
    
    # Clean sequences (remove non-amino-acid characters)
    standard_aa = set("ACDEFGHIKLMNPQRSTVWY")
    
    if vh_seq:
        vh_seq = "".join(c for c in vh_seq if c in standard_aa)
        if len(vh_seq) < 50:
            vh_seq = None
    
    if vl_seq:
        vl_seq = "".join(c for c in vl_seq if c in standard_aa)
        if len(vl_seq) < 40:
            vl_seq = None
    
    return vh_seq, vl_seq, inn_name


def process_therapeutic_entry(
    entry: Dict[str, Any],
    germline_fr_map_vh: Dict[str, str],
    germline_fr_map_vl: Dict[str, str],
) -> Dict[str, Any]:
    """
    Process a single therapeutic antibody entry.
    
    Returns:
        Result dict with VH/VL germline assignments and metadata
    """
    result = {
        "inn": None,
        "vh": {
            "sequence": None,
            "germline_match": None,
            "identity": 0.0,
            "success": False,
        },
        "vl": {
            "sequence": None,
            "germline_match": None,
            "identity": 0.0,
            "success": False,
        },
        "error": None,
    }
    
    if not HAS_ANARCII:
        result["error"] = "ANARCII not available"
        return result
    
    # Extract sequences and INN
    vh_seq, vl_seq, inn_name = extract_sequences_from_entry(entry)
    result["inn"] = inn_name
    
    if not vh_seq and not vl_seq:
        result["error"] = "No VH or VL sequence found"
        return result
    
    # Process VH
    if vh_seq:
        try:
            numbering_rows = imgt_number_anarcii(vh_seq)
            fr1_fr3 = extract_fr1_fr3_from_numbering(numbering_rows)
            
            if fr1_fr3:
                germline_match, identity = match_to_germline(fr1_fr3, germline_fr_map_vh)
                result["vh"]["sequence"] = vh_seq
                result["vh"]["germline_match"] = germline_match
                result["vh"]["identity"] = identity
                result["vh"]["success"] = germline_match is not None and identity > 0.7  # Threshold
        except Exception as e:
            result["vh"]["error"] = str(e)
    
    # Process VL
    if vl_seq:
        try:
            numbering_rows = imgt_number_anarcii(vl_seq)
            fr1_fr3 = extract_fr1_fr3_from_numbering(numbering_rows)
            
            if fr1_fr3:
                germline_match, identity = match_to_germline(fr1_fr3, germline_fr_map_vl)
                result["vl"]["sequence"] = vl_seq
                result["vl"]["germline_match"] = germline_match
                result["vl"]["identity"] = identity
                result["vl"]["success"] = germline_match is not None and identity > 0.7  # Threshold
        except Exception as e:
            result["vl"]["error"] = str(e)
    
    return result


def load_framework_library(yaml_path: Path) -> Tuple[List[Dict[str, Any]], Dict[str, str]]:
    """
    Load framework library YAML and build germline FR map.
    
    Returns:
        (frameworks_list, {germline_id: fr1_fr3_sequence})
    """
    frameworks = []
    germline_fr_map = {}
    
    if not yaml_path.exists():
        return frameworks, germline_fr_map
    
    with open(yaml_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
        if data and "frameworks" in data:
            frameworks = data["frameworks"]
            
            for fw in frameworks:
                germline_id = fw.get("germline")
                fr_seq = fw.get("fr_sequence_fr1_fr3")
                
                if germline_id and fr_seq and fr_seq != "TODO":
                    germline_fr_map[germline_id] = fr_seq
    
    return frameworks, germline_fr_map


def compute_statistics(
    results: List[Dict[str, Any]],
    vh_frameworks: List[Dict[str, Any]],
    vl_frameworks: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """Compute framework usage statistics from therapeutic results."""
    stats = {
        "total_entries": len(results),
        "vh_assignments": defaultdict(int),
        "vl_assignments": defaultdict(int),
        "vh_vl_pairings": defaultdict(int),
        "framework_evidence": {},
    }
    
    # Count assignments
    for result in results:
        if result["vh"]["success"] and result["vh"]["germline_match"]:
            germline = result["vh"]["germline_match"]
            stats["vh_assignments"][germline] += 1
        
        if result["vl"]["success"] and result["vl"]["germline_match"]:
            germline = result["vl"]["germline_match"]
            stats["vl_assignments"][germline] += 1
        
        # Track pairings
        if (result["vh"]["success"] and result["vh"]["germline_match"] and
            result["vl"]["success"] and result["vl"]["germline_match"]):
            pairing = f"{result['vh']['germline_match']} + {result['vl']['germline_match']}"
            stats["vh_vl_pairings"][pairing] += 1
    
    # Build evidence for Tier1/Tier2 frameworks
    tier1_tier2_vh = {}
    tier1_tier2_vl = {}
    
    for fw in vh_frameworks:
        tags = fw.get("tags", [])
        if "tier1" in tags or "tier2" in tags:
            germline = fw.get("germline")
            if germline:
                tier1_tier2_vh[germline] = fw
    
    for fw in vl_frameworks:
        tags = fw.get("tags", [])
        if "tier1" in tags or "tier2" in tags:
            germline = fw.get("germline")
            if germline:
                tier1_tier2_vl[germline] = fw
    
    # Collect example therapeutics for each Tier1/Tier2 framework
    for germline, fw in tier1_tier2_vh.items():
        example_inns = []
        count = stats["vh_assignments"].get(germline, 0)
        
        for result in results:
            if (result["vh"]["success"] and 
                result["vh"]["germline_match"] == germline and 
                result["inn"]):
                example_inns.append(result["inn"])
        
        stats["framework_evidence"][f"VH:{germline}"] = {
            "germline": germline,
            "chain": "VH",
            "count_in_dataset": count,
            "example_therapeutics": example_inns[:10],  # Top 10 examples
            "notes": f"Based on Thera-SAbDab dataset. {count} therapeutic antibodies matched to this germline.",
        }
    
    for germline, fw in tier1_tier2_vl.items():
        example_inns = []
        count = stats["vl_assignments"].get(germline, 0)
        
        for result in results:
            if (result["vl"]["success"] and 
                result["vl"]["germline_match"] == germline and 
                result["inn"]):
                example_inns.append(result["inn"])
        
        stats["framework_evidence"][f"VL:{germline}"] = {
            "germline": germline,
            "chain": "VL",
            "count_in_dataset": count,
            "example_therapeutics": example_inns[:10],  # Top 10 examples
            "notes": f"Based on Thera-SAbDab dataset. {count} therapeutic antibodies matched to this germline.",
        }
    
    return stats


def generate_evidence_report(
    stats: Dict[str, Any],
    output_path: Path,
) -> str:
    """Generate markdown evidence report."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    lines = [
        "# Framework Library Therapeutic Evidence",
        "",
        f"**Generated:** {timestamp}",
        f"**Source:** Thera-SAbDab therapeutic antibody database",
        "",
        "## Summary",
        "",
        f"- **Total Therapeutic Entries Processed:** {stats['total_entries']}",
        f"- **VH Germlines with Matches:** {len(stats['vh_assignments'])}",
        f"- **VL Germlines with Matches:** {len(stats['vl_assignments'])}",
        f"- **Tier1/Tier2 Frameworks with Evidence:** {len(stats['framework_evidence'])}",
        "",
        "## VH Germline Distribution",
        "",
        "| Germline | Count |",
        "|----------|-------|",
    ]
    
    # Sort VH assignments by count
    vh_sorted = sorted(stats["vh_assignments"].items(), key=lambda x: x[1], reverse=True)
    for germline, count in vh_sorted[:20]:  # Top 20
        lines.append(f"| {germline} | {count} |")
    
    lines.extend([
        "",
        "## VL Germline Distribution",
        "",
        "| Germline | Count |",
        "|----------|-------|",
    ])
    
    # Sort VL assignments by count
    vl_sorted = sorted(stats["vl_assignments"].items(), key=lambda x: x[1], reverse=True)
    for germline, count in vl_sorted[:20]:  # Top 20
        lines.append(f"| {germline} | {count} |")
    
    lines.extend([
        "",
        "## Common VH–VL Pairings",
        "",
        "| Pairing | Count |",
        "|---------|-------|",
    ])
    
    # Sort pairings by count
    pairing_sorted = sorted(stats["vh_vl_pairings"].items(), key=lambda x: x[1], reverse=True)
    for pairing, count in pairing_sorted[:20]:  # Top 20
        lines.append(f"| {pairing} | {count} |")
    
    lines.extend([
        "",
        "## Tier1/Tier2 Framework Evidence",
        "",
    ])
    
    # Group by chain
    vh_evidence = {k: v for k, v in stats["framework_evidence"].items() if k.startswith("VH:")}
    vl_evidence = {k: v for k, v in stats["framework_evidence"].items() if k.startswith("VL:")}
    
    if vh_evidence:
        lines.extend([
            "### VH Frameworks",
            "",
            "| Framework | Count | Example Therapeutics |",
            "|-----------|-------|---------------------|",
        ])
        
        for fw_id, evidence in sorted(vh_evidence.items()):
            germline = evidence["germline"]
            count = evidence["count_in_dataset"]
            examples = ", ".join(evidence["example_therapeutics"][:5])  # First 5
            if not examples:
                examples = "N/A"
            lines.append(f"| {germline} | {count} | {examples} |")
        
        lines.append("")
    
    if vl_evidence:
        lines.extend([
            "### VL Frameworks",
            "",
            "| Framework | Count | Example Therapeutics |",
            "|-----------|-------|---------------------|",
        ])
        
        for fw_id, evidence in sorted(vl_evidence.items()):
            germline = evidence["germline"]
            count = evidence["count_in_dataset"]
            examples = ", ".join(evidence["example_therapeutics"][:5])  # First 5
            if not examples:
                examples = "N/A"
            lines.append(f"| {germline} | {count} | {examples} |")
        
        lines.append("")
    
    lines.extend([
        "## Notes",
        "",
        "- All statistics are based solely on Thera-SAbDab dataset.",
        "- Germline assignments are based on FR1-FR3 sequence identity matching.",
        "- Only Tier1/Tier2 frameworks from our library are shown in evidence section.",
        "- Example therapeutics list is limited to first 10 matches per framework.",
    ])
    
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(
        description="Build therapeutic evidence layer for framework library"
    )
    parser.add_argument(
        "--thera_file",
        type=Path,
        help="Path to Thera-SAbDab CSV or XLSX file (if not provided, will search in core/data/thera_sabdab/)",
    )
    parser.add_argument(
        "--thera_dir",
        type=Path,
        default=PROJECT_ROOT / "core" / "data" / "thera_sabdab",
        help="Directory containing Thera-SAbDab files",
    )
    parser.add_argument(
        "--vh_yaml",
        type=Path,
        default=PROJECT_ROOT / "core" / "data" / "framework_library" / "vh_frameworks.yaml",
        help="Path to VH frameworks YAML file",
    )
    parser.add_argument(
        "--vl_yaml",
        type=Path,
        default=PROJECT_ROOT / "core" / "data" / "framework_library" / "vl_frameworks.yaml",
        help="Path to VL frameworks YAML file",
    )
    parser.add_argument(
        "--output_report",
        type=Path,
        default=PROJECT_ROOT / "docs" / "framework_library_thera_evidence.md",
        help="Path to output markdown report",
    )
    parser.add_argument(
        "--output_json",
        type=Path,
        default=PROJECT_ROOT / "core" / "data" / "framework_library" / "thera_stats.json",
        help="Path to output JSON statistics",
    )
    
    args = parser.parse_args()
    
    print("=" * 80)
    print("Build Therapeutic Evidence Layer for Framework Library")
    print("=" * 80)
    print()
    
    if not HAS_ANARCII:
        print("❌ ERROR: ANARCII is not available. Install with: pip install anarcii")
        return 1
    
    # Step 1: Find Thera-SAbDab file
    print("[1/5] Locating Thera-SAbDab file...")
    
    if args.thera_file:
        thera_path = args.thera_file
    else:
        # Search in thera_dir
        thera_dir = args.thera_dir
        thera_dir.mkdir(parents=True, exist_ok=True)
        
        # Look for CSV or XLSX files
        csv_files = list(thera_dir.glob("*.csv"))
        xlsx_files = list(thera_dir.glob("*.xlsx")) if HAS_XLSX else []
        
        if csv_files:
            thera_path = csv_files[0]
        elif xlsx_files:
            thera_path = xlsx_files[0]
        else:
            print(f"  ❌ No Thera-SAbDab file found in {thera_dir}")
            print(f"  Please download from https://opig.stats.ox.ac.uk/webapps/thera-sabdab/")
            print(f"  and place CSV or XLSX file in {thera_dir}")
            return 1
    
    if not thera_path.exists():
        print(f"  ❌ File not found: {thera_path}")
        return 1
    
    print(f"  ✅ Found: {thera_path}")
    print()
    
    # Step 2: Load Thera-SAbDab data
    print("[2/5] Loading Thera-SAbDab data...")
    
    if thera_path.suffix.lower() == ".csv":
        entries = load_thera_sabdab_csv(thera_path)
    elif thera_path.suffix.lower() == ".xlsx":
        entries = load_thera_sabdab_xlsx(thera_path)
    else:
        print(f"  ❌ Unsupported file format: {thera_path.suffix}")
        return 1
    
    print(f"  ✅ Loaded {len(entries)} entries")
    print()
    
    # Step 3: Load framework libraries
    print("[3/5] Loading framework libraries...")
    
    vh_frameworks, vh_germline_fr_map = load_framework_library(args.vh_yaml)
    vl_frameworks, vl_germline_fr_map = load_framework_library(args.vl_yaml)
    
    print(f"  ✅ Loaded {len(vh_frameworks)} VH frameworks ({len(vh_germline_fr_map)} with FR sequences)")
    print(f"  ✅ Loaded {len(vl_frameworks)} VL frameworks ({len(vl_germline_fr_map)} with FR sequences)")
    print()
    
    if not vh_germline_fr_map and not vl_germline_fr_map:
        print("  ⚠️  Warning: No FR sequences found in framework libraries.")
        print("  Run build_framework_library_from_imgt.py first to populate FR sequences.")
        print()
    
    # Step 4: Process therapeutic entries
    print("[4/5] Processing therapeutic entries...")
    
    results = []
    success_count = 0
    
    for i, entry in enumerate(entries, 1):
        if i % 100 == 0:
            print(f"  Processing: {i}/{len(entries)}")
        
        result = process_therapeutic_entry(entry, vh_germline_fr_map, vl_germline_fr_map)
        results.append(result)
        
        if result["vh"]["success"] or result["vl"]["success"]:
            success_count += 1
    
    print(f"  ✅ Processed {len(results)} entries ({success_count} with successful assignments)")
    print()
    
    # Step 5: Compute statistics and generate outputs
    print("[5/5] Computing statistics and generating outputs...")
    
    stats = compute_statistics(results, vh_frameworks, vl_frameworks)
    
    # Save JSON
    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output_json, "w", encoding="utf-8") as f:
        json.dump(stats, f, indent=2, ensure_ascii=False)
    print(f"  ✅ Saved JSON: {args.output_json}")
    
    # Generate and save report
    report_content = generate_evidence_report(stats, args.output_report)
    args.output_report.parent.mkdir(parents=True, exist_ok=True)
    with open(args.output_report, "w", encoding="utf-8") as f:
        f.write(report_content)
    print(f"  ✅ Saved report: {args.output_report}")
    print()
    
    print("=" * 80)
    print("✅ Build complete!")
    print("=" * 80)
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
