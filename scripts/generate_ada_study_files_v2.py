#!/usr/bin/env python3
"""
ADA Study Files v2 — generates 6 study-ready files:
  confirmed_ada.xlsx / .md / .pdf
  need_fulltext.xlsx  / .md / .pdf

New in v2:
  - Target & Disease columns
  - ADA value column  +  ADA evidence chain column  (separate)
  - PMID shown explicitly when available
  - ADA clinical context extracted from evidence_chain
"""
from __future__ import annotations
import json, re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (BaseDocTemplate, Frame, HRFlowable,
                                 PageBreak, PageTemplate, Paragraph,
                                 Spacer, Table, TableStyle)

# ─── Paths ───────────────────────────────────────────────────────────────────
REPO     = Path(__file__).resolve().parents[1]
FTF_DIR  = REPO / "data/ADA_reliable_package/final_three_files"
CLIN_DB  = REPO / "data/ADA_reliable_package/clinical_db/clinical_ada_db_data.json"
OUT_DIR  = REPO / "data/ADA_reliable_package/study_materials"
OUT_DIR.mkdir(parents=True, exist_ok=True)

TODAY    = datetime.now().strftime("%Y-%m-%d")

# ─── Chinese font ─────────────────────────────────────────────────────────────
for _fp in [r"C:\Windows\Fonts\msyh.ttc", r"C:\Windows\Fonts\msyhbd.ttc",
            r"C:\Windows\Fonts\simhei.ttf", r"C:\Windows\Fonts\simsun.ttc"]:
    if Path(_fp).exists():
        try:
            pdfmetrics.registerFont(TTFont("CNFont", _fp))
            CN_FONT = "CNFont"; break
        except Exception: pass
else:
    CN_FONT = "Helvetica"

# ══════════════════════════════════════════════════════════════════════════════
#  ANTIBODY LOOKUP TABLE  (Target · Disease)
# ══════════════════════════════════════════════════════════════════════════════

AB_INFO: dict[str, tuple[str, str]] = {
    # name_lower : (target, indications)
    "abatacept":         ("CD80/CD86 (CTLA4-Ig)",   "Rheumatoid arthritis, Juvenile idiopathic arthritis, Psoriatic arthritis"),
    "adalimumab":        ("TNF-α",                   "Rheumatoid arthritis, Crohn's disease, Psoriasis, Ankylosing spondylitis, UC, Uveitis"),
    "alemtuzumab":       ("CD52",                    "Relapsing-remitting multiple sclerosis, B-CLL"),
    "alirocumab":        ("PCSK9",                   "Hypercholesterolemia, Cardiovascular risk reduction"),
    "atezolizumab":      ("PD-L1",                   "NSCLC, Bladder cancer, Triple-negative breast cancer, HCC"),
    "atoltivimab":       ("Ebola virus GP",          "Ebola virus disease"),
    "avelumab":          ("PD-L1",                   "Merkel cell carcinoma, Urothelial carcinoma"),
    "basiliximab":       ("IL-2Rα (CD25)",           "Renal transplant rejection prophylaxis"),
    "belimumab":         ("BLyS/BAFF",               "Systemic lupus erythematosus (SLE), Lupus nephritis"),
    "bevacizumab":       ("VEGF-A",                  "Colorectal cancer, NSCLC, Glioblastoma, RCC, Cervical cancer, Ovarian cancer"),
    "bimekizumab":       ("IL-17A/IL-17F",           "Plaque psoriasis, Psoriatic arthritis, Ankylosing spondylitis"),
    "brolucizumab":      ("VEGF-A",                  "Neovascular (wet) AMD, Diabetic macular edema"),
    "burosumab":         ("FGF23",                   "X-linked hypophosphatemia, Tumor-induced osteomalacia"),
    "caplacizumab":      ("vWF A1 domain",           "Acquired thrombotic thrombocytopenic purpura (aTTP)"),
    "cemiplimab":        ("PD-1",                    "Cutaneous squamous cell carcinoma, BCC, NSCLC"),
    "certolizumab pegol":("TNF-α",                   "Rheumatoid arthritis, Crohn's disease, Psoriatic arthritis, Ankylosing spondylitis"),
    "cetuximab":         ("EGFR",                    "Colorectal cancer, Head and neck squamous cell carcinoma"),
    "daclizumab":        ("IL-2Rα (CD25)",           "Multiple sclerosis (withdrawn 2018)"),
    "daratumumab":       ("CD38",                    "Multiple myeloma"),
    "denosumab":         ("RANK-L",                  "Osteoporosis, Bone metastases, Giant cell tumor of bone"),
    "dinutuximab":       ("GD2",                     "Neuroblastoma"),
    "dupilumab":         ("IL-4Rα (IL-4/IL-13)",     "Atopic dermatitis, Asthma, CRS with nasal polyps, EoE, Prurigo nodularis"),
    "durvalumab":        ("PD-L1",                   "NSCLC, SCLC, Biliary tract cancer, HCC"),
    "eculizumab":        ("Complement C5",           "PNH, aHUS, Neuromyelitis optica, Myasthenia gravis"),
    "efalizumab":        ("CD11a (LFA-1)",           "Plaque psoriasis (withdrawn 2009)"),
    "elotuzumab":        ("SLAMF7 (CS1)",            "Relapsed/refractory multiple myeloma"),
    "emicizumab":        ("FIXa/FX (bispecific)",    "Hemophilia A (with/without inhibitors)"),
    "eptinezumab":       ("CGRP",                    "Migraine prevention"),
    "evolocumab":        ("PCSK9",                   "Hypercholesterolemia, Homozygous familial hypercholesterolemia, CV risk"),
    "faricimab":         ("VEGF-A / Ang-2",          "Neovascular AMD, Diabetic macular edema"),
    "favezelimab":       ("LAG-3",                   "Relapsed/refractory CLL (in combination)"),
    "fremanezumab":      ("CGRP",                    "Migraine prevention"),
    "galcanezumab":      ("CGRP",                    "Migraine prevention, Cluster headache"),
    "gemtuzumab ozogamicin": ("CD33",                "Acute myeloid leukemia"),
    "golimumab":         ("TNF-α",                   "Rheumatoid arthritis, Ankylosing spondylitis, Psoriatic arthritis, UC"),
    "guselkumab":        ("IL-23p19",                "Plaque psoriasis, Psoriatic arthritis"),
    "idarucizumab":      ("Dabigatran",              "Reversal of dabigatran anticoagulation"),
    "infliximab":        ("TNF-α",                   "Rheumatoid arthritis, Crohn's disease, UC, Ankylosing spondylitis, Psoriasis"),
    "ipilimumab":        ("CTLA-4",                  "Melanoma, RCC, NSCLC, HCC, MSI-H CRC (combination)"),
    "isatuximab":        ("CD38",                    "Multiple myeloma"),
    "ixekizumab":        ("IL-17A",                  "Plaque psoriasis, Psoriatic arthritis, Ankylosing spondylitis"),
    "lecanemab":         ("Amyloid-β protofibrils",  "Early Alzheimer's disease"),
    "mepolizumab":       ("IL-5",                    "Severe eosinophilic asthma, EGPA, Hypereosinophilic syndrome"),
    "mogamulizumab":     ("CCR4",                    "Mycosis fungoides, Sézary syndrome"),
    "natalizumab":       ("α4-integrin (VLA-4)",     "Relapsing-remitting multiple sclerosis, Crohn's disease"),
    "naxitamab":         ("GD2",                     "Relapsed/refractory high-risk neuroblastoma"),
    "necitumumab":       ("EGFR",                    "Metastatic squamous NSCLC"),
    "nirsevimab":        ("RSV F-protein",           "RSV prevention in infants and young children"),
    "nivolumab":         ("PD-1",                    "Melanoma, NSCLC, RCC, Hodgkin lymphoma, HCC, MSI-H CRC, Gastric cancer"),
    "obinutuzumab":      ("CD20",                    "CLL, Follicular lymphoma"),
    "ocrelizumab":       ("CD20",                    "Multiple sclerosis (RRMS, PPMS)"),
    "ofatumumab":        ("CD20",                    "CLL, Relapsing MS"),
    "omalizumab":        ("IgE",                     "Allergic asthma, Chronic urticaria, Nasal polyps"),
    "ozoralizumab":      ("TNF-α (nanobody-based)",  "Rheumatoid arthritis"),
    "palivizumab":       ("RSV F-protein",           "RSV prevention in premature infants and high-risk children"),
    "panitumumab":       ("EGFR",                    "Metastatic colorectal cancer (RAS wild-type)"),
    "pembrolizumab":     ("PD-1",                    "Melanoma, NSCLC, HNSCC, Bladder cancer, MSI-H tumors, Cervical cancer, HCC"),
    "pertuzumab":        ("HER2 (domain II)",        "HER2+ breast cancer (neoadjuvant, adjuvant, metastatic)"),
    "polatuzumab vedotin":("CD79b (ADC)",            "Diffuse large B-cell lymphoma"),
    "ramucirumab":       ("VEGFR-2",                 "Gastric/GEJ cancer, NSCLC, Colorectal cancer, HCC"),
    "ranibizumab":       ("VEGF-A",                  "Neovascular AMD, Macular edema, Diabetic macular edema"),
    "ravulizumab":       ("Complement C5",           "PNH, aHUS, Myasthenia gravis, NMOSD"),
    "rilonacept":        ("IL-1α/IL-1β (trap)",      "Cryopyrin-associated periodic syndrome (CAPS), Recurrent pericarditis"),
    "risankizumab":      ("IL-23p19",                "Plaque psoriasis, Psoriatic arthritis, Crohn's disease, UC"),
    "rituximab":         ("CD20",                    "NHL, CLL, RA, ANCA-associated vasculitis, Pemphigus vulgaris"),
    "romosozumab":       ("Sclerostin",              "Osteoporosis (postmenopausal women)"),
    "sacituzumab govitecan":("Trop-2 (ADC)",         "Triple-negative breast cancer, Urothelial carcinoma"),
    "sarilumab":         ("IL-6Rα",                  "Rheumatoid arthritis"),
    "satralizumab":      ("IL-6R",                   "Neuromyelitis optica spectrum disorder (NMOSD)"),
    "secukinumab":       ("IL-17A",                  "Plaque psoriasis, Psoriatic arthritis, Ankylosing spondylitis, nr-axSpA"),
    "siltuximab":        ("IL-6",                    "Idiopathic multicentric Castleman's disease"),
    "sintilimab":        ("PD-1",                    "Hodgkin lymphoma, NSCLC, HCC, Gastric cancer"),
    "sirukumab":         ("IL-6",                    "Rheumatoid arthritis (development discontinued)"),
    "spesolimab":        ("IL-36R",                  "Generalized pustular psoriasis"),
    "sutimlimab":        ("Complement C1s",          "Cold agglutinin disease"),
    "tanezumab":         ("NGF",                     "Chronic low back pain, Osteoarthritis (development discontinued)"),
    "tezepelumab":       ("TSLP",                    "Severe asthma"),
    "tislelizumab":      ("PD-1",                    "Hodgkin lymphoma, Urothelial carcinoma, HCC, NSCLC"),
    "tocilizumab":       ("IL-6R",                   "Rheumatoid arthritis, Giant cell arteritis, CRS, Systemic JIA"),
    "toripalimab":       ("PD-1",                    "Nasopharyngeal carcinoma, Melanoma (China approval)"),
    "tralokinumab":      ("IL-13",                   "Atopic dermatitis"),
    "trastuzumab":       ("HER2",                    "HER2+ breast cancer, HER2+ gastric/GEJ cancer"),
    "tremelimumab":      ("CTLA-4",                  "Hepatocellular carcinoma (combination), NSCLC (combination)"),
    "ustekinumab":       ("IL-12/IL-23 p40",         "Plaque psoriasis, Psoriatic arthritis, Crohn's disease, UC"),
    "vedolizumab":       ("α4β7 integrin",           "Ulcerative colitis, Crohn's disease"),
    "zanidatamab":       ("HER2",                    "Biliary tract cancer, Gastroesophageal junction adenocarcinoma"),
    "zinpentraximab":    ("Pentraxin-3",             "IgA nephropathy (investigational)"),
    "zolbetuximab":      ("Claudin 18.2",            "HER2-negative gastric/GEJ adenocarcinoma"),
    # additional entries likely in the 116
    "abciximab":         ("GPIIb/IIIa",              "Prevention of cardiac ischemic complications (PCI)"),
    "ado-trastuzumab emtansine": ("HER2 (ADC)",      "HER2+ breast cancer"),
    "amivantamab":       ("EGFR/MET (bispecific)",   "NSCLC with EGFR exon 20 insertion mutations"),
    "anifrolumab":       ("IFNAR1 (Type I IFN-R)",   "Systemic lupus erythematosus"),
    "arcitumomab":       ("CEA",                     "Colorectal cancer imaging (diagnostic)"),
    "belantamab mafodotin": ("BCMA (ADC)",           "Relapsed/refractory multiple myeloma"),
    "benralizumab":      ("IL-5Rα",                  "Severe eosinophilic asthma"),
    "bezlotoxumab":      ("C. difficile toxin B",    "Prevention of C. difficile infection recurrence"),
    "blinatumomab":      ("CD19/CD3 (BiTE)",         "Relapsed/refractory B-ALL"),
    "brentuximab vedotin":("CD30 (ADC)",             "Hodgkin lymphoma, PTCL, CTCL"),
    "brodalumab":        ("IL-17RA",                 "Plaque psoriasis"),
    "cadonilimab":       ("PD-1/CTLA-4 (bispecific)","Cervical cancer"),
    "canakinumab":       ("IL-1β",                   "CAPS, SJIA, AOSD, Gout flares"),
    "crizanlizumab":     ("P-selectin",              "Sickle cell disease vaso-occlusive crises"),
    "clesrovimab":       ("RSV F-protein",           "RSV prevention in infants"),
    "darinaparsin":      ("CD38/CD138",              "Myeloma (investigational)"),
    "depemokimab":       ("IL-5",                    "Severe eosinophilic asthma (ultra-long-acting)"),
    "dostarlimab":       ("PD-1",                    "dMMR/MSI-H endometrial cancer, Solid tumors"),
    "erenumab":          ("CGRP-R",                  "Migraine prevention"),
    "etrolizumab":       ("α4β7/αEβ7 integrin",      "Ulcerative colitis (development discontinued)"),
    "gemtuzumab":        ("CD33",                    "Acute myeloid leukemia"),
    "ibritumomab tiuxetan": ("CD20",                 "Relapsed/refractory follicular NHL, transformed NHL"),
    "infliximab biosimilar": ("TNF-α",               "Same as infliximab"),
    "itepekimab":        ("IL-33",                   "Asthma, COPD"),
    "ixekizumab":        ("IL-17A",                  "Plaque psoriasis, Psoriatic arthritis, Ankylosing spondylitis"),
    "lanadelumab":       ("Plasma kallikrein",        "Hereditary angioedema prevention"),
    "lebrikizumab":      ("IL-13",                   "Atopic dermatitis"),
    "loncastuximab tesirine": ("CD19 (ADC)",         "Relapsed/refractory DLBCL"),
    "marstacimab":       ("Tissue factor pathway inhibitor", "Hemophilia A/B with inhibitors"),
    "moxetumomab pasudotox": ("CD22",                "Relapsed/refractory hairy cell leukemia"),
    "narsoplimab":       ("MASP-2",                  "HSCT-associated TMA"),
    "nemolizumab":       ("IL-31RA",                 "Atopic dermatitis, Prurigo nodularis"),
    "obiltoxaximab":     ("Anthrax protective antigen", "Anthrax (Bacillus anthracis)"),
    "olaratumab":        ("PDGFRα",                  "Soft tissue sarcoma (withdrawn)"),
    "oteseconazole":     ("CYP51",                   "Recurrent vulvovaginal candidiasis (small molecule, not antibody)"),
    "ozanezumab":        ("NgR1/Nogo-A receptor",    "ALS (development discontinued)"),
    "polatuzumab":       ("CD79b (ADC)",             "DLBCL"),
    "ravagalimab":       ("CD40",                    "Crohn's disease (development discontinued)"),
    "rimegepant":        ("CGRP-R",                  "Migraine (small molecule, not antibody)"),
    "rozanolixizumab":   ("FcRn",                    "Immune thrombocytopenia, Generalized myasthenia gravis"),
    "ruplizumab":        ("CD154 (CD40L)",            "Lupus nephritis (discontinued)"),
    "satralizumab":      ("IL-6R",                   "Neuromyelitis optica spectrum disorder (NMOSD)"),
    "setrusumab":        ("Sclerostin",              "Osteogenesis imperfecta"),
    "spartalizumab":     ("PD-1",                    "Solid tumors (development discontinued)"),
    "spesolimab":        ("IL-36R",                  "Generalized pustular psoriasis"),
    "sutimlimab":        ("C1s complement protease", "Cold agglutinin disease"),
    "tafolecimab":       ("PCSK9",                   "Hypercholesterolemia (China approval)"),
    "tarlatamab":        ("DLL3/CD3 (BiTE)",         "Extensive-stage SCLC"),
    "tebentafusp":       ("gp100/CD3",               "Uveal melanoma"),
    "teprotumumab":      ("IGF-1R",                  "Thyroid eye disease"),
    "tezepelumab":       ("TSLP",                    "Severe uncontrolled asthma"),
    "tiragolumab":       ("TIGIT",                   "NSCLC (combination, investigational)"),
    "tovetumab":         ("HGF/c-MET",               "Solid tumors (investigational)"),
    "ublituximab":       ("CD20",                    "Multiple sclerosis (RRMS, SPMS)"),
    "vaborbactam":       ("β-lactamase",             "Bacterial infection (not antibody)"),
    "vibostolimab":      ("TIGIT",                   "Solid tumors (investigational)"),
    "volrustomig":       ("PD-1/TIGIT (bispecific)", "NSCLC (investigational)"),
    "zanubrutinib":      ("BTK (small molecule)",    "B-cell malignancies"),
    "ziralimumab":       ("CD127 (IL-7Rα)",          "Solid tumors (investigational)"),
    "axatilimab":        ("CSF-1R",                  "Chronic graft-versus-host disease"),
    "domvanalimab":      ("TIGIT",                   "NSCLC (combination, investigational)"),
    "amivantamab":       ("EGFR/MET (bispecific)",   "NSCLC (exon 20 insertions)"),
    "toripalimab":       ("PD-1",                    "Nasopharyngeal carcinoma, Melanoma, Esophageal squamous cell carcinoma"),
    # ── antibodies identified as missing during field audit ──────────────────
    "anakinra":          ("IL-1R antagonist (IL-1Ra recombinant)",
                          "Rheumatoid arthritis, SJIA, AOSD, MAS, COVID-19 CRS, FMF"),
    "astegolimab":       ("IL-33",                   "Severe uncontrolled asthma (Phase III)"),
    "camrelizumab":      ("PD-1",                    "Hodgkin lymphoma, HCC, NPC, NSCLC, Esophageal SCC (China NMPA approval)"),
    "cilgavimab":        ("SARS-CoV-2 spike protein","COVID-19 pre-exposure prophylaxis (combination with tixagevimab = Evusheld)"),
    "clazakizumab":      ("IL-6",                    "Rheumatoid arthritis (development discontinued); kidney transplant rejection (investigational)"),
    "concizumab":        ("TFPI (Tissue factor pathway inhibitor)",
                          "Hemophilia A/B with inhibitors, Non-inhibitor hemophilia A"),
    "dazukibart":        ("IFN-β",                   "Myasthenia gravis (investigational)"),
    "donanemab":         ("Amyloid-β N3pG (pyroglutamate)",
                          "Early symptomatic Alzheimer's disease"),
    "enfortumab vedotin":("Nectin-4 (ADC)",          "Locally advanced or metastatic urothelial carcinoma"),
    "enfortumab":        ("Nectin-4 (ADC)",          "Locally advanced or metastatic urothelial carcinoma"),
    "frunevetmab":       ("NGF (Nerve growth factor)","Feline osteoarthritis-associated pain (veterinary, Solensia)"),
    "itolizumab":        ("CD6",                     "Plaque psoriasis (India), Acute GVHD, COVID-19 cytokine storm"),
    "lokivetmab":        ("IL-31",                   "Canine atopic dermatitis (veterinary, Cytopoint)"),
    "reslizumab":        ("IL-5",                    "Severe eosinophilic asthma"),
    "retifanlimab":      ("PD-1",                    "Merkel cell carcinoma, Squamous cell anal cancer, Endometrial cancer"),
    "tildrakizumab":     ("IL-23p19",                "Moderate-to-severe plaque psoriasis"),
    "tisotumab vedotin": ("Tissue factor / TF (ADC)","Recurrent or metastatic cervical cancer"),
    "tisotumab":         ("Tissue factor / TF (ADC)","Recurrent or metastatic cervical cancer"),
    "tixagevimab":       ("SARS-CoV-2 spike protein","COVID-19 pre-exposure prophylaxis (combination with cilgavimab = Evusheld)"),
    "zenocutuzumab":     ("HER2/HER3 (bispecific, NRG1+ cancers)",
                          "NRG1 fusion-positive solid tumors (NSCLC, pancreatic, other)"),
    "bimagrumab":        ("ActRIIA/ActRIIB (anti-myostatin/activin)",
                          "Inclusion body myositis, Type 2 diabetes with obesity"),
    "budigalimab":       ("PD-1",                    "Head and neck squamous cell carcinoma, Cervical cancer (investigational)"),
    "ebronucimab":       ("PCSK9",                   "Hypercholesterolemia (investigational, China)"),
    "ecromeximab":       ("GD3 ganglioside",         "Metastatic melanoma (development discontinued)"),
    "elezanumab":        ("RGMa (Repulsive guidance molecule A)",
                          "Multiple sclerosis, Spinal cord injury (investigational)"),
    "enuzovimab":        ("Lassa virus glycoprotein","Lassa fever treatment (investigational)"),
    "etaracizumab":      ("αvβ3 integrin",           "Metastatic melanoma, Prostate cancer (development discontinued)"),
    "exidavnemab":       ("α-synuclein",             "Parkinson's disease (investigational)"),
    "fulranumab":        ("NGF (Nerve growth factor)","Osteoarthritis pain, Moderate-to-severe cancer pain (development discontinued)"),
    "margetuximab":      ("HER2 (Fc-optimized anti-HER2)",
                          "HER2-positive metastatic breast cancer"),
}


def get_target_disease(name: str) -> tuple[str, str]:
    key = name.lower().strip()
    if key in AB_INFO:
        return AB_INFO[key]
    # partial match
    for k, v in AB_INFO.items():
        if key.startswith(k) or k.startswith(key):
            return v
    return ("—", "—")


# ══════════════════════════════════════════════════════════════════════════════
#  LOAD DATA
# ══════════════════════════════════════════════════════════════════════════════

def load_final(fname: str) -> list[dict]:
    blob = json.loads((FTF_DIR / fname).read_text(encoding="utf-8"))
    return sorted(blob["entries"], key=lambda x: x.get("antibody_name", "").lower())


def load_clinical_db() -> dict[str, dict]:
    """Returns dict[drugname_lower -> primary_record]"""
    blob = json.loads(CLIN_DB.read_text(encoding="utf-8"))
    result: dict[str, dict] = {}
    for k, v in blob["records"].items():
        result[k.lower()] = v.get("primary_record", {})
    return result


CONFIRMED   = load_final("confirmed_ada.json")
NEED_FT     = load_final("need_fulltext.json")
CLIN_RECORDS = load_clinical_db()   # for evidence_chain


def get_clin_rec(name: str) -> dict:
    return CLIN_RECORDS.get(name.lower().strip(), {})


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def clean(v) -> str:
    if v is None: return ""
    if isinstance(v, list): return "; ".join(str(x) for x in v if x)
    return str(v).strip()


def shorten(s, n=80) -> str:
    s = clean(s)
    return (s[:n] + "…") if len(s) > n else s


def extract_pmids(entry: dict, clin: dict) -> str:
    """Get clean PMID string, checking all sources."""
    pmids: list[str] = []
    # from final file
    for p in (entry.get("pmids_extracted") or []):
        s = str(p).strip()
        if s and s not in pmids:
            pmids.append(s)
    # from verification_source  (e.g. "PMID41199323")
    vsrc = entry.get("verification_source", "")
    if vsrc:
        for m in re.findall(r"(?:PMID[:\s]*)(\d{7,10})", vsrc, re.I):
            if m not in pmids:
                pmids.append(m)
    # from evidence_chain / evidence_source in clin record
    for field in ["evidence_source", "source_url"]:
        val = clin.get(field) or ""
        for m in re.findall(r"(?:PMID[:\s]*)(\d{7,10})", val, re.I):
            if m not in pmids:
                pmids.append(m)
    # from evidence_chain text
    chain = clin.get("evidence_chain", "")
    for m in re.findall(r"PMID[:\s]*(\d{7,10})", chain, re.I):
        if m not in pmids:
            pmids.append(m)
    return "; ".join(pmids) if pmids else ""


def extract_ada_context(entry: dict, clin: dict) -> str:
    """
    Extract the ADA clinical context.
    For AI-batch Tier B: return transparent note — no false prose.
    For manually-verified Tier A: extract context from evidence_chain.
    """
    extr      = entry.get("ada_value_extraction", "")
    ev_source = entry.get("evidence_source", "")
    ai_batch  = ("ai_extracted_from_real_url_human_review_needed" in extr
                 or "" in ev_source)

    if ai_batch:
        # Clinical context prose is AI-generated — do not reproduce
        src_type = entry.get("source_type") or ""
        vstatus  = entry.get("verification_status", "")
        matched  = clean(entry.get("verification_matched_pcts", ""))
        if vstatus.startswith("verified") and matched:
            return (f"ADA incidence {matched}% confirmed present in source by automated "
                    f"text matching. Source: {src_type}. "
                    f"Clinical details (population, regimen, assay) must be read from the original URL.")
        return (f"Source: {src_type}. Clinical context (population/regimen/assay) "
                f"not auto-extracted — open citation URL for details.")

    # Manually verified
    chain = clin.get("evidence_chain", "")
    if not chain:
        return ""
    lines = chain.split("\n")
    ctx_lines = []
    for line in lines:
        line = line.strip()
        if not line or line.startswith("##") or line.startswith("---"):
            continue
        if re.search(r'\bADA\b|\bimmunog|\banti-drug\b|\banti‑drug\b', line, re.I):
            clean_line = re.sub(r'\*+', '', line).strip()
            if len(clean_line) > 30:
                ctx_lines.append(clean_line)
    result = " ".join(ctx_lines[:3])
    if len(result) > 350:
        result = result[:350] + "…"
    return result


def extract_evidence_chain_summary(entry: dict, clin: dict) -> str:
    """
    Return a compact, honest evidence chain summary.
    For Tier B AI-batch entries: label as AI-synthesized + show real URL source type.
    For Tier A manually-verified entries: show the actual key statement.
    """
    extr      = entry.get("ada_value_extraction", "")
    ev_source = entry.get("evidence_source", "")
    src_type  = entry.get("source_type") or clin.get("source_type") or ""
    vstatus   = entry.get("verification_status", "")
    tier      = entry.get("class_evidence_tier", "?")

    ai_batch = ("ai_extracted_from_real_url_human_review_needed" in extr
                or "" in ev_source)

    if ai_batch:
        # The chain text is AI-paraphrased. Be explicit.
        verified_note = ""
        if vstatus.startswith("verified"):
            matched = clean(entry.get("verification_matched_pcts", ""))
            verified_note = f"; ADA value {matched}% auto-confirmed in source text"
        return (f"[AI-summarized from verified URL (data real, text is summary not direct quote)] "
                f"Source type: {src_type}{verified_note}. "
                f"→ Open citation URL to read primary immunogenicity data.")

    # Manually verified — extract from evidence_chain
    chain = clin.get("evidence_chain", "")
    if not chain:
        return f"{src_type}" if src_type else "See citation URL"

    m = re.search(r'### Data Summary\s*\n(.*?)(?:\n\n|\Z)', chain, re.S)
    if m:
        summary = re.sub(r'\*+', '', m.group(1)).strip()
        return shorten(summary, 220)

    bold = re.findall(r'\*\*([^*]+)\*\*', chain)
    if bold:
        return shorten(" | ".join(bold[:2]), 220)
    return shorten(chain, 220)


def get_reference(entry: dict, clin: dict) -> str:
    """
    Return the best reference string:
      - PMID(s) if available (e.g. "33908636")
      - else short source description + URL domain
    """
    pmids = extract_pmids(entry, clin)
    if pmids:
        return pmids

    # No PMID → build a concise source label from source_type + URL
    src_type = (entry.get("source_type") or clin.get("source_type") or "").strip()
    urls = entry.get("citation_urls") or []
    if isinstance(urls, str):
        urls = [u.strip() for u in urls.split(";") if u.strip()]

    # Extract domain from first URL
    domain = ""
    if urls:
        m = re.match(r"https?://(?:www\.)?([^/]+)", urls[0])
        if m:
            domain = m.group(1)

    vstatus = entry.get("verification_status", "")
    # Map to clean label
    if "fda" in domain or "accessdata" in domain or "dailymed" in domain:
        return f"FDA label (DailyMed/AccessData)"
    if "pmc.ncbi" in domain:
        return f"PMC full text: {domain}"
    if "pubmed" in domain or "ncbi" in domain:
        return f"PubMed: {domain}"
    if "ema.europa" in domain:
        return f"EMA EPAR: {domain}"
    if src_type:
        return f"{src_type}" + (f" [{domain}]" if domain else "")
    if domain:
        return f"URL: {domain}"
    return "See citation URL"


def ver_label(s: str) -> str:
    MAP = {
        "verified_text_match":                     "Abstract/URL direct match",
        "verified_pmc_fulltext":                   "PMC full-text confirmed",
        "verified_fda_label_dailymed_spl":         "FDA DailyMed SPL label",
        "verified_partial_primary_value_confirmed":"Primary value confirmed",
        "verified_dailymed_spl":                   "DailyMed (brand name)",
        "verified_dailymed_spl_pmid_corrected":    "DailyMed (PMID corrected)",
        "verified_pubmed_fulltext":                "PubMed full-text search",
        "verified_pmid_corrected_match":           "Corrected PMID match",
        "verified_qualitative_no_pct_to_match":    "Qualitative: no ADA detected",
    }
    return MAP.get(s, s.replace("_", " "))


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL STYLES
# ══════════════════════════════════════════════════════════════════════════════

THIN    = Side(style="thin", color="C0C0C0")
BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP    = Alignment(wrap_text=True, vertical="top")
CENTER  = Alignment(horizontal="center", vertical="top", wrap_text=True)
HDR_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
BODY_F   = Font(name="Calibri", size=9)
BOLD_F   = Font(name="Calibri", bold=True, size=9)
LINK_F   = Font(name="Calibri", size=8, color="1155CC", underline="single")

HDR_BLUE  = PatternFill("solid", fgColor="1F4E79")
HDR_GREEN = PatternFill("solid", fgColor="1E5C37")
TIERA_F   = PatternFill("solid", fgColor="DEEAF1")
TIERB_F   = PatternFill("solid", fgColor="E2EFDA")
YELLOW_F  = PatternFill("solid", fgColor="FFF2CC")
ORANGE_F  = PatternFill("solid", fgColor="FCE4D6")
WHITE_F   = PatternFill("solid", fgColor="FFFFFF")


def _hdr_row(ws, row: int, headers: list[str], fill: PatternFill):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HDR_FONT; c.fill = fill; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[row].height = 30


def _title_row(ws, cols: int, text: str, color: str):
    ws.merge_cells(f"A1:{get_column_letter(cols)}1")
    c = ws["A1"]
    c.value     = text
    c.font      = Font(name="Calibri", bold=True, size=13, color=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", fgColor="F2F2F2")
    ws.row_dimensions[1].height = 26


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — CONFIRMED
# ══════════════════════════════════════════════════════════════════════════════

CONF_HEADERS = [
    "#", "Antibody Name",
    "Target", "Disease / Indication",
    "ADA Value",
    "Evidence Chain Note",
    "ADA Clinical Context",
    "Verification Method", "Matched %",
    "Reference (PMID / Source)", "Citation URL (Primary Source)",
    "Tier", "Chain Origin", "Annotation",
]
CONF_WIDTHS = [4, 22, 26, 40, 32, 52, 52, 26, 14, 28, 55, 5, 28, 52]


def make_excel_confirmed(entries: list[dict], out: Path):
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Confirmed ADA (80)"

    _title_row(ws, len(CONF_HEADERS),
               f"InSynBio ADA Database — Confirmed Entries (n={len(entries)})  |  {TODAY}",
               "1F4E79")
    _hdr_row(ws, 2, CONF_HEADERS, HDR_BLUE)
    for i, w in enumerate(CONF_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for ri, e in enumerate(entries, 3):
        clin   = get_clin_rec(e.get("antibody_name", ""))
        tgt, dis = get_target_disease(e.get("antibody_name", ""))
        tier   = e.get("class_evidence_tier", "?")
        fill   = TIERA_F if tier == "A" else TIERB_F

        extr      = e.get("ada_value_extraction", "")
        ev_source = e.get("evidence_source", "")
        ai_batch  = ("ai_extracted_from_real_url_human_review_needed" in extr
                     or "" in ev_source)
        chain_origin = ("AI-summarized from verified URL — ADA value confirmed in source" if ai_batch
                        else "Manually verified / human-curated")

        ref    = get_reference(e, clin)
        urls   = clean(e.get("citation_urls") or e.get("verification_source_tried") or "")
        ada_chain_sum = extract_evidence_chain_summary(e, clin)
        ada_ctx = extract_ada_context(e, clin)

        vals = [
            ri - 2,
            e.get("antibody_name", ""),
            tgt, dis,
            clean(e.get("ada_value_display")),
            ada_chain_sum,
            ada_ctx,
            ver_label(e.get("verification_status", "")),
            clean(e.get("verification_matched_pcts")),
            ref,
            shorten(urls, 160),
            tier,
            chain_origin,
            shorten(e.get("ada_value_annotation") or "", 160),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = fill
            c.alignment = WRAP
            c.border    = BORDER
            c.font      = BOLD_F if ci == 2 else BODY_F
            # Highlight chain origin column in orange for AI-batch rows
            if ci == 13 and ai_batch:
                c.fill = PatternFill("solid", fgColor="FFF2CC")
                c.font = Font(name="Calibri", size=9, italic=True, color="7F6000")
        ws.row_dimensions[ri].height = 65

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(CONF_HEADERS))}{len(entries) + 2}"

    # Sheet 2: Stats
    ws2 = wb.create_sheet("Verification Stats")
    _title_row(ws2, 3, "Verification Method Breakdown", "1F4E79")
    _hdr_row(ws2, 2, ["Method", "Count", "Example Antibodies"], HDR_BLUE)
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 8
    ws2.column_dimensions["C"].width = 60
    counts: dict[str, list] = {}
    for e in entries:
        k = ver_label(e.get("verification_status", ""))
        counts.setdefault(k, []).append(e["antibody_name"])
    for ri, (k, names) in enumerate(sorted(counts.items(), key=lambda x: -len(x[1])), 3):
        ws2.cell(row=ri, column=1, value=k).font = BODY_F
        ws2.cell(row=ri, column=2, value=len(names)).font = BOLD_F
        ws2.cell(row=ri, column=3, value=", ".join(names[:8]) + ("…" if len(names) > 8 else "")).font = BODY_F
        for ci in range(1, 4):
            ws2.cell(row=ri, column=ci).border = BORDER
            ws2.cell(row=ri, column=ci).alignment = WRAP
        ws2.row_dimensions[ri].height = 20
    ws2.freeze_panes = "A3"

    wb.save(out)
    print(f"  ✓ {out.name}")


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL — NEED FULLTEXT
# ══════════════════════════════════════════════════════════════════════════════

NFT_HEADERS = [
    "#", "Antibody Name",
    "Target", "Disease / Indication",
    "Claimed ADA Value",
    "Evidence Chain Note",
    "Tier", "Chain Origin", "Category",
    "Why Manual?", "Action Required",
    "Reference (PMID / Source)", "Manual Check URL(s)",
]
NFT_WIDTHS = [4, 22, 26, 40, 32, 52, 5, 28, 16, 55, 50, 28, 70]


def make_excel_needft(entries: list[dict], out: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Need Full-Text (36)"

    _title_row(ws, len(NFT_HEADERS),
               f"InSynBio ADA — Need Manual Verification (n={len(entries)})  |  {TODAY}",
               "1E5C37")
    _hdr_row(ws, 2, NFT_HEADERS, HDR_GREEN)
    for i, w in enumerate(NFT_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for ri, e in enumerate(entries, 3):
        clin   = get_clin_rec(e.get("antibody_name", ""))
        tgt, dis = get_target_disease(e.get("antibody_name", ""))
        reason = e.get("manual_check_reason", "")
        cat    = "A — 403 Paywall" if "403" in reason else "B — EMA/Trial/Abandoned"
        fill   = YELLOW_F if "403" in reason else ORANGE_F

        extr      = e.get("ada_value_extraction", "")
        ev_source = e.get("evidence_source", "")
        ai_batch  = ("ai_extracted_from_real_url_human_review_needed" in extr
                     or "" in ev_source)
        chain_origin = ("AI-summarized from URL — ADA value awaiting manual confirmation" if ai_batch
                        else "Manually verified / human-curated")

        ref    = get_reference(e, clin)
        mc_urls = clean(e.get("manual_check_urls") or e.get("citation_urls") or "")
        ada_chain_sum = extract_evidence_chain_summary(e, clin)

        vals = [
            ri - 2,
            e.get("antibody_name", ""),
            tgt, dis,
            clean(e.get("ada_value_display")),
            ada_chain_sum,
            e.get("class_evidence_tier", "?"),
            chain_origin,
            cat,
            shorten(reason, 120),
            shorten(e.get("suggested_action") or "Open URL → search Section 6.2 Immunogenicity", 100),
            ref,
            shorten(mc_urls, 200),
        ]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = fill
            c.alignment = WRAP
            c.border    = BORDER
            c.font      = BOLD_F if ci == 2 else BODY_F
        ws.row_dimensions[ri].height = 70

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(NFT_HEADERS))}{len(entries) + 2}"
    wb.save(out)
    print(f"  ✓ {out.name}")


# ══════════════════════════════════════════════════════════════════════════════
#  MARKDOWN
# ══════════════════════════════════════════════════════════════════════════════

def make_md_confirmed(entries: list[dict], out: Path):
    tier_b_ai = sum(1 for e in entries
                    if ("ai_extracted_from_real_url" in (e.get("ada_value_extraction") or "")
                        or "" in (e.get("evidence_source") or "")))
    tier_a_n  = sum(1 for e in entries if e.get("class_evidence_tier") == "A")
    tier_b_n  = sum(1 for e in entries if e.get("class_evidence_tier") == "B")

    L = [
        "# InSynBio ADA Database — Confirmed Entries",
        "",
        f"> **Generated**: {TODAY}  |  **n = {len(entries)}**  |  Source: InSynBio AbEngineCore ADA Verification Pipeline v3.0",
        "",
        "## Evidence Quality Classification",
        "",
        "| Class | n | Description |",
        "|-------|---|-------------|",
        f"| **Tier A** | {tier_a_n} | PMID / FDA official label / CT.gov anchor. Evidence chain text from primary source. |",
        f"| **Tier B — manually verified** | {tier_b_n - tier_b_ai} | Real URL; ADA value and chain text manually verified by human. |",
        f"| **Tier B — AI-synthesized** | {tier_b_ai} | Real URL; ADA value confirmed by automated text matching. **Evidence chain prose is AI-generated from URL content — not a verbatim quotation.** To access primary data, open the Citation URL. |",
        "",
        "> **Note on AI-summarized evidence narratives (Tier B, {tier_b_ai} entries):**",
        "> The 'Evidence Chain Note' column for these entries contains AI-generated prose summarizing",
        "> the URL source content. It is NOT a verbatim excerpt — the data is real, but the text is a paraphrase.",
        "> The ADA numeric value has been independently confirmed to appear in the source page by",
        "> automated text matching (`verified_text_match` / `verified_partial_primary_value_confirmed`).",
        "> The Citation URL is the authoritative reference — open it to read the original immunogenicity data.",
        "",
        "---",
        "",
        "## Summary Table",
        "",
        "| # | Antibody | Target | Disease / Indication | ADA Value | Reference | Tier | Verification |",
        "|---|----------|--------|----------------------|-----------|---------|------|--------------|",
    ]
    for i, e in enumerate(entries, 1):
        clin = get_clin_rec(e["antibody_name"])
        tgt, dis = get_target_disease(e["antibody_name"])
        ref   = get_reference(e, clin)
        ada   = shorten(clean(e.get("ada_value_display")), 50)
        vstatus = ver_label(e.get("verification_status", ""))
        tier  = e.get("class_evidence_tier", "?")
        L.append(f"| {i} | **{e['antibody_name']}** | {tgt} | {shorten(dis,45)} | {ada} | {ref} | {tier} | {vstatus} |")

    L += ["", "---", "", "## Detailed Entries", ""]

    for i, e in enumerate(entries, 1):
        clin    = get_clin_rec(e["antibody_name"])
        tgt, dis = get_target_disease(e["antibody_name"])
        ref     = get_reference(e, clin)
        ada     = clean(e.get("ada_value_display"))
        chain   = extract_evidence_chain_summary(e, clin)
        ctx     = extract_ada_context(e, clin)
        annot   = e.get("ada_value_annotation") or ""
        tier    = e.get("class_evidence_tier", "?")
        vstatus = ver_label(e.get("verification_status", ""))
        matched = clean(e.get("verification_matched_pcts"))
        urls    = clean(e.get("citation_urls") or "")
        extr    = e.get("ada_value_extraction", "")
        ev_src  = e.get("evidence_source", "")
        ai_batch = ("ai_extracted_from_real_url_human_review_needed" in extr
                    or "" in ev_src)
        chain_origin = ("⚠ AI-synthesized from URL source (not direct quote)" if ai_batch
                        else "✓ Manually verified / human-curated")

        L += [
            f"### {i}. {e['antibody_name']}",
            "",
            f"| Field | Value |",
            f"|-------|-------|",
            f"| **Target** | {tgt} |",
            f"| **Disease / Indication** | {dis} |",
            f"| **ADA Value** | {ada} |",
            f"| **Evidence Chain Note** | {shorten(chain, 200)} |",
            f"| **Chain Origin** | {chain_origin} |",
            f"| **ADA Clinical Context** | {shorten(ctx, 240)} |",
            f"| **Verification** | {vstatus} |",
            f"| **Matched %** | {matched} |",
            f"| **Tier** | {tier} |",
        ]
        if ref:
            L.append(f"| **Reference** | {ref} |")
        if urls:
            L.append(f"| **Citation URL (Primary Source)** | {shorten(urls,180)} |")
        if annot:
            L.append(f"| **⚠ Annotation** | {annot} |")
        L.append("")

    out.write_text("\n".join(L), encoding="utf-8")
    print(f"  ✓ {out.name}")


def make_md_needft(entries: list[dict], out: Path):
    cat_a = sum(1 for e in entries if "403" in (e.get("manual_check_reason") or ""))
    cat_b = len(entries) - cat_a

    L = [
        "# InSynBio ADA Database — Need Manual Verification",
        "",
        f"> **Generated**: {TODAY}  |  **n = {len(entries)}**",
        "",
        "## Overview",
        "",
        "These antibodies have plausible ADA values that could not be auto-verified.",
        "A reviewer with institutional access can confirm them via the provided URLs.",
        "",
        "| Category | n | Description |",
        "|----------|---|-------------|",
        f"| **A — HTTP 403** | {cat_a} | URL blocked by paywall — needs institutional subscription |",
        f"| **B — EMA/Trial/Abandoned** | {cat_b} | EMA EPAR PDF / clinical trial record / withdrawn drug |",
        "",
        "---",
        "",
        "## Summary Table",
        "",
        "| # | Antibody | Target | Disease | ADA Value | Reference | Cat | Check URL |",
        "|---|----------|--------|---------|-----------|------|-----|-----------|",
    ]
    for i, e in enumerate(entries, 1):
        clin = get_clin_rec(e["antibody_name"])
        tgt, dis = get_target_disease(e["antibody_name"])
        ref   = get_reference(e, clin)
        ada   = shorten(clean(e.get("ada_value_display")), 45)
        reason = e.get("manual_check_reason", "")
        cat   = "A" if "403" in reason else "B"
        mc    = (e.get("manual_check_urls") or e.get("citation_urls") or [])
        url1  = (mc[0] if isinstance(mc, list) else mc.split(";")[0].strip()) if mc else ""
        L.append(f"| {i} | **{e['antibody_name']}** | {tgt} | {shorten(dis,30)} | {ada} | {ref} | {cat} | {url1[:70]} |")

    L += ["", "---", "", "## Detailed Entries", ""]

    for i, e in enumerate(entries, 1):
        clin    = get_clin_rec(e["antibody_name"])
        tgt, dis = get_target_disease(e["antibody_name"])
        ref     = get_reference(e, clin)
        ada     = clean(e.get("ada_value_display"))
        chain   = extract_evidence_chain_summary(e, clin)
        reason  = e.get("manual_check_reason", "")
        cat     = "A — HTTP 403 (institutional access)" if "403" in reason else "B — EMA/Clinical/Abandoned"
        action  = e.get("suggested_action") or "Open URL → Section 6.2 Immunogenicity"
        mc_urls = e.get("manual_check_urls") or e.get("citation_urls") or []
        if isinstance(mc_urls, str):
            mc_urls = [u.strip() for u in mc_urls.split(";") if u.strip()]
        extr   = e.get("ada_value_extraction", "")
        ev_src = e.get("evidence_source", "")
        ai_batch = ("ai_extracted_from_real_url_human_review_needed" in extr
                    or "" in ev_src)
        chain_origin = ("⚠ AI-summarized from URL — ADA value awaiting manual confirmation" if ai_batch
                        else "✓ Manually verified / human-curated")

        L += [
            f"### {i}. {e['antibody_name']}",
            "",
            f"| Field | Value |",
            f"|-------|-------|",
            f"| **Target** | {tgt} |",
            f"| **Disease / Indication** | {dis} |",
            f"| **Claimed ADA Value** | {ada} |",
            f"| **Evidence Chain Note** | {shorten(chain,200)} |",
            f"| **Chain Origin** | {chain_origin} |",
            f"| **Tier** | {e.get('class_evidence_tier','?')} |",
            f"| **Category** | {cat} |",
            f"| **Reason** | {reason} |",
            f"| **Action** | {action} |",
        ]
        if ref:
            L.append(f"| **Reference** | {ref} |")
        for u in mc_urls[:3]:
            L.append(f"| **URL** | <{u}> |")
        L.append("")

    out.write_text("\n".join(L), encoding="utf-8")
    print(f"  ✓ {out.name}")


# ══════════════════════════════════════════════════════════════════════════════
#  PDF (reportlab)
# ══════════════════════════════════════════════════════════════════════════════

def _styles():
    f = CN_FONT
    return {
        "title":  ParagraphStyle("title",  fontName=f, fontSize=17, spaceAfter=6,
                                  textColor=colors.HexColor("#1F4E79"), alignment=TA_CENTER, leading=22),
        "sub":    ParagraphStyle("sub",    fontName=f, fontSize=9.5, spaceAfter=4,
                                  textColor=colors.HexColor("#404040"), alignment=TA_CENTER),
        "h1":     ParagraphStyle("h1",     fontName=f, fontSize=12, spaceBefore=12, spaceAfter=5,
                                  textColor=colors.HexColor("#1F4E79"), leading=15),
        "body":   ParagraphStyle("body",   fontName=f, fontSize=8, spaceAfter=2, leading=11),
        "note":   ParagraphStyle("note",   fontName=f, fontSize=7, textColor=colors.HexColor("#666"), leading=9),
        "cell":   ParagraphStyle("cell",   fontName=f, fontSize=7.5, leading=10),
        "bold":   ParagraphStyle("bold",   fontName=f, fontSize=7.5, leading=10),
    }


def _pg(canvas, doc):
    canvas.saveState()
    canvas.setFont(CN_FONT, 7.5)
    canvas.setFillColor(colors.HexColor("#888"))
    canvas.drawRightString(A4[0] - 1.2*cm, 0.8*cm, f"Page {doc.page}")
    canvas.drawString(1.2*cm, 0.8*cm, f"InSynBio ADA Database · {TODAY}")
    canvas.restoreState()


def _ts(hdr_color):
    return TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0), hdr_color),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), CN_FONT),
        ("FONTSIZE",     (0, 0), (-1, 0), 7.5),
        ("FONTNAME",     (0, 1), (-1, -1), CN_FONT),
        ("FONTSIZE",     (0, 1), (-1, -1), 7),
        ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#EBF3FB"), colors.white]),
        ("GRID",         (0, 0), (-1, -1), 0.25, colors.HexColor("#BBBBBB")),
        ("TOPPADDING",   (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 3),
        ("LEFTPADDING",  (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ])


def _build_doc(out_path: Path, page_fn):
    doc = BaseDocTemplate(str(out_path), pagesize=A4,
                          leftMargin=1.2*cm, rightMargin=1.2*cm,
                          topMargin=2*cm, bottomMargin=2*cm)
    frame = Frame(doc.leftMargin, doc.bottomMargin,
                  doc.width, doc.height, id="normal")
    doc.addPageTemplates([PageTemplate(id="main", frames=frame, onPage=page_fn)])
    return doc


def make_pdf_confirmed(entries: list[dict], out: Path):
    S = _styles()
    hdr_color = colors.HexColor("#1F4E79")
    doc = _build_doc(out, _pg)
    story = []

    # ── Cover ──
    story += [
        Spacer(1, 1.5*cm),
        Paragraph("InSynBio ADA Database", S["title"]),
        Paragraph("Confirmed ADA Entries", ParagraphStyle("t2", fontName=CN_FONT, fontSize=13,
            textColor=colors.HexColor("#2E74B5"), alignment=TA_CENTER, spaceAfter=5)),
        Paragraph(f"n = {len(entries)} antibodies  ·  {TODAY}", S["sub"]),
        HRFlowable(width="100%", thickness=1, color=hdr_color, spaceAfter=14),
    ]

    # ── Verification breakdown ──
    story.append(Paragraph("Verification Method Breakdown", S["h1"]))
    counts: dict[str, int] = {}
    for e in entries:
        counts[ver_label(e.get("verification_status",""))] = \
            counts.get(ver_label(e.get("verification_status","")), 0) + 1
    bk = [["Verification Method", "n"]] + \
         [[k, str(v)] for k, v in sorted(counts.items(), key=lambda x: -x[1])]
    bt = Table(bk, colWidths=[13*cm, 2*cm])
    bt.setStyle(_ts(hdr_color))
    story += [bt, Spacer(1, 0.3*cm), PageBreak()]

    # ── Main table ──
    story.append(Paragraph("Full Antibody Table", S["h1"]))
    cw = [0.6*cm, 2.8*cm, 2.6*cm, 3.5*cm, 4.2*cm, 2.5*cm, 1.4*cm]
    hdr = ["#", "Antibody", "Target", "Disease", "ADA Value", "PMID(s)", "Tier / Method"]
    rows = [hdr]
    for i, e in enumerate(entries, 1):
        clin    = get_clin_rec(e["antibody_name"])
        tgt, dis = get_target_disease(e["antibody_name"])
        ref     = get_reference(e, clin)
        ada     = clean(e.get("ada_value_display"))
        annot   = e.get("ada_value_annotation") or ""
        if annot:
            ada += f"\n[{shorten(annot,55)}]"
        vstatus = ver_label(e.get("verification_status", ""))
        tier    = e.get("class_evidence_tier", "?")
        rows.append([str(i), e["antibody_name"], shorten(tgt,28), shorten(dis,38),
                     shorten(ada,60), ref, f"{tier} / {vstatus[:16]}"])
    t = Table(rows, colWidths=cw, repeatRows=1)
    ts = _ts(hdr_color)
    for ri, e in enumerate(entries, 1):
        bg = colors.HexColor("#DEEAF1") if e.get("class_evidence_tier") == "A" \
             else colors.HexColor("#E2EFDA")
        ts.add("BACKGROUND", (0, ri), (-1, ri), bg)
    t.setStyle(ts)
    story.append(t)
    story += [
        Spacer(1, 0.4*cm),
        Paragraph("Tier A = PMID / FDA label / CT.gov anchor · Tier B = real URL anchor  |  "
                  "Values in [brackets] have contextual annotations requiring interpretation.", S["note"]),
    ]

    # ── Detailed per-antibody section ──
    story += [PageBreak(), Paragraph("Detailed Evidence per Antibody", S["h1"])]
    for i, e in enumerate(entries, 1):
        clin    = get_clin_rec(e["antibody_name"])
        tgt, dis = get_target_disease(e["antibody_name"])
        ref     = get_reference(e, clin)
        ada     = clean(e.get("ada_value_display"))
        chain   = extract_evidence_chain_summary(e, clin)
        ctx     = shorten(extract_ada_context(e, clin), 280)
        annot   = e.get("ada_value_annotation") or ""
        tier    = e.get("class_evidence_tier", "?")
        vstatus = ver_label(e.get("verification_status", ""))
        urls    = shorten(clean(e.get("citation_urls") or ""), 110)
        matched = clean(e.get("verification_matched_pcts"))

        rows2 = [
            ["Target",             tgt],
            ["Disease",            dis],
            ["ADA Value",          ada],
            ["ADA Evidence Chain", shorten(chain, 200)],
            ["ADA Clinical Context", ctx],
            ["Verification",       vstatus],
            ["Matched %",          matched],
            ["Tier",               tier],
        ]
        if ref:
            rows2.append(["PMID(s)", ref])
        if urls:
            rows2.append(["URL", urls])
        if annot:
            rows2.append(["⚠ Annotation", shorten(annot, 200)])

        story.append(Paragraph(f"{i}. {e['antibody_name']}", ParagraphStyle(
            "entry_title", fontName=CN_FONT, fontSize=9.5, spaceBefore=8, spaceAfter=2,
            textColor=hdr_color)))
        det = Table(rows2, colWidths=[3.5*cm, 14*cm])
        dts = TableStyle([
            ("FONTNAME",     (0, 0), (-1, -1), CN_FONT),
            ("FONTSIZE",     (0, 0), (-1, -1), 7.5),
            ("FONTNAME",     (0, 0), (0, -1),  CN_FONT),
            ("TEXTCOLOR",    (0, 0), (0, -1),  colors.HexColor("#1F4E79")),
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
            ("GRID",         (0, 0), (-1, -1), 0.2, colors.HexColor("#CCCCCC")),
            ("BACKGROUND",   (0, 0), (0, -1),  colors.HexColor("#EBF3FB")),
            ("TOPPADDING",   (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ])
        det.setStyle(dts)
        story.append(det)

    doc.build(story)
    print(f"  ✓ {out.name}")


def make_pdf_needft(entries: list[dict], out: Path):
    S = _styles()
    hdr_color = colors.HexColor("#1E5C37")
    doc = _build_doc(out, _pg)
    story = []

    cat_a = sum(1 for e in entries if "403" in (e.get("manual_check_reason") or ""))
    cat_b = len(entries) - cat_a

    story += [
        Spacer(1, 1.5*cm),
        Paragraph("InSynBio ADA Database", S["title"]),
        Paragraph("Need Manual Verification", ParagraphStyle("t2", fontName=CN_FONT, fontSize=13,
            textColor=colors.HexColor("#1E5C37"), alignment=TA_CENTER, spaceAfter=5)),
        Paragraph(f"n = {len(entries)} antibodies  ·  {TODAY}", S["sub"]),
        HRFlowable(width="100%", thickness=1, color=hdr_color, spaceAfter=14),
    ]

    cat_data = [
        ["Category", "n", "Description"],
        ["A — HTTP 403", str(cat_a), "URL blocked by paywall (NEJM / ScienceDirect). Open with institutional subscription."],
        ["B — EMA/Trial/Abandoned", str(cat_b), "EMA EPAR PDF / clinical trial record / withdrawn drug. Open document and search immunogenicity."],
    ]
    ct = Table(cat_data, colWidths=[4*cm, 1.3*cm, 12.5*cm])
    ct.setStyle(_ts(hdr_color))
    story += [ct, Spacer(1, 0.4*cm), PageBreak()]

    story.append(Paragraph("Full List with Check URLs", S["h1"]))
    cw = [0.6*cm, 2.8*cm, 2.6*cm, 3.5*cm, 4*cm, 1.4*cm, 2.9*cm]
    hdr = ["#", "Antibody", "Target", "Disease", "Claimed ADA", "PMID(s)", "Cat / Tier"]
    rows = [hdr]
    for i, e in enumerate(entries, 1):
        clin    = get_clin_rec(e["antibody_name"])
        tgt, dis = get_target_disease(e["antibody_name"])
        ref     = get_reference(e, clin)
        reason  = e.get("manual_check_reason", "")
        cat     = "A" if "403" in reason else "B"
        rows.append([
            str(i), e["antibody_name"], shorten(tgt,28), shorten(dis,38),
            shorten(clean(e.get("ada_value_display")),55),
            ref, f"{cat} / {e.get('class_evidence_tier','?')}",
        ])
    t = Table(rows, colWidths=cw, repeatRows=1)
    ts2 = _ts(hdr_color)
    for ri, e in enumerate(entries, 1):
        reason = e.get("manual_check_reason", "")
        bg = colors.HexColor("#FFF2CC") if "403" in reason else colors.HexColor("#FCE4D6")
        ts2.add("BACKGROUND", (0, ri), (-1, ri), bg)
    t.setStyle(ts2)
    story.append(t)
    story += [Spacer(1, 0.4*cm), PageBreak()]

    story.append(Paragraph("Detailed Entries with Check URLs", S["h1"]))
    for i, e in enumerate(entries, 1):
        clin    = get_clin_rec(e["antibody_name"])
        tgt, dis = get_target_disease(e["antibody_name"])
        ref     = get_reference(e, clin)
        ada     = clean(e.get("ada_value_display"))
        chain   = extract_evidence_chain_summary(e, clin)
        reason  = e.get("manual_check_reason", "")
        cat     = "A — HTTP 403" if "403" in reason else "B — EMA/Clinical/Abandoned"
        action  = e.get("suggested_action") or "Open URL → Section 6.2 Immunogenicity"
        mc_urls = e.get("manual_check_urls") or e.get("citation_urls") or []
        if isinstance(mc_urls, str):
            mc_urls = [u.strip() for u in mc_urls.split(";") if u.strip()]

        rows2 = [
            ["Target",             tgt],
            ["Disease",            dis],
            ["Claimed ADA",        ada],
            ["ADA Evidence Chain", shorten(chain, 200)],
            ["Tier",               e.get("class_evidence_tier", "?")],
            ["Category",           cat],
            ["Action",             shorten(action, 150)],
        ]
        if ref:
            rows2.append(["PMID(s)", ref])
        for u in mc_urls[:2]:
            rows2.append(["Check URL", shorten(u, 110)])

        story.append(Paragraph(f"{i}. {e['antibody_name']}", ParagraphStyle(
            "et2", fontName=CN_FONT, fontSize=9.5, spaceBefore=8, spaceAfter=2,
            textColor=hdr_color)))
        det = Table(rows2, colWidths=[3.5*cm, 14*cm])
        dts = TableStyle([
            ("FONTNAME",     (0, 0), (-1, -1), CN_FONT),
            ("FONTSIZE",     (0, 0), (-1, -1), 7.5),
            ("TEXTCOLOR",    (0, 0), (0, -1),  colors.HexColor("#1E5C37")),
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
            ("GRID",         (0, 0), (-1, -1), 0.2, colors.HexColor("#CCCCCC")),
            ("BACKGROUND",   (0, 0), (0, -1),  colors.HexColor("#E2EFDA")),
            ("TOPPADDING",   (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING",(0, 0), (-1, -1), 2),
            ("LEFTPADDING",  (0, 0), (-1, -1), 4),
        ])
        det.setStyle(dts)
        story.append(det)

    doc.build(story)
    print(f"  ✓ {out.name}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print("=== Generating Excel files ===")
    make_excel_confirmed(CONFIRMED, OUT_DIR / "confirmed_ada.xlsx")
    make_excel_needft(NEED_FT,     OUT_DIR / "need_fulltext.xlsx")

    print("=== Generating Markdown files ===")
    make_md_confirmed(CONFIRMED, OUT_DIR / "confirmed_ada.md")
    make_md_needft(NEED_FT,     OUT_DIR / "need_fulltext.md")

    print("=== Generating PDF files ===")
    make_pdf_confirmed(CONFIRMED, OUT_DIR / "confirmed_ada.pdf")
    make_pdf_needft(NEED_FT,     OUT_DIR / "need_fulltext.pdf")

    print(f"\nAll 6 files → {OUT_DIR}")
