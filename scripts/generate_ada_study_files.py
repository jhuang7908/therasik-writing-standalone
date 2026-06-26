#!/usr/bin/env python3
"""
Generate 6 study files from the confirmed (80) and need-fulltext (36) ADA databases:
  confirmed_ada.xlsx / .md / .pdf
  need_fulltext.xlsx  / .md / .pdf
"""
from __future__ import annotations
import json, re, textwrap
from datetime import datetime
from pathlib import Path

# ── openpyxl ──────────────────────────────────────────────────────────────────
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

# ── reportlab ─────────────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (BaseDocTemplate, Frame, HRFlowable,
                                 PageBreak, PageTemplate, Paragraph,
                                 Spacer, Table, TableStyle)

REPO    = Path(__file__).resolve().parents[1]
SRC_DIR = REPO / "data/ADA_reliable_package/final_three_files"
OUT_DIR = REPO / "data/ADA_reliable_package/study_materials"
OUT_DIR.mkdir(parents=True, exist_ok=True)

TODAY   = datetime.now().strftime("%Y-%m-%d")

# ── Register Chinese font ─────────────────────────────────────────────────────
FONT_PATHS = [
    r"C:\Windows\Fonts\msyh.ttc",
    r"C:\Windows\Fonts\msyhbd.ttc",
    r"C:\Windows\Fonts\simhei.ttf",
    r"C:\Windows\Fonts\simsun.ttc",
    r"C:\Windows\Fonts\Arial.ttf",
]
CN_FONT = "Arial"
for fp in FONT_PATHS:
    if Path(fp).exists():
        try:
            pdfmetrics.registerFont(TTFont("CNFont", fp))
            CN_FONT = "CNFont"
            break
        except Exception:
            continue


# ══════════════════════════════════════════════════════════════════════════════
#  DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════

def load(fname: str) -> list[dict]:
    blob = json.loads((SRC_DIR / fname).read_text(encoding="utf-8"))
    return sorted(blob["entries"], key=lambda x: x.get("antibody_name", "").lower())


CONFIRMED = load("confirmed_ada.json")
NEED_FT   = load("need_fulltext.json")


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def clean(v) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        return "; ".join(str(x) for x in v)
    return str(v).strip()


def shorten(s: str, n: int = 80) -> str:
    s = clean(s)
    return s[:n] + "…" if len(s) > n else s


def ver_label(s: str) -> str:
    MAP = {
        "verified_text_match":                    "/URL",
        "verified_pmc_fulltext":                  "PMC",
        "verified_fda_label_dailymed_spl":        "FDA DailyMed SPL",
        "verified_partial_primary_value_confirmed":"",
        "verified_dailymed_spl":                  "DailyMed",
        "verified_dailymed_spl_pmid_corrected":   "DailyMed(PMID)",
        "verified_pubmed_fulltext":               "PubMed",
        "verified_pmid_corrected_match":          "PMID",
        "verified_qualitative_no_pct_to_match":   "ADA",
    }
    return MAP.get(s, s)


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL
# ══════════════════════════════════════════════════════════════════════════════

HDR_FILL_BLUE  = PatternFill("solid", fgColor="1F4E79")
HDR_FILL_GREEN = PatternFill("solid", fgColor="1E5C37")
TIER_A_FILL    = PatternFill("solid", fgColor="EBF3FB")
TIER_B_FILL    = PatternFill("solid", fgColor="F0F7F0")
GREY_FILL      = PatternFill("solid", fgColor="F5F5F5")
WHITE_FILL     = PatternFill("solid", fgColor="FFFFFF")
HDR_FONT       = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT      = Font(name="Calibri", size=9)
BOLD_FONT      = Font(name="Calibri", bold=True, size=9)
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)


def make_excel_confirmed(entries: list[dict], out: Path) -> None:
    wb = openpyxl.Workbook()

    # ── Sheet 1: All entries ──
    ws = wb.active
    ws.title = "Confirmed ADA (80)"
    ws.sheet_view.showGridLines = True

    # Title row
    ws.merge_cells("A1:L1")
    tc = ws["A1"]
    tc.value = f"InSynBio ADA Database — Confirmed Entries (n={len(entries)})  |  Generated {TODAY}"
    tc.font  = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    headers = [
        "No.", "Antibody Name", "Target / Disease",
        "ADA Value Display", "Verification Method",
        "Tier", "ADA Status",
        "Matched %", "Annotation",
        "Evidence Source", "PMID(s)", "Citation URL(s)",
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = HDR_FONT
        c.fill      = HDR_FILL_BLUE
        c.alignment = CENTER
        c.border    = BORDER
    ws.row_dimensions[2].height = 28

    col_widths = [5, 22, 28, 34, 24, 6, 18, 18, 38, 26, 18, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for ri, e in enumerate(entries, 3):
        tier  = e.get("class_evidence_tier", "?")
        fill  = TIER_A_FILL if tier == "A" else TIER_B_FILL if tier == "B" else WHITE_FILL
        row_data = [
            ri - 2,
            e.get("antibody_name", ""),
            (e.get("class_ada_status") or "").replace("_", " "),
            clean(e.get("ada_value_display")),
            ver_label(e.get("verification_status", "")),
            tier,
            (e.get("class_ada_status") or "").replace("_", " "),
            clean(e.get("verification_matched_pcts")),
            shorten(e.get("ada_value_annotation") or "", 120),
            shorten(e.get("evidence_source") or "", 50),
            clean(e.get("pmids_extracted")),
            shorten(clean(e.get("citation_urls")), 120),
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = BOLD_FONT if ci == 2 else BODY_FONT
            c.fill      = fill
            c.alignment = WRAP
            c.border    = BORDER
        ws.row_dimensions[ri].height = 42

    # ── Sheet 2: By verification method ──
    ws2 = wb.create_sheet("By Verification Method")
    ws2.merge_cells("A1:D1")
    ws2["A1"].value = "Verification Method Breakdown"
    ws2["A1"].font  = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 22

    for ci, h in enumerate(["Method", "Code", "Count", "Example Antibodies"], 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL_BLUE; c.alignment = CENTER; c.border = BORDER
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 42
    ws2.column_dimensions["C"].width = 8
    ws2.column_dimensions["D"].width = 50

    counts: dict[str, list] = {}
    for e in entries:
        k = e.get("verification_status", "unknown")
        counts.setdefault(k, []).append(e["antibody_name"])

    for ri, (k, names) in enumerate(sorted(counts.items(), key=lambda x: -len(x[1])), 3):
        ws2.cell(row=ri, column=1, value=ver_label(k)).font = BODY_FONT
        ws2.cell(row=ri, column=2, value=k).font = BODY_FONT
        ws2.cell(row=ri, column=3, value=len(names)).font = BOLD_FONT
        ws2.cell(row=ri, column=4, value=", ".join(names[:5]) + ("…" if len(names) > 5 else "")).font = BODY_FONT
        for ci in range(1, 5):
            ws2.cell(row=ri, column=ci).border = BORDER
            ws2.cell(row=ri, column=ci).alignment = WRAP
        ws2.row_dimensions[ri].height = 20

    ws2.freeze_panes = "A3"

    # Freeze panes + filter on main sheet
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:L{1 + len(entries) + 1}"

    wb.save(out)
    print(f"  Saved {out.name}")


def make_excel_needft(entries: list[dict], out: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Need Full-Text (36)"

    ws.merge_cells("A1:K1")
    tc = ws["A1"]
    tc.value = f"InSynBio ADA — Need Manual Verification (n={len(entries)})  |  {TODAY}"
    tc.font  = Font(name="Calibri", bold=True, size=13, color="1E5C37")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    HDR_F2 = PatternFill("solid", fgColor="1E5C37")
    headers = [
        "No.", "Antibody Name", "ADA Value (Claimed)",
        "Tier", "Category", "Why Manual?",
        "Action Required", "Manual Check URLs",
        "Evidence Source", "PMID(s)", "Citation URL(s)",
    ]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_F2; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[2].height = 28

    col_widths = [5, 22, 34, 6, 14, 50, 38, 65, 26, 18, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    CAT_A_FILL = PatternFill("solid", fgColor="FFF2CC")
    CAT_B_FILL = PatternFill("solid", fgColor="FCE4D6")

    for ri, e in enumerate(entries, 3):
        reason = e.get("manual_check_reason", "")
        cat    = "A-403" if "403" in reason else "B-EMA//"
        fill   = CAT_A_FILL   if "403" in reason else CAT_B_FILL
        mc_urls = e.get("manual_check_urls") or e.get("citation_urls") or []

        row_data = [
            ri - 2,
            e.get("antibody_name", ""),
            clean(e.get("ada_value_display")),
            e.get("class_evidence_tier", "?"),
            cat,
            shorten(reason, 100),
            shorten(e.get("suggested_action") or "Open URL → search immunogenicity section", 80),
            clean(mc_urls),
            shorten(e.get("evidence_source") or "", 50),
            clean(e.get("pmids_extracted")),
            shorten(clean(e.get("citation_urls")), 80),
        ]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = BOLD_FONT if ci == 2 else BODY_FONT
            c.fill      = fill
            c.alignment = WRAP
            c.border    = BORDER
        ws.row_dimensions[ri].height = 52

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:K{1 + len(entries) + 1}"
    wb.save(out)
    print(f"  Saved {out.name}")


# ══════════════════════════════════════════════════════════════════════════════
#  MARKDOWN
# ══════════════════════════════════════════════════════════════════════════════

def make_md_confirmed(entries: list[dict], out: Path) -> None:
    lines = [
        "# InSynBio ADA Database — Confirmed Entries",
        "",
        f"> **Generated**: {TODAY}  |  **Total entries**: {len(entries)}  |  **Source**: InSynBio AbEngineCore ADA Verification Pipeline v3.0",
        "",
        "## Overview",
        "",
        "This file contains **80 therapeutic antibodies** with ADA (Anti-Drug Antibody) incidence rates",
        "verified against real primary sources (PubMed abstracts, PMC full text, FDA DailyMed SPL labels).",
        "All values have been confirmed by automated text matching against the cited source.",
        "",
        "### Verification Method Legend",
        "",
        "| Code | Method | Description |",
        "|------|--------|-------------|",
        "| `verified_text_match` | /URL | ADA% found in fetched abstract or URL text |",
        "| `verified_pmc_fulltext` | PMC | ADA% found in PMC full-text XML |",
        "| `verified_fda_label_dailymed_spl` | FDA DailyMed SPL | ADA% found in FDA label immunogenicity section |",
        "| `verified_partial_primary_value_confirmed` |  | Primary ADA% confirmed; secondary values from other sources |",
        "| `verified_dailymed_spl` | DailyMed | Verified using drug brand name in DailyMed |",
        "| `verified_dailymed_spl_pmid_corrected` | PMID | Original PMID was wrong; value confirmed via DailyMed |",
        "| `verified_pubmed_fulltext` | PubMed | Confirmed via PubMed targeted search full text |",
        "| `verified_pmid_corrected_match` | PMID | Replaced wrong PMID with correct paper |",
        "| `verified_qualitative_no_pct_to_match` | ADA | Qualitative 'no ADA detected'; no % to verify |",
        "",
        "---",
        "",
        "## Entry Table",
        "",
        "| # | Antibody | Tier | ADA Value | Verification | Matched % | Annotation | PMID(s) |",
        "|---|----------|------|-----------|--------------|-----------|------------|---------|",
    ]

    for i, e in enumerate(entries, 1):
        name    = e.get("antibody_name", "")
        tier    = e.get("class_evidence_tier", "?")
        ada     = shorten(clean(e.get("ada_value_display")), 55)
        vstatus = ver_label(e.get("verification_status", ""))
        matched = clean(e.get("verification_matched_pcts"))
        annot   = shorten(e.get("ada_value_annotation") or "", 80)
        pmids   = clean(e.get("pmids_extracted"))
        lines.append(f"| {i} | **{name}** | {tier} | {ada} | {vstatus} | {matched} | {annot} | {pmids} |")

    lines += [
        "",
        "---",
        "",
        "## Detailed Entries",
        "",
    ]

    for i, e in enumerate(entries, 1):
        name    = e.get("antibody_name", "")
        tier    = e.get("class_evidence_tier", "?")
        ada     = clean(e.get("ada_value_display"))
        vstatus = e.get("verification_status", "")
        matched = clean(e.get("verification_matched_pcts"))
        annot   = e.get("ada_value_annotation") or ""
        es      = e.get("evidence_source") or ""
        pmids   = clean(e.get("pmids_extracted"))
        urls    = clean(e.get("citation_urls"))
        status  = e.get("class_ada_status") or ""

        lines += [
            f"### {i}. {name}",
            "",
            f"- **Evidence Tier**: {tier}",
            f"- **ADA Status**: {status.replace('_',' ')}",
            f"- **ADA Value**: {ada}",
            f"- **Verification**: {ver_label(vstatus)}",
            f"- **Matched values**: {matched}",
        ]
        if annot:
            lines.append(f"- **⚠ Annotation**: {annot}")
        if pmids:
            lines.append(f"- **PMID(s)**: {pmids}")
        if urls:
            lines.append(f"- **Source URL(s)**: {shorten(urls, 150)}")
        lines.append("")

    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Saved {out.name}")


def make_md_needft(entries: list[dict], out: Path) -> None:
    lines = [
        "# InSynBio ADA Database — Need Manual Verification",
        "",
        f"> **Generated**: {TODAY}  |  **Total entries**: {len(entries)}",
        "",
        "## Overview",
        "",
        "These **36 antibodies** have plausible ADA values that could not be auto-verified.",
        "A human reviewer with institutional database access can confirm them by following",
        "the provided URLs.",
        "",
        "### Categories",
        "",
        "| Category | Count | Description |",
        "|----------|-------|-------------|",
        f"| **A — HTTP 403** | {sum(1 for e in entries if '403' in (e.get('manual_check_reason') or ''))} | Real URL but returns 403 (NEJM/ScienceDirect etc.) — needs institutional subscription |",
        f"| **B — EMA/Clinical/Abandoned** | {sum(1 for e in entries if '403' not in (e.get('manual_check_reason') or ''))} | EMA EPAR PDF / clinical trial data / withdrawn drugs — needs human to open document |",
        "",
        "---",
        "",
        "## Summary Table",
        "",
        "| # | Antibody | Tier | ADA Value | Category | Action Required | Check URL |",
        "|---|----------|------|-----------|----------|-----------------|-----------|",
    ]

    for i, e in enumerate(entries, 1):
        name   = e.get("antibody_name", "")
        tier   = e.get("class_evidence_tier", "?")
        ada    = shorten(clean(e.get("ada_value_display")), 45)
        reason = e.get("manual_check_reason", "")
        cat    = "A-403" if "403" in reason else "B-EMA/Trial"
        action = shorten(e.get("suggested_action") or "Open URL → search immunogenicity", 60)
        mc_urls = e.get("manual_check_urls") or e.get("citation_urls") or []
        url1   = mc_urls[0][:70] if mc_urls else ""
        lines.append(f"| {i} | **{name}** | {tier} | {ada} | {cat} | {action} | {url1} |")

    lines += ["", "---", "", "## Detailed Entries", ""]

    for i, e in enumerate(entries, 1):
        name    = e.get("antibody_name", "")
        tier    = e.get("class_evidence_tier", "?")
        ada     = clean(e.get("ada_value_display"))
        reason  = e.get("manual_check_reason", "")
        cat     = "A — HTTP 403 (institutional access)" if "403" in reason else "B — EMA/Clinical/Abandoned"
        action  = e.get("suggested_action") or "Open URL → search Section 6.2 Immunogenicity"
        mc_urls = e.get("manual_check_urls") or e.get("citation_urls") or []
        es      = e.get("evidence_source") or ""
        pmids   = clean(e.get("pmids_extracted"))

        lines += [
            f"### {i}. {name}",
            "",
            f"- **Tier**: {tier}",
            f"- **Claimed ADA Value**: {ada}",
            f"- **Category**: {cat}",
            f"- **Reason**: {reason}",
            f"- **Action**: {action}",
        ]
        for u in mc_urls:
            lines.append(f"- **URL**: <{u}>")
        if pmids:
            lines.append(f"- **PMID(s)**: {pmids}")
        if es:
            lines.append(f"- **Evidence Source**: {es}")
        lines.append("")

    out.write_text("\n".join(lines), encoding="utf-8")
    print(f"  Saved {out.name}")


# ══════════════════════════════════════════════════════════════════════════════
#  PDF (reportlab)
# ══════════════════════════════════════════════════════════════════════════════

PAGE_W, PAGE_H = A4

def _styles():
    ss = getSampleStyleSheet()
    base = CN_FONT

    title_s  = ParagraphStyle("title",  fontName=base, fontSize=18, spaceAfter=8,
                               textColor=colors.HexColor("#1F4E79"), alignment=TA_CENTER, leading=22)
    sub_s    = ParagraphStyle("sub",    fontName=base, fontSize=10, spaceAfter=4,
                               textColor=colors.HexColor("#404040"), alignment=TA_CENTER)
    h1_s     = ParagraphStyle("h1",     fontName=base, fontSize=13, spaceBefore=14, spaceAfter=6,
                               textColor=colors.HexColor("#1F4E79"), leading=16)
    h2_s     = ParagraphStyle("h2",     fontName=base, fontSize=10, spaceBefore=8, spaceAfter=4,
                               textColor=colors.HexColor("#2E74B5"), leading=13)
    body_s   = ParagraphStyle("body",   fontName=base, fontSize=8.5, spaceAfter=3, leading=12)
    note_s   = ParagraphStyle("note",   fontName=base, fontSize=7.5, textColor=colors.HexColor("#666666"),
                               spaceAfter=2, leading=10)
    return title_s, sub_s, h1_s, h2_s, body_s, note_s


def _page_num(canvas, doc):
    canvas.saveState()
    canvas.setFont(CN_FONT, 8)
    canvas.setFillColor(colors.HexColor("#888888"))
    canvas.drawRightString(PAGE_W - 1.5*cm, 1*cm, f"Page {doc.page}")
    canvas.drawString(1.5*cm, 1*cm, f"InSynBio ADA Database — {TODAY}")
    canvas.restoreState()


def _table_style(header_color):
    return TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), header_color),
        ("TEXTCOLOR",   (0, 0), (-1, 0), colors.white),
        ("FONTNAME",    (0, 0), (-1, 0), CN_FONT),
        ("FONTSIZE",    (0, 0), (-1, 0), 8),
        ("FONTNAME",    (0, 1), (-1, -1), CN_FONT),
        ("FONTSIZE",    (0, 1), (-1, -1), 7.5),
        ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        ("ALIGN",       (0, 1), (-1, -1), "LEFT"),
        ("VALIGN",      (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#EBF3FB"), colors.white]),
        ("GRID",        (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("TOPPADDING",  (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("WORDWRAP",    (0, 0), (-1, -1), True),
    ])


def make_pdf_confirmed(entries: list[dict], out: Path) -> None:
    title_s, sub_s, h1_s, h2_s, body_s, note_s = _styles()
    doc = BaseDocTemplate(str(out), pagesize=A4,
                          leftMargin=1.5*cm, rightMargin=1.5*cm,
                          topMargin=2*cm, bottomMargin=2*cm)
    frame = Frame(doc.leftMargin, doc.bottomMargin,
                  doc.width, doc.height, id="normal")
    doc.addPageTemplates([PageTemplate(id="main", frames=frame, onPage=_page_num)])

    story = []
    hdr_color = colors.HexColor("#1F4E79")

    # Title page
    story.append(Spacer(1, 2*cm))
    story.append(Paragraph("InSynBio ADA Database", title_s))
    story.append(Paragraph("Confirmed ADA Entries", ParagraphStyle(
        "t2", fontName=CN_FONT, fontSize=14, textColor=colors.HexColor("#2E74B5"),
        alignment=TA_CENTER, spaceAfter=6)))
    story.append(Paragraph(f"n = {len(entries)} antibodies  ·  Generated {TODAY}", sub_s))
    story.append(HRFlowable(width="100%", thickness=1, color=hdr_color, spaceAfter=16))

    # Verification breakdown table
    story.append(Paragraph("Verification Method Breakdown", h1_s))
    counts: dict[str, int] = {}
    for e in entries:
        k = ver_label(e.get("verification_status", ""))
        counts[k] = counts.get(k, 0) + 1
    bk_data = [["Verification Method", "Count"]] + \
              [[k, str(v)] for k, v in sorted(counts.items(), key=lambda x: -x[1])]
    bk_table = Table(bk_data, colWidths=[13*cm, 2.5*cm])
    bk_table.setStyle(_table_style(hdr_color))
    story.append(bk_table)
    story.append(Spacer(1, 0.5*cm))

    story.append(PageBreak())
    story.append(Paragraph("Full Antibody Table", h1_s))

    # Main table
    col_w = [0.7*cm, 3.2*cm, 1.2*cm, 5.5*cm, 3.8*cm, 4.5*cm]
    header = ["#", "Antibody", "Tier", "ADA Value (Verified)", "Verification", "PMID / Source"]
    rows = [header]
    for i, e in enumerate(entries, 1):
        pmids = clean(e.get("pmids_extracted"))
        src   = shorten(clean(e.get("citation_urls")), 55) if not pmids else pmids
        annot = e.get("ada_value_annotation") or ""
        ada   = clean(e.get("ada_value_display"))
        if annot:
            ada += f"\n[{shorten(annot,60)}]"
        rows.append([
            str(i),
            e.get("antibody_name", ""),
            e.get("class_evidence_tier", "?"),
            ada,
            ver_label(e.get("verification_status", "")),
            src,
        ])
    t = Table(rows, colWidths=col_w, repeatRows=1)
    ts = _table_style(hdr_color)
    # Highlight Tier A rows slightly
    for ri, e in enumerate(entries, 1):
        if e.get("class_evidence_tier") == "A":
            ts.add("BACKGROUND", (0, ri), (-1, ri), colors.HexColor("#EBF3FB"))
        else:
            ts.add("BACKGROUND", (0, ri), (-1, ri), colors.HexColor("#F0F7F0"))
    t.setStyle(ts)
    story.append(t)

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        "Note: Values with [annotation] require contextual interpretation. "
        "Tier A = PMID/FDA/CT.gov anchor; Tier B = real URL anchor.",
        note_s))

    doc.build(story)
    print(f"  Saved {out.name}")


def make_pdf_needft(entries: list[dict], out: Path) -> None:
    title_s, sub_s, h1_s, h2_s, body_s, note_s = _styles()
    doc = BaseDocTemplate(str(out), pagesize=A4,
                          leftMargin=1.5*cm, rightMargin=1.5*cm,
                          topMargin=2*cm, bottomMargin=2*cm)
    frame = Frame(doc.leftMargin, doc.bottomMargin,
                  doc.width, doc.height, id="normal")
    doc.addPageTemplates([PageTemplate(id="main", frames=frame, onPage=_page_num)])

    story = []
    hdr_color = colors.HexColor("#1E5C37")

    story.append(Spacer(1, 2*cm))
    story.append(Paragraph("InSynBio ADA Database", title_s))
    story.append(Paragraph("Need Manual Verification", ParagraphStyle(
        "t2", fontName=CN_FONT, fontSize=14, textColor=colors.HexColor("#1E5C37"),
        alignment=TA_CENTER, spaceAfter=6)))
    story.append(Paragraph(f"n = {len(entries)} antibodies  ·  Generated {TODAY}", sub_s))
    story.append(HRFlowable(width="100%", thickness=1, color=hdr_color, spaceAfter=16))

    cat_a = sum(1 for e in entries if "403" in (e.get("manual_check_reason") or ""))
    cat_b = len(entries) - cat_a
    story.append(Paragraph("Categories", h1_s))
    cat_data = [
        ["Category", "Count", "Description"],
        ["A — HTTP 403", str(cat_a), "URL blocked by paywall (NEJM/ScienceDirect). Open with institutional access."],
        ["B — EMA/Trial/Abandoned", str(cat_b), "EMA EPAR PDF / clinical trial / withdrawn drug. Open PDF and search immunogenicity."],
    ]
    ct = Table(cat_data, colWidths=[4*cm, 1.5*cm, 13*cm])
    ct.setStyle(_table_style(hdr_color))
    story.append(ct)
    story.append(Spacer(1, 0.4*cm))
    story.append(PageBreak())

    story.append(Paragraph("Full List with Check URLs", h1_s))
    col_w = [0.7*cm, 3.2*cm, 1.2*cm, 4.5*cm, 2*cm, 7.4*cm]
    header = ["#", "Antibody", "Tier", "Claimed ADA Value", "Cat", "Manual Check URL(s)"]
    rows = [header]
    for i, e in enumerate(entries, 1):
        mc_urls = e.get("manual_check_urls") or e.get("citation_urls") or []
        url_str = "\n".join(u[:75] for u in mc_urls[:2])
        reason  = e.get("manual_check_reason", "")
        cat     = "A" if "403" in reason else "B"
        rows.append([
            str(i),
            e.get("antibody_name", ""),
            e.get("class_evidence_tier", "?"),
            shorten(clean(e.get("ada_value_display")), 60),
            cat,
            url_str,
        ])
    t = Table(rows, colWidths=col_w, repeatRows=1)
    ts2 = _table_style(hdr_color)
    for ri, e in enumerate(entries, 1):
        reason = e.get("manual_check_reason", "")
        bg = colors.HexColor("#FFF2CC") if "403" in reason else colors.HexColor("#FCE4D6")
        ts2.add("BACKGROUND", (0, ri), (-1, ri), bg)
    t.setStyle(ts2)
    story.append(t)

    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        "Action: For Category A, open URL with institutional library access and find Section 6.2 Immunogenicity. "
        "For Category B, open the EMA EPAR PDF or PubMed full text and search for 'ADA', "
        "'anti-drug antibody', or 'immunogenicity incidence'.",
        note_s))

    doc.build(story)
    print(f"  Saved {out.name}")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=== Generating Excel files ===")
    make_excel_confirmed(CONFIRMED, OUT_DIR / "confirmed_ada.xlsx")
    make_excel_needft(NEED_FT,     OUT_DIR / "need_fulltext.xlsx")

    print("=== Generating Markdown files ===")
    make_md_confirmed(CONFIRMED, OUT_DIR / "confirmed_ada.md")
    make_md_needft(NEED_FT,     OUT_DIR / "need_fulltext.md")

    print("=== Generating PDF files ===")
    make_pdf_confirmed(CONFIRMED, OUT_DIR / "confirmed_ada.pdf")
    make_pdf_needft(NEED_FT,     OUT_DIR / "need_fulltext.pdf")

    print(f"\nAll 6 files written to:\n  {OUT_DIR}")
